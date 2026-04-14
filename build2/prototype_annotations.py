#!/usr/bin/env python3
"""
prototype_annotations.py — Add sticky-note annotation boxes to Excel sheets.

Post-processes an .xlsx file to inject OOXML drawing elements (rectangles
with text) anchored to cell ranges. Works as a step after openpyxl saves.

Usage (standalone test):
    python3 build/prototype_annotations.py

Produces build/test_annotations.xlsx — open in Excel to verify rendering.
"""

import openpyxl
from openpyxl.styles import Font
import zipfile
import os
import re


# ── OOXML constants ──

_DRAWING_REL = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing'
_DRAWING_CT  = 'application/vnd.openxmlformats-officedocument.drawing+xml'
_XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
_A   = 'http://schemas.openxmlformats.org/drawingml/2006/main'


def _esc(text):
    return text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def _make_drawing_xml(shapes):
    """Build a complete wsDr drawing document containing rectangle shapes."""
    anchors = []
    for i, s in enumerate(shapes, start=2):
        paras = []
        for line in s.get('lines', ['']):
            # **bold** or _italic_ markup
            bold = line.startswith('**') and line.endswith('**')
            text = line[2:-2] if bold else line
            italic = text.startswith('_') and text.endswith('_')
            if italic:
                text = text[1:-1]

            # Empty line → empty paragraph
            if not text and not bold:
                paras.append('<a:p><a:endParaRPr lang="en-US" sz="800"/></a:p>')
                continue

            b = ' b="1"' if bold else ''
            it = ' i="1"' if italic else ''
            color = '808080' if italic and not bold else '333333'
            paras.append(
                f'<a:p><a:r>'
                f'<a:rPr lang="en-US" sz="800"{b}{it}>'
                f'<a:solidFill><a:srgbClr val="{color}"/></a:solidFill>'
                f'<a:latin typeface="Arial"/>'
                f'</a:rPr>'
                f'<a:t>{_esc(text)}</a:t>'
                f'</a:r></a:p>'
            )

        fill = s.get('fill', 'FFFFF0')
        border = s.get('border', 'C0C0C0')

        anchors.append(
            f'<xdr:twoCellAnchor>'
            f'<xdr:from>'
            f'<xdr:col>{s["from_col"]}</xdr:col><xdr:colOff>57150</xdr:colOff>'
            f'<xdr:row>{s["from_row"]}</xdr:row><xdr:rowOff>28575</xdr:rowOff>'
            f'</xdr:from>'
            f'<xdr:to>'
            f'<xdr:col>{s["to_col"]}</xdr:col><xdr:colOff>0</xdr:colOff>'
            f'<xdr:row>{s["to_row"]}</xdr:row><xdr:rowOff>0</xdr:rowOff>'
            f'</xdr:to>'
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr>'
            f'<xdr:cNvPr id="{i}" name="Note {i}"/>'
            f'<xdr:cNvSpPr/>'
            f'</xdr:nvSpPr>'
            f'<xdr:spPr>'
            f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
            f'<a:solidFill><a:srgbClr val="{fill}"/></a:solidFill>'
            f'<a:ln w="9525">'
            f'<a:solidFill><a:srgbClr val="{border}"/></a:solidFill>'
            f'</a:ln>'
            f'</xdr:spPr>'
            f'<xdr:txBody>'
            f'<a:bodyPr vertOverflow="clip" wrap="square" '
            f'lIns="91440" tIns="45720" rIns="91440" bIns="45720" anchor="t"/>'
            f'<a:lstStyle/>'
            f'{"".join(paras)}'
            f'</xdr:txBody>'
            f'</xdr:sp>'
            f'<xdr:clientData/>'
            f'</xdr:twoCellAnchor>'
        )

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<xdr:wsDr xmlns:xdr="{_XDR}" xmlns:a="{_A}">'
        f'{"".join(anchors)}'
        '</xdr:wsDr>'
    )


def add_sticky_notes(xlsx_path, notes):
    """
    Add rectangle annotation boxes to sheets in an existing xlsx file.

    notes: list of dicts with keys:
        sheet    (str)  — target sheet name
        from_col (int)  — 0-based column of top-left anchor
        from_row (int)  — 0-based row of top-left anchor
        to_col   (int)  — 0-based column of bottom-right anchor
        to_row   (int)  — 0-based row of bottom-right anchor
        lines    (list) — text lines; **bold**, _italic_ markup supported
        fill     (str)  — hex RGB fill color  (default FFFFF0)
        border   (str)  — hex RGB border color (default C0C0C0)
    """
    tmp = xlsx_path + '.tmp'

    by_sheet = {}
    for n in notes:
        by_sheet.setdefault(n['sheet'], []).append(n)

    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        all_names = set(zin.namelist())

        # ── Map sheet names → file paths ──
        wb_xml  = zin.read('xl/workbook.xml').decode('utf-8')
        wb_rels = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')

        sheet_rids = {}
        for m in re.finditer(r'<sheet\b(.*?)/>', wb_xml):
            attrs = m.group(1)
            nm = re.search(r'name="([^"]+)"', attrs)
            rid = re.search(r'r:id="([^"]+)"', attrs)
            if nm and rid:
                sheet_rids[nm.group(1)] = rid.group(1)

        rid_targets = {}
        for m in re.finditer(r'<Relationship\b(.*?)/>', wb_rels):
            attrs = m.group(1)
            id_m = re.search(r'Id="([^"]+)"', attrs)
            tgt  = re.search(r'Target="([^"]+)"', attrs)
            if id_m and tgt:
                rid_targets[id_m.group(1)] = tgt.group(1)

        sheet_paths = {}
        for name, rid in sheet_rids.items():
            t = rid_targets.get(rid, '')
            if t:
                # Handle both absolute (/xl/...) and relative (worksheets/...) targets
                sheet_paths[name] = t.lstrip('/') if t.startswith('/') else 'xl/' + t

        # ── Next available drawing number ──
        existing_dwg = [n for n in all_names if re.match(r'xl/drawings/drawing\d+\.xml$', n)]
        next_num = len(existing_dwg) + 1

        # ── Collect modifications ──
        modifications = {}   # existing file path → new bytes
        additions = {}       # new file path → bytes

        ct = zin.read('[Content_Types].xml').decode('utf-8')

        for sheet_name, shapes in by_sheet.items():
            sp = sheet_paths.get(sheet_name)
            if not sp:
                print(f'  Warning: sheet "{sheet_name}" not found, skipping')
                continue

            dnum = next_num
            next_num += 1
            dpath = f'xl/drawings/drawing{dnum}.xml'

            # Drawing XML
            additions[dpath] = _make_drawing_xml(shapes).encode('utf-8')

            # Sheet rels
            sb = os.path.basename(sp)
            sd = os.path.dirname(sp)
            rels_path = f'{sd}/_rels/{sb}.rels'

            if rels_path in all_names:
                rc = zin.read(rels_path).decode('utf-8')
                max_rid = max((int(x) for x in re.findall(r'Id="rId(\d+)"', rc)), default=0)
                new_rid = f'rId{max_rid + 1}'
                rel_el = (f'<Relationship Id="{new_rid}" '
                          f'Type="{_DRAWING_REL}" '
                          f'Target="../drawings/drawing{dnum}.xml"/>')
                rc = rc.replace('</Relationships>', f'{rel_el}</Relationships>')
                modifications[rels_path] = rc.encode('utf-8')
            else:
                new_rid = 'rId1'
                rc = (
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                    f'<Relationship Id="{new_rid}" '
                    f'Type="{_DRAWING_REL}" '
                    f'Target="../drawings/drawing{dnum}.xml"/>'
                    '</Relationships>'
                )
                additions[rels_path] = rc.encode('utf-8')

            # Add <drawing> ref to sheet XML (include xmlns:r — may not be on <worksheet>)
            _R_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
            sx = zin.read(sp).decode('utf-8')
            sx = sx.replace('</worksheet>',
                            f'<drawing xmlns:r="{_R_NS}" r:id="{new_rid}"/></worksheet>')
            modifications[sp] = sx.encode('utf-8')

            # Content type
            ct = ct.replace('</Types>',
                            f'<Override PartName="/{dpath}" '
                            f'ContentType="{_DRAWING_CT}"/></Types>')

        modifications['[Content_Types].xml'] = ct.encode('utf-8')

        # ── Write output ──
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                if item in modifications:
                    zout.writestr(item, modifications[item])
                else:
                    zout.writestr(item, zin.read(item))
            for path, content in additions.items():
                zout.writestr(path, content)

    os.replace(tmp, xlsx_path)
    print(f'  Injected annotations on {len(by_sheet)} sheet(s) in {os.path.basename(xlsx_path)}')


# ── Standalone test ──

def test():
    test_path = 'build/test_annotations.xlsx'

    # Create a simple workbook with some data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Test Sheet'
    ws['A1'] = 'Annotation Prototype'
    ws['A1'].font = Font(name='Arial', size=10, bold=True)
    for i in range(2, 20):
        ws.cell(row=i, column=1, value=f'Row {i}')
        ws.cell(row=i, column=2, value=i * 1000)
        ws.cell(row=i, column=3, value=i * 500)
    for c in range(1, 4):
        ws.column_dimensions[chr(64 + c)].width = 18
    wb.save(test_path)
    wb.close()
    print(f'Created {test_path}')

    # Inject annotations
    add_sticky_notes(test_path, [
        {
            'sheet': 'Test Sheet',
            'from_col': 4, 'from_row': 1,
            'to_col': 10, 'to_row': 8,
            'lines': [
                '**USCG Data Gap**',
                '',
                'Coast Guard programs (OPC, FRC, PSC, WCC) appear in',
                'sections (B)-(D) with program-level totals from the',
                'DHS Congressional Justification.',
                '',
                '_P-5c/P-8a cost breakdowns are not available._',
                '_USCG uses DHS Capital Investment Exhibits, not DoD P-series._',
            ],
            'fill': 'FFFFF0',
            'border': 'C0C0C0',
        },
        {
            'sheet': 'Test Sheet',
            'from_col': 4, 'from_row': 10,
            'to_col': 10, 'to_row': 15,
            'lines': [
                '**Green note**',
                '',
                'Second box with a different color scheme.',
                'Supports **bold** and _italic_ markup.',
            ],
            'fill': 'E8F5E9',
            'border': '81C784',
        },
    ])

    # Verify openpyxl can still load it
    wb2 = openpyxl.load_workbook(test_path)
    print(f'Verified loadable: {wb2.sheetnames}')
    wb2.close()
    print(f'\nOpen {test_path} in Excel to check rendering.')


if __name__ == '__main__':
    test()
