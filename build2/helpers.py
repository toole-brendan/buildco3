"""
helpers.py — Cell-writing, layout, and SUMIFS cascade helpers.
"""

from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

from build2.styles import (
    NOWRAP, ROW_HEIGHT,
    F_TITLE, F_PURPOSE, F_SUBSEC, F_SUBSUBSEC, F_HDR, F_DATA, F_TOTAL, F_PCT_T, F_KPI,
    BG_BLACK, BG_DGRAY, BG_GRAY, BG_TEAL,
    B_HDR, B_TOT, NUM_FMT, PCT_FMT,
)


# ── SUMIFS cascade builders ──────────────────────────────────

def make_cascade(best_value_range):
    """Build cf / cf_neg / cf_inner using a pre-computed best-value column
    and the JB_RC (Row Class) helper column."""
    bv = best_value_range

    def cf(*conditions):
        cond = (',' + ','.join(conditions)) if conditions else ''
        return f'=SUMIFS({bv}{cond},JB_RC,"ADD")'

    def cf_neg(*conditions):
        cond = (',' + ','.join(conditions)) if conditions else ''
        return f'=-SUMIFS({bv}{cond},JB_RC,"ADD")'

    def cf_inner(*conditions):
        cond = (',' + ','.join(conditions)) if conditions else ''
        return f'SUMIFS({bv}{cond},JB_RC,"ADD")'

    return cf, cf_neg, cf_inner


def make_altview_cascade(best_value_range):
    """Build cf / cf_inner for [ALT_VIEW] rows using pre-computed best-value column."""
    bv = best_value_range

    def cf(*conditions):
        cond = (',' + ','.join(conditions)) if conditions else ''
        return f'=SUMIFS({bv}{cond},JB_RC,"ALT")'

    def cf_inner(*conditions):
        cond = (',' + ','.join(conditions)) if conditions else ''
        return f'SUMIFS({bv}{cond},JB_RC,"ALT")'

    return cf, cf_inner


# ── Cell & layout helpers ────────────────────────────────────

def cw(ws, c, w):
    ws.column_dimensions[get_column_letter(c)].width = w


def wc(ws, r, c, v, font=None, fill=None, fmt=None, border=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if fmt:    cell.number_format = fmt
    if border: cell.border = border
    cell.alignment = NOWRAP


def band(ws, r, max_col, text, font, fill):
    """Fill a full-width band (title or section header)."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = fill
        cell.alignment = NOWRAP
    wc(ws, r, 1, text, font=font, fill=fill)


def title_band(ws, r, text, mc):
    band(ws, r, mc, text, F_TITLE, BG_BLACK)


def subsec_band(ws, r, text, mc):
    band(ws, r, mc, text, F_SUBSEC, BG_DGRAY)


def subsubsec_band(ws, r, text, mc):
    band(ws, r, mc, text, F_SUBSUBSEC, BG_GRAY)


def purpose_row(ws, r, text):
    wc(ws, r, 1, text, font=F_PURPOSE)


def hdr_row(ws, r, headers, cs=1, fill=None):
    for i, (t, w) in enumerate(headers):
        c = cs + i
        wc(ws, r, c, t, font=F_HDR, fill=fill, border=B_HDR)
        cw(ws, c, w)


def total_cell(ws, r, c, v):
    wc(ws, r, c, v, font=F_TOTAL, fmt=NUM_FMT, border=B_TOT)


def total_label(ws, r, c, v):
    wc(ws, r, c, v, font=F_TOTAL, border=B_TOT)


def pct_total(ws, r, c, v):
    wc(ws, r, c, v, font=F_PCT_T, fmt=PCT_FMT, border=B_TOT)


def span_top_border(ws, r, mc):
    """Extend medium top border across all cells in row to last active column."""
    for c in range(1, mc + 1):
        cell = ws.cell(row=r, column=c)
        old = cell.border
        cell.border = Border(top=Side(style='medium'),
                             left=old.left, right=old.right, bottom=old.bottom)


def apply_group_borders(ws, r1, r2, col_start, n_usn, n_cg):
    """Add thin left borders at USN/USCG/Total group boundaries for mekko tables."""
    tc = col_start + n_usn + n_cg
    boundaries = [col_start]
    if n_cg > 0:
        boundaries.append(col_start + n_usn)
    boundaries.append(tc)
    for c in boundaries:
        for r in range(r1, r2 + 1):
            cell = ws.cell(row=r, column=c)
            old = cell.border
            cell.border = Border(left=Side(style='thin'),
                                 top=old.top, right=old.right, bottom=old.bottom)


def svc_group(types, svc_map):
    usn = [t for t in types if svc_map.get(t, '') in ('USN', 'MSC', '')]
    cg = [t for t in types if svc_map.get(t, '') == 'USCG']
    return usn, cg


def svc_order(types, svc_map):
    """Return types ordered USN/MSC first, USCG second."""
    usn, cg = svc_group(types, svc_map)
    return usn + cg


def write_svc_header(ws, r, types, svc_map, col_start=2):
    """Write USN / USCG label row for mekko sections with sub-subsection fill.
    Returns (ordered_types, n_usn, n_cg)."""
    usn, cg = svc_group(types, svc_map)
    ordered = usn + cg
    span = col_start + len(ordered)
    for c in range(1, span + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = BG_GRAY
        cell.alignment = NOWRAP
    if usn: wc(ws, r, col_start, 'USN', font=F_SUBSUBSEC, fill=BG_GRAY)
    if cg:  wc(ws, r, col_start + len(usn), 'USCG', font=F_SUBSUBSEC, fill=BG_GRAY)
    return ordered, len(usn), len(cg)


def finish_sheet(ws):
    """Apply sheet-level formatting: gridlines off, universal row height."""
    ws.sheet_view.showGridLines = False
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT
