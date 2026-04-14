#!/usr/bin/env python3
"""
build_from_data.py — Build market-sizing sheets from an assembled data workbook.

Reads "J Book Items Cons." (Sheet 1) from a pre-assembled workbook (with derived
columns already in place) and generates all market-sizing sheets for every
configured fiscal year. Idempotent: strips and rebuilds calculated sheets on
each run.

Usage:
    python3 -m build2.build_from_data
"""

import openpyxl
import shutil
import os

from build2.config import WORKBOOK_SRC, VALIDATION_SRC, BASE_SHEETS, FY_CONFIGS
from build2.styles import ROW_HEIGHT
from build2.helpers import make_cascade, make_altview_cascade
from build2.data_reader import read_data
from build2.named_ranges import add_helper_columns, create_named_ranges
from build2.config import next_output_path

from build2.sheets.divider import create_divider
from build2.sheets.total_funding import create_total_funding
from build2.sheets.newbuild_tam import create_newbuild_tam
from build2.sheets.newbuild_sam import create_newbuild_sam
from build2.sheets.newbuild_cost_detail import create_newbuild_cost_detail
from build2.sheets.mro_tam import create_mro_tam
from build2.sheets.mro_sam import create_mro_sam
from build2.sheets.competitive_dynamics import create_competitive_dynamics
from build2.sheets.validation import import_validation_sheet

import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__)))
from build2.prototype_annotations import add_sticky_notes


def main():
    out_path = next_output_path()
    print(f'Source: {WORKBOOK_SRC}')
    print(f'Output: {out_path}')
    shutil.copy2(WORKBOOK_SRC, out_path)

    wb = openpyxl.load_workbook(out_path)

    # Strip all non-base sheets
    for nm in list(wb.sheetnames):
        if nm not in BASE_SHEETS:
            del wb[nm]
    print(f'Base sheets kept: {wb.sheetnames}')

    # Helper columns (Row Class, FY best-value) + named ranges
    add_helper_columns(wb)
    create_named_ranges(wb)

    # Build each fiscal year
    generated = []
    fy26_sam_info = None; fy26_detail_info = None
    for fyc in FY_CONFIGS:
        label  = fyc['label']
        prefix = fyc['prefix']
        tab_color = fyc['tab_color']
        print(f'\n--- {label} ---')

        d = read_data(WORKBOOK_SRC, fyc['read_indices'])
        print(f'  TAM types: {len(d["all_tam_types"])}, SAM types: {len(d["all_sam_types"])}')
        print(f'  NB non-zero: {len(d["nb_nz"])}, NB SAM non-zero: {len(d["nb_sam_nz"])}')
        print(f'  MRO non-zero: {len(d["mro_nz"])}, MRO SAM non-zero: {len(d["mro_sam_nz"])}')
        print(f'  MRO buckets with data: {d["mro_bkts_with_data"]}')
        print(f'  NB hulls non-zero: {len(d["nb_hull_nz"])}, P-5c hulls: {len(d["p5c_hull_types"])}, P-8a hulls: {len(d["p8a_hull_types"])}')

        # Divider tab
        divider_name = create_divider(wb, label, tab_color)

        cf, cf_neg, cf_inner = make_cascade(fyc['best_value_range'])
        acf, acf_inner = make_altview_cascade(fyc['best_value_range'])

        args = (wb, d, label, prefix, cf, cf_neg, cf_inner)
        create_total_funding(*args)
        create_newbuild_tam(*args)
        sam_info = create_newbuild_sam(*args, altview_cf=acf, altview_inner=acf_inner)
        mro_tam_info = create_mro_tam(*args)
        create_mro_sam(*args, mro_tam_info=mro_tam_info)

        # Supplemental cost detail (currently FY26 only — P-5c/P-8a data availability)
        cost_detail_name = None
        detail_info = None
        if d.get('p5c_hull_types'):
            detail_info = create_newbuild_cost_detail(wb, d, label, prefix, acf_inner)
            cost_detail_name = f'{prefix} Newbuild SAM Cost Detail'

        if prefix == 'FY26':
            fy26_sam_info = sam_info
            fy26_detail_info = detail_info

        # Apply tab colors to all sheets in this FY group
        fy_sheets = [
            f'{prefix} Total Funding', f'{prefix} Newbuild TAM',
            f'{prefix} Newbuild SAM',  f'{prefix} MRO TAM',
            f'{prefix} MRO SAM',
        ]
        if cost_detail_name:
            fy_sheets.append(cost_detail_name)
        for sn in fy_sheets:
            wb[sn].sheet_properties.tabColor = tab_color

        generated.append(divider_name)
        generated.extend([
            f'{prefix} Total Funding', f'{prefix} Newbuild TAM',
            f'{prefix} Newbuild SAM',
        ])
        if cost_detail_name:
            generated.append(cost_detail_name)
        generated.extend([
            f'{prefix} MRO TAM',
            f'{prefix} MRO SAM',
        ])

    # Competitive Dynamics
    print('\n--- Competitive Dynamics ---')
    create_competitive_dynamics(wb)
    generated.append('Competitive Dynamics')

    # Validation sheet (imported from standalone file)
    if os.path.exists(VALIDATION_SRC):
        print('\n--- Validation ---')
        import_validation_sheet(wb, VALIDATION_SRC)
        generated.append('Validation')
    else:
        print(f'\nWarning: {VALIDATION_SRC} not found — skipping Validation sheet')

    # Order: Sheet 1, then generated sheets in FY order, then remaining base sheets
    base_first = 'J Book Items Cons.'
    base_rest  = [s for s in wb.sheetnames if s in BASE_SHEETS and s != base_first]
    ordered    = [base_first] + generated + base_rest
    wb._sheets = [wb[s] for s in ordered]

    wb.save(out_path)

    # Post-save: inject sticky-note annotations via OOXML
    annotations = []

    if fy26_sam_info and fy26_sam_info.get('section_e_row'):
        er = fy26_sam_info['section_e_row'] - 1
        tc = fy26_sam_info['tc_e']
        annotations.append({
            'sheet': 'FY26 Newbuild SAM',
            'from_col': tc, 'from_row': er,
            'to_col': tc + 4, 'to_row': er + 12,
            'lines': [
                '**USCG Data Gap**',
                '',
                'Sections (E) and (F) cover Navy (SCN)',
                'hull programs only. Coast Guard cutters',
                '(OPC, FRC, PSC, WCC) appear in (B)-(D)',
                'with program-level totals but lack P-5c',
                'and P-8a cost category breakdowns.',
                '',
                '_USCG uses DHS Capital Investment Exhibits,_',
                '_not DoD P-series — no component-level_',
                '_cost data is published._',
            ],
        })
    if fy26_detail_info and fy26_detail_info.get('section_b_row'):
        br = fy26_detail_info['section_b_row'] - 1
        annotations.append({
            'sheet': 'FY26 Newbuild SAM Cost Detail',
            'from_col': 5, 'from_row': br,
            'to_col': 10, 'to_row': br + 9,
            'lines': [
                '**About This Table**',
                '',
                'Individual GFE systems from Navy Exhibit',
                'P-8a — radars, weapons, electronics,',
                'propulsion — grouped by cost category',
                'within each hull. Gross values (pre-AP/SFF)',
                'from SCN justification books. Use for',
                'system market sizing and cross-program',
                'commonality analysis.',
            ],
        })
    if annotations:
        add_sticky_notes(out_path, annotations)

    print(f'\nSaved {out_path}')
    print(f'Sheet order: {[s for s in wb.sheetnames]}')

    # Verify formula lengths
    wb2 = openpyxl.load_workbook(out_path)
    max_len = 0; max_cell = ''
    for nm in generated:
        for row in wb2[nm].iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith('=') and len(v) > max_len:
                    max_len = len(v); max_cell = f'{nm}!R{cell.row}C{cell.column}'
    wb2.close()
    status = 'OK' if max_len < 8192 else 'OVER LIMIT!'
    print(f'Max formula: {max_len} chars at {max_cell} — {status}')


if __name__ == '__main__':
    main()
