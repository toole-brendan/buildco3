"""
newbuild_tam.py — Newbuild TAM sheet builder.
"""

from openpyxl.utils import get_column_letter

from build2.config import TAM_CATEGORIES, TAM_EXCLUDED_CATEGORIES
from build2.styles import F_DATA, F_HDR, F_PCT, F_TOTAL, B_HDR, NUM_FMT, PCT_FMT
from build2.helpers import (
    cw, wc, title_band, subsec_band, purpose_row, hdr_row,
    total_label, total_cell, pct_total, span_top_border,
    apply_group_borders, svc_order, write_svc_header, finish_sheet,
)


def create_newbuild_tam(wb, d, label, prefix, cf, cf_neg, cf_inner):
    ws = wb.create_sheet(f'{prefix} Newbuild TAM')
    mc = max(4, 2 + len(d['nb_nz']) + 1)
    cw(ws, 1, 30); cw(ws, 2, 12)

    r = 1; title_band(ws, r, f'Newbuild TAM \u2014 {label} ($K)', mc)
    r = 2; purpose_row(ws, r,
        'Total Funding narrowed to new construction only and only oceangoing vessels — '
        'excludes small craft (Combatant Crafts, Support Crafts) and USCG Boats. '
        'Limited to line items with clear vessel type designation; unattributed fleet-wide items are excluded.')

    # (A) Bridge
    r = 4; subsec_band(ws, r, '(A) Bridge: Total NC Funding \u2192 Newbuild TAM', 2)
    r += 1; wc(ws, r, 1, 'Total New Construction Funding', font=F_DATA)
    wc(ws, r, 2, cf('JB_B,1'), font=F_DATA, fmt=NUM_FMT)
    r += 1; wc(ws, r, 1, '  Less: Unattributed (fleet-wide NC items)', font=F_DATA)
    wc(ws, r, 2, cf_neg('JB_B,1', 'JB_V,""'), font=F_DATA, fmt=NUM_FMT)
    for cat in TAM_EXCLUDED_CATEGORIES:
        r += 1; wc(ws, r, 1, f'  Less: {cat}', font=F_DATA)
        wc(ws, r, 2, cf_neg('JB_B,1', f'JB_V,"{cat}"'), font=F_DATA, fmt=NUM_FMT)
    r += 1; total_label(ws, r, 1, 'Newbuild TAM')
    tam_f = '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES))
    total_cell(ws, r, 2, '=' + tam_f)
    span_top_border(ws, r, 2)

    # (B) All vessel types
    r += 3; subsec_band(ws, r, '(B) Newbuild TAM by Vessel Type', 4)
    all_types = svc_order(d['all_tam_types'], d['type_svc'])
    r += 1; hdr_row(ws, r, [('Vessel Type', 30), ('Service', 8), (f'{label} ($K)', 12), ('% of TAM', 10)])
    ft = r + 1
    for vt in all_types:
        r += 1; wc(ws, r, 1, vt, font=F_DATA)
        wc(ws, r, 2, d['type_svc'].get(vt, ''), font=F_DATA)
        wc(ws, r, 3, cf('JB_B,1', f'JB_W,"{vt}"'), font=F_DATA, fmt=NUM_FMT)
    lt = r; r += 1; tr = r
    total_label(ws, r, 1, 'Newbuild TAM'); total_label(ws, r, 2, '')
    total_cell(ws, r, 3, f'=SUM(C{ft}:C{lt})')
    pct_total(ws, r, 4, 1.0)
    span_top_border(ws, r, 4)
    for rn in range(ft, lt + 1):
        wc(ws, rn, 4, f'=IFERROR(C{rn}/C${tr},0)', font=F_PCT, fmt=PCT_FMT)

    # (C) Mekko
    mk = svc_order(d['nb_nz'], d['type_svc'])
    r += 3; subsec_band(ws, r, '(C) Newbuild TAM \u2014 DD&C vs AP/LLTM by Vessel Type (Mekko)', 2 + len(mk))
    r += 1; svc_r = r; mk, n_usn, n_cg = write_svc_header(ws, r, d['nb_nz'], d['type_svc'])
    r += 1; wc(ws, r, 1, 'Sub-Category', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
    for i, vt in enumerate(mk):
        c = 2 + i; wc(ws, r, c, vt, font=F_HDR, border=B_HDR); cw(ws, c, 12)
    tc = 2 + len(mk); wc(ws, r, tc, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc, 12)
    dr = r + 1; ar = r + 2; orr = r + 3; tkr = r + 4

    wc(ws, tkr, 1, 'Total ($K)', font=F_TOTAL, border=B_HDR)
    for i, vt in enumerate(mk):
        total_cell(ws, tkr, 2 + i, cf('JB_B,1', f'JB_W,"{vt}"'))
    total_cell(ws, tkr, tc, f'=SUM({get_column_letter(2)}{tkr}:{get_column_letter(tc-1)}{tkr})')

    ddc = ['"Full ship DD&C"', '"Full ship DD&C / LLTM"', '"Full vessel procurement"']
    wc(ws, dr, 1, 'Full Ship DD&C', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        parts = '+'.join(cf_inner('JB_B,1', f'JB_W,"{vt}"', f'JB_F,{s}') for s in ddc)
        wc(ws, dr, c, f'=IFERROR(({parts})/{cl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)
    tcl = get_column_letter(tc)
    wc(ws, dr, tc, f'=IFERROR(1-{tcl}{ar},0)', font=F_PCT, fmt=PCT_FMT)

    wc(ws, ar, 1, 'Advance Procurement / LLTM', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        ap = cf_inner('JB_B,1', f'JB_W,"{vt}"', 'JB_F,"Advance procurement / LLTM"')
        wc(ws, ar, c, f'=IFERROR(({ap})/{cl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)
    ap_cat = '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"', 'JB_F,"Advance procurement / LLTM"') for c in sorted(TAM_CATEGORIES))
    wc(ws, ar, tc, f'=IFERROR(({ap_cat})/{tcl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)

    wc(ws, orr, 1, 'Other NC Sub-Categories', font=F_DATA)
    for i in range(len(mk)):
        c = 2 + i; cl = get_column_letter(c)
        wc(ws, orr, c, f'=IFERROR(1-{cl}{dr}-{cl}{ar},0)', font=F_PCT, fmt=PCT_FMT)
    wc(ws, orr, tc, f'=IFERROR(1-{tcl}{dr}-{tcl}{ar},0)', font=F_PCT, fmt=PCT_FMT)

    span_top_border(ws, tkr, tc)
    apply_group_borders(ws, svc_r, tkr, 2, n_usn, n_cg)
    r = tkr

    # (D) Newbuild TAM by Hull Program
    hull_tam_nz = d.get('nb_hull_tam_nz', [])
    if hull_tam_nz:
        r += 3
        subsec_band(ws, r, f'(D) Newbuild TAM by Hull Program \u2014 {label} ($K)', 4)
        r += 1
        purpose_row(ws, r, 'Same funding as (B) broken down to individual hull programs.')
        r += 1
        hdr_row(ws, r, [('Hull Program', 30), ('Service', 8), (f'{label} ($K)', 12), ('% of TAM', 10)])
        ft_d = r + 1
        for hl in hull_tam_nz:
            r += 1
            vt_label = d['hull_type'].get(hl, '')
            wc(ws, r, 1, f'{hl} ({vt_label})' if vt_label else hl, font=F_DATA)
            wc(ws, r, 2, d['hull_svc'].get(hl, ''), font=F_DATA)
            wc(ws, r, 3, cf('JB_B,1', f'JB_H,"{hl}"'), font=F_DATA, fmt=NUM_FMT)
        lt_d = r; r += 1; tr_d = r
        total_label(ws, r, 1, 'Newbuild TAM'); total_label(ws, r, 2, '')
        total_cell(ws, r, 3, f'=SUM(C{ft_d}:C{lt_d})')
        pct_total(ws, r, 4, 1.0)
        span_top_border(ws, r, 4)
        for rn in range(ft_d, lt_d + 1):
            wc(ws, rn, 4, f'=IFERROR(C{rn}/C${tr_d},0)', font=F_PCT, fmt=PCT_FMT)

    finish_sheet(ws)
