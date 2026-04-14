"""
validation.py — Import the Validation sheet from a standalone xlsx.
"""

import openpyxl
from copy import copy


def import_validation_sheet(wb, src_path):
    """Copy the Validation sheet from a standalone xlsx into the workbook."""
    src_wb = openpyxl.load_workbook(src_path)
    ws_src = src_wb['Validation']
    ws_dst = wb.create_sheet('Validation')

    for row in ws_src.iter_rows(min_row=1, max_row=ws_src.max_row, max_col=ws_src.max_column):
        for cell in row:
            new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

    for col_letter, dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col_letter].width = dim.width
    for row_num, dim in ws_src.row_dimensions.items():
        ws_dst.row_dimensions[row_num].height = dim.height

    ws_dst.sheet_view.showGridLines = False
    src_wb.close()
    print(f'  Imported {ws_src.max_row} rows from {src_path}')
