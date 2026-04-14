"""
mro_sam.py — MRO SAM sheet builder.
"""

from openpyxl.utils import get_column_letter

from build2.config import TAM_CATEGORIES, MRO_BKTS
from build2.styles import F_DATA, F_HDR, F_PCT, F_TOTAL, F_GREEN, B_HDR, NUM_FMT, PCT_FMT
from build2.helpers import (
    cw, wc, title_band, subsec_band, purpose_row, hdr_row,
    total_label, total_cell, pct_total, span_top_border,
    apply_group_borders, svc_order, write_svc_header, finish_sheet,
)


def create_mro_sam(wb, d, label, prefix, cf, cf_neg, cf_inner, mro_tam_info=None):
    ws = wb.create_sheet(f'{prefix} MRO SAM')
    mc = max(6, 2 + len(d['mro_sam_nz']) + 1)
    cw(ws, 1, 30); cw(ws, 2, 12)

    tam_sheet = f'{prefix} MRO TAM'

    r = 1; title_band(ws, r, f'MRO SAM \u2014 {label} ($K)', mc)
    if prefix == 'FY27':
        r = 2; purpose_row(ws, r,
            'MRO TAM narrowed further: excludes Submarines, Aircraft Carriers & UUVs, and draws work types down to '
            'scheduled depot maintenance and modernization/alteration installation — the outsourceable categories '
            'where a company could compete for repair or upgrade work. Justification books (P-5c, P-8a) not yet '
            'released for FY27 — visibility is at program level only; cost element granularity is not yet available.')
    else:
        r = 2; purpose_row(ws, r,
            'MRO TAM narrowed further: excludes Submarines, Aircraft Carriers & UUVs, and draws work types down to '
            'scheduled depot maintenance and modernization/alteration installation — the outsourceable categories '
            'where a company could compete for repair or upgrade work. P-5c and P-8a cost elements are the target '
            'for component-level visibility, but granularity is harder to isolate for O&M-funded work than for '
            'procurement-funded new construction.')

    # (A) Bridge
    r = 4; subsec_band(ws, r, '(A) Bridge: MRO TAM \u2192 MRO SAM', 2)
    r += 1; wc(ws, r, 1, 'MRO TAM', font=F_DATA)
    tam_ref_row = mro_tam_info['bkt_total_row'] if mro_tam_info else 18
    wc(ws, r, 2, f"='{tam_sheet}'!B{tam_ref_row}", font=F_GREEN, fmt=NUM_FMT)

    excl_rows = []
    for vt in ['Submarines', 'Aircraft Carriers', 'Unmanned Undersea Vehicles']:
        r += 1; excl_rows.append(r)
        wc(ws, r, 1, f'  Less: {vt}', font=F_DATA)
        parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_W,"{vt}"') for b in MRO_BKTS)
        wc(ws, r, 2, f'=-({parts})', font=F_DATA, fmt=NUM_FMT)

    r += 1; nowork_row = r
    wc(ws, r, 1, '  Less: Non-outsourceable work types', font=F_DATA)

    r += 1; sam_result_row = r
    total_label(ws, r, 1, 'MRO SAM')
    sam_parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_V,"{c}"') for b in ['2', '4'] for c in sorted(TAM_CATEGORIES))
    excl_parts = '+'.join(
        f'-({cf_inner(f"JB_B,{b}", f"""JB_W,"{vt}" """.strip())})'
        for b in ['2', '4'] for vt in ['Submarines', 'Aircraft Carriers', 'Unmanned Undersea Vehicles']
    )
    total_cell(ws, r, 2, f'={sam_parts}{excl_parts}')
    span_top_border(ws, r, 2)

    excl_sum = '+'.join(f'B{er}' for er in excl_rows)
    wc(ws, nowork_row, 2, f'=-(B5+{excl_sum}-B{sam_result_row})', font=F_DATA, fmt=NUM_FMT)

    # (B) All SAM vessel types
    r += 3; subsec_band(ws, r, '(B) MRO SAM by Vessel Type', 6)
    all_types = svc_order(d['all_sam_mro_sorted'], d['type_svc'])
    r += 1; hdr_row(ws, r, [('Vessel Type', 30), ('Service', 8), ('SDM ($K)', 12), ('Mod ($K)', 12), ('Total ($K)', 12), ('% of SAM', 10)])
    ft = r + 1
    for vt in all_types:
        r += 1; wc(ws, r, 1, vt, font=F_DATA)
        wc(ws, r, 2, d['type_svc'].get(vt, ''), font=F_DATA)
        wc(ws, r, 3, cf('JB_B,2', f'JB_W,"{vt}"'), font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, 4, cf('JB_B,4', f'JB_W,"{vt}"'), font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, 5, f'=C{r}+D{r}', font=F_DATA, fmt=NUM_FMT)
    lt = r; r += 1; tr = r
    total_label(ws, r, 1, 'MRO SAM'); total_label(ws, r, 2, '')
    for c in [3, 4, 5]:
        total_cell(ws, r, c, f'=SUM({get_column_letter(c)}{ft}:{get_column_letter(c)}{lt})')
    pct_total(ws, r, 6, 1.0)
    span_top_border(ws, r, 6)
    for rn in range(ft, lt + 1):
        wc(ws, rn, 6, f'=IFERROR(E{rn}/E${tr},0)', font=F_PCT, fmt=PCT_FMT)

    # (C) Mekko
    mk = svc_order(d['mro_sam_nz'], d['type_svc'])
    r += 3; subsec_band(ws, r, '(C) MRO SAM \u2014 SDM vs Modernization by Vessel Type (Mekko)', 2 + len(mk))
    r += 1; svc_r = r; mk, n_usn, n_cg = write_svc_header(ws, r, d['mro_sam_nz'], d['type_svc'], col_start=2)
    r += 1; wc(ws, r, 1, 'Work Type', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
    for i, vt in enumerate(mk):
        c = 2 + i; wc(ws, r, c, vt, font=F_HDR, border=B_HDR); cw(ws, c, 12)
    tc = 2 + len(mk); wc(ws, r, tc, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc, 12)
    sr = r + 1; mr = r + 2; tkr = r + 3

    wc(ws, tkr, 1, 'Total ($K)', font=F_TOTAL, border=B_HDR)
    for i, vt in enumerate(mk):
        c = 2 + i
        parts = cf_inner('JB_B,2', f'JB_W,"{vt}"') + '+' + cf_inner('JB_B,4', f'JB_W,"{vt}"')
        total_cell(ws, tkr, c, '=' + parts)
    total_cell(ws, tkr, tc, f'=SUM({get_column_letter(2)}{tkr}:{get_column_letter(tc-1)}{tkr})')

    tcl = get_column_letter(tc)
    wc(ws, sr, 1, 'Scheduled Depot Maintenance', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        wc(ws, sr, c, f'=IFERROR(({cf_inner("JB_B,2", f"""JB_W,"{vt}" """.strip())})/{cl}${tkr},0)',
           font=F_PCT, fmt=PCT_FMT)
    wc(ws, sr, tc, f'=IFERROR(1-{tcl}{mr},0)', font=F_PCT, fmt=PCT_FMT)

    wc(ws, mr, 1, 'Modernization', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        wc(ws, mr, c, f'=IFERROR(({cf_inner("JB_B,4", f"""JB_W,"{vt}" """.strip())})/{cl}${tkr},0)',
           font=F_PCT, fmt=PCT_FMT)
    mod_parts = '+'.join(cf_inner('JB_B,4', f'JB_W,"{vt}"') for vt in mk)
    wc(ws, mr, tc, f'=IFERROR(({mod_parts})/{tcl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)

    span_top_border(ws, tkr, tc)
    apply_group_borders(ws, svc_r, tkr, 2, n_usn, n_cg)

    finish_sheet(ws)
