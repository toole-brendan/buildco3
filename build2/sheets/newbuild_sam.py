"""
newbuild_sam.py — Newbuild SAM sheet builder.
"""

from openpyxl.utils import get_column_letter

from build2.config import (
    TAM_CATEGORIES, SAM_EXCLUDED_TYPES, P5C_COST_CATEGORIES, P5C_VALID,
)
from build2.styles import (
    F_DATA, F_HDR, F_PCT, F_TOTAL, F_KPI, F_GREEN, F_GRAY,
    BG_GRAY, BG_TEAL,
    B_HDR, NUM_FMT, PCT_FMT,
)
from build2.helpers import (
    cw, wc, title_band, subsec_band, subsubsec_band, purpose_row, hdr_row,
    total_label, total_cell, pct_total, span_top_border,
    apply_group_borders, svc_order, write_svc_header, finish_sheet,
)


def create_newbuild_sam(wb, d, label, prefix, cf, cf_neg, cf_inner, altview_cf=None, altview_inner=None):
    ws = wb.create_sheet(f'{prefix} Newbuild SAM')
    mc = max(4, 2 + len(d['nb_sam_nz']) + 1)
    cw(ws, 1, 30); cw(ws, 2, 12)

    r = 1; title_band(ws, r, f'Newbuild SAM \u2014 {label} ($K)', mc)
    if prefix == 'FY27':
        r = 2; purpose_row(ws, r,
            'Newbuild TAM narrowed further: excludes Submarines and Aircraft Carriers '
            '(single-yard / nuclear-restricted programs) and Unmanned Undersea Vehicles. '
            'Focused on outsourceable new construction where a company could compete as subprime '
            'on module fabrication or systems integration — scoped to full-ship DD&C funding '
            '(not advance procurement). Justification books (P-5c, P-8a) not yet released for FY27 — '
            'visibility is at program level only, not cost element level.')
    else:
        r = 2; purpose_row(ws, r,
            'Newbuild TAM narrowed further: excludes Submarines and Aircraft Carriers '
            '(single-yard / nuclear-restricted programs) and Unmanned Undersea Vehicles. '
            'Focused on outsourceable new construction where a company could compete as subprime '
            'on module fabrication or systems integration — scoped to full-ship DD&C funding '
            '(not advance procurement) with P-5c and P-8a cost elements providing '
            'component-level visibility where available.')

    # (A) Bridge
    r = 4; subsec_band(ws, r, '(A) Bridge: Newbuild TAM \u2192 Newbuild SAM', 2)
    r += 1; bridge_first = r
    wc(ws, r, 1, 'Newbuild TAM', font=F_DATA)
    wc(ws, r, 2, '=' + '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES)), font=F_DATA, fmt=NUM_FMT)
    for vt in ['Submarines', 'Aircraft Carriers', 'Unmanned Undersea Vehicles']:
        r += 1; wc(ws, r, 1, f'  Less: {vt}', font=F_DATA)
        wc(ws, r, 2, cf_neg('JB_B,1', f'JB_W,"{vt}"'), font=F_DATA, fmt=NUM_FMT)
    r += 1; total_label(ws, r, 1, 'Newbuild SAM')
    total_cell(ws, r, 2, f'=SUM(B{bridge_first}:B{r-1})')
    span_top_border(ws, r, 2)

    # (B) All SAM vessel types
    r += 3; subsec_band(ws, r, '(B) Newbuild SAM by Vessel Type', 4)
    all_types = svc_order(d['all_sam_types'], d['type_svc'])
    r += 1; hdr_row(ws, r, [('Vessel Type', 30), ('Service', 8), (f'{label} ($K)', 12), ('% of SAM', 10)])
    ft = r + 1
    for vt in all_types:
        r += 1; wc(ws, r, 1, vt, font=F_DATA)
        wc(ws, r, 2, d['type_svc'].get(vt, ''), font=F_DATA)
        wc(ws, r, 3, cf('JB_B,1', f'JB_W,"{vt}"'), font=F_DATA, fmt=NUM_FMT)
    lt = r; r += 1; tr = r
    total_label(ws, r, 1, 'Newbuild SAM'); total_label(ws, r, 2, '')
    total_cell(ws, r, 3, f'=SUM(C{ft}:C{lt})')
    pct_total(ws, r, 4, 1.0)
    span_top_border(ws, r, 4)
    for rn in range(ft, lt + 1):
        wc(ws, rn, 4, f'=IFERROR(C{rn}/C${tr},0)', font=F_PCT, fmt=PCT_FMT)

    # (C) Mekko
    mk = svc_order(d['nb_sam_nz'], d['type_svc'])
    r += 3; subsec_band(ws, r, '(C) Newbuild SAM \u2014 DD&C vs AP/LLTM by Vessel Type (Mekko)', 2 + len(mk))
    r += 1; svc_r = r; mk, n_usn, n_cg = write_svc_header(ws, r, d['nb_sam_nz'], d['type_svc'])
    r += 1; wc(ws, r, 1, 'Sub-Category', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
    for i, vt in enumerate(mk):
        c = 2 + i; wc(ws, r, c, vt, font=F_HDR, border=B_HDR); cw(ws, c, 12)
    tc = 2 + len(mk); wc(ws, r, tc, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc, 12)
    dr = r + 1; ar = r + 2; tkr = r + 3

    wc(ws, tkr, 1, 'Total ($K)', font=F_TOTAL, border=B_HDR)
    for i, vt in enumerate(mk):
        total_cell(ws, tkr, 2 + i, cf('JB_B,1', f'JB_W,"{vt}"'))
    total_cell(ws, tkr, tc, f'=SUM({get_column_letter(2)}{tkr}:{get_column_letter(tc-1)}{tkr})')

    ddc = ['"Full ship DD&C"', '"Full ship DD&C / LLTM"', '"Full vessel procurement"']
    wc(ws, dr, 1, 'Full Ship DD&C', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        parts = '+'.join(cf_inner('JB_B,1', f'JB_W,"{vt}"', f'JB_F,{s}') for s in ddc)
        wc(ws, dr, c, f'=IFERROR(({parts})/{cl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)
    tcl = get_column_letter(tc)
    wc(ws, dr, tc, f'=IFERROR(1-{tcl}{ar},0)', font=F_PCT, fmt=PCT_FMT)

    wc(ws, ar, 1, 'Advance Procurement / LLTM', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        ap = cf_inner('JB_B,1', f'JB_W,"{vt}"', 'JB_F,"Advance procurement / LLTM"')
        wc(ws, ar, c, f'=IFERROR(({ap})/{cl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)
    ap_cat = '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"', 'JB_F,"Advance procurement / LLTM"') for c in sorted(TAM_CATEGORIES))
    wc(ws, ar, tc, f'=IFERROR(({ap_cat})/{tcl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)

    span_top_border(ws, tkr, tc)
    apply_group_borders(ws, svc_r, tkr, 2, n_usn, n_cg)
    r = tkr

    # (D) Newbuild SAM by Hull Program
    hull_nz = d.get('nb_hull_nz', [])
    if hull_nz:
        r += 3
        subsec_band(ws, r, f'(D) Newbuild SAM by Hull Program \u2014 {label} ($K)', 4)
        r += 1
        purpose_row(ws, r, 'Same funding as (B) broken down to individual hull programs \u2014 the level at which budget line items are traceable.')
        r += 1
        hdr_row(ws, r, [('Hull Program', 30), ('Service', 8), (f'{label} ($K)', 12), ('% of SAM', 10)])
        ft_d = r + 1
        for hl in hull_nz:
            r += 1
            vt_label = d['hull_type'].get(hl, '')
            wc(ws, r, 1, f'{hl} ({vt_label})' if vt_label else hl, font=F_DATA)
            wc(ws, r, 2, d['hull_svc'].get(hl, ''), font=F_DATA)
            wc(ws, r, 3, cf('JB_B,1', f'JB_H,"{hl}"'), font=F_DATA, fmt=NUM_FMT)
        lt_d = r; r += 1; tr_d = r
        total_label(ws, r, 1, 'Newbuild SAM'); total_label(ws, r, 2, '')
        total_cell(ws, r, 3, f'=SUM(C{ft_d}:C{lt_d})')
        pct_total(ws, r, 4, 1.0)
        span_top_border(ws, r, 4)
        for rn in range(ft_d, lt_d + 1):
            wc(ws, rn, 4, f'=IFERROR(C{rn}/C${tr_d},0)', font=F_PCT, fmt=PCT_FMT)

    # (E) P-5c Cost Category Breakdown by Hull & (F) Proportional Allocation
    _section_e_row = None; _tc_e = None

    p5c_hulls = d.get('p5c_hull_types', [])
    if p5c_hulls and altview_inner:
        cats = [c for c in P5C_COST_CATEGORIES
                if any(d['p5c_hull'].get((h, c), 0) > 0 for h in p5c_hulls)]

        n_hulls = len(p5c_hulls)
        tc_e = 2 + n_hulls
        _tc_e = tc_e

        # (E) Gross P-5c Cost Category Breakdown by Hull
        r += 3; _section_e_row = r
        subsec_band(ws, r, f'(E) P-5c Gross Cost Category Breakdown by Hull \u2014 {label} ($K)', tc_e)
        r += 1
        purpose_row(ws, r, 'Gross total-ship-estimate costs from Exhibit P-5c. These are pre-AP/SFF and will NOT match SAM net totals.')
        r += 1
        svc_r_e = r
        p5c_ordered, n_usn_e, n_cg_e = write_svc_header(ws, r, p5c_hulls, d['hull_svc'])

        r += 1
        wc(ws, r, 1, 'Cost Category', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
        for i, hl in enumerate(p5c_ordered):
            c = 2 + i
            wc(ws, r, c, hl, font=F_HDR, border=B_HDR); cw(ws, c, 14)
        wc(ws, r, tc_e, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc_e, 14)

        first_e = r + 1
        for cat in cats:
            r += 1
            wc(ws, r, 1, cat, font=F_DATA)
            for i, hl in enumerate(p5c_ordered):
                c = 2 + i
                formula = altview_inner('JB_B,1', f'JB_EX,"P-5c"', f'JB_CC,"{cat}"', f'JB_H,"{hl}"')
                wc(ws, r, c, '=' + formula, font=F_DATA, fmt=NUM_FMT)
            wc(ws, r, tc_e, f'=SUM({get_column_letter(2)}{r}:{get_column_letter(tc_e-1)}{r})',
               font=F_DATA, fmt=NUM_FMT)
        last_e = r

        r += 1; tkr_e = r
        wc(ws, r, 1, 'Gross Total ($K)', font=F_TOTAL, border=B_HDR)
        for c in range(2, tc_e + 1):
            total_cell(ws, r, c,
                       f'=SUM({get_column_letter(c)}{first_e}:{get_column_letter(c)}{last_e})')
        span_top_border(ws, r, tc_e)
        apply_group_borders(ws, svc_r_e, tkr_e, 2, n_usn_e, n_cg_e)

        # Percentage sub-table
        r += 2
        subsubsec_band(ws, r, 'Cost category mix (% of gross total per hull program)', tc_e)
        r += 1
        wc(ws, r, 1, 'Cost Category', font=F_HDR, border=B_HDR)
        for i, hl in enumerate(p5c_ordered):
            wc(ws, r, 2 + i, hl, font=F_HDR, border=B_HDR)
        wc(ws, r, tc_e, 'Total', font=F_HDR, border=B_HDR)

        pct_first_e = r + 1
        for ci, cat in enumerate(cats):
            r += 1
            data_row = first_e + ci
            wc(ws, r, 1, cat, font=F_DATA)
            for c in range(2, tc_e + 1):
                cl = get_column_letter(c)
                wc(ws, r, c, f'=IFERROR({cl}{data_row}/{cl}${tkr_e},0)',
                   font=F_PCT, fmt=PCT_FMT)
        pct_last_e = r
        r += 1
        wc(ws, r, 1, 'Total', font=F_TOTAL, border=B_HDR)
        for c in range(2, tc_e + 1):
            pct_total(ws, r, c, 1.0)
        span_top_border(ws, r, tc_e)
        apply_group_borders(ws, pct_first_e - 2, r, 2, n_usn_e, n_cg_e)

        # (F) Proportional Cost Category Allocation by Hull to SAM Net Totals
        r += 3
        subsec_band(ws, r, f'(F) Proportional Cost Category Allocation to SAM Net Totals \u2014 {label} ($K)', tc_e)
        r += 1
        purpose_row(ws, r, 'P-5c percentage mix from (E) applied to SAM net budget-authority totals from (D). Totals reconcile to SAM.')
        r += 1
        svc_r_f = r
        p5c_ordered_f, n_usn_f, n_cg_f = write_svc_header(ws, r, p5c_hulls, d['hull_svc'])

        r += 1
        wc(ws, r, 1, 'Cost Category', font=F_HDR, border=B_HDR)
        for i, hl in enumerate(p5c_ordered_f):
            c = 2 + i
            wc(ws, r, c, hl, font=F_HDR, border=B_HDR); cw(ws, c, 14)
        wc(ws, r, tc_e, 'Total', font=F_HDR, border=B_HDR)

        r += 1; sam_ref_r = r
        wc(ws, r, 1, 'SAM Net Total ($K)', font=F_KPI, fill=BG_TEAL)
        for i, hl in enumerate(p5c_ordered_f):
            c = 2 + i
            wc(ws, r, c, cf('JB_B,1', f'JB_H,"{hl}"'), font=F_KPI, fill=BG_TEAL, fmt=NUM_FMT)
        wc(ws, r, tc_e, f'=SUM({get_column_letter(2)}{r}:{get_column_letter(tc_e-1)}{r})',
           font=F_KPI, fill=BG_TEAL, fmt=NUM_FMT)

        for ci, cat in enumerate(cats):
            r += 1
            pct_row = pct_first_e + ci
            wc(ws, r, 1, cat, font=F_DATA)
            for c in range(2, tc_e + 1):
                cl = get_column_letter(c)
                wc(ws, r, c, f'=IFERROR({cl}${sam_ref_r}*{cl}{pct_row},0)',
                   font=F_DATA, fmt=NUM_FMT)

        apply_group_borders(ws, svc_r_f, r, 2, n_usn_f, n_cg_f)

    finish_sheet(ws)
    return {'section_e_row': _section_e_row, 'tc_e': _tc_e}
