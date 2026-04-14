"""
mro_tam.py — MRO TAM sheet builder.
"""

from openpyxl.utils import get_column_letter

from build2.config import TAM_CATEGORIES, TAM_EXCLUDED_CATEGORIES, MRO_BKTS, BUCKET_NAMES, BUCKET_SHORT
from build2.styles import F_DATA, F_HDR, F_PCT, F_TOTAL, B_HDR, NUM_FMT, PCT_FMT
from build2.helpers import (
    cw, wc, title_band, subsec_band, purpose_row, hdr_row,
    total_label, total_cell, pct_total, span_top_border,
    apply_group_borders, svc_order, write_svc_header, finish_sheet,
)


def create_mro_tam(wb, d, label, prefix, cf, cf_neg, cf_inner):
    ws = wb.create_sheet(f'{prefix} MRO TAM')
    mc = max(4, 2 + len(d['mro_nz']) + 1)
    cw(ws, 1, 30); cw(ws, 2, 12)

    bkts = d['mro_bkts_with_data']

    r = 1; title_band(ws, r, f'MRO TAM \u2014 {label} ($K)', mc)
    r = 2; purpose_row(ws, r,
        'Total Funding narrowed to all post-construction work — depot maintenance, continuous & emergent repair, '
        'modernization & alterations, major life-cycle events, sustainment engineering, and availability support — '
        'and only oceangoing vessels. Same vessel scope as Newbuild TAM; limited to line items with clear vessel type designation.')

    # (A) Bridge
    r = 4; subsec_band(ws, r, '(A) Bridge: Total MRO Funding \u2192 MRO TAM', 2)
    r += 1; wc(ws, r, 1, 'Total MRO Funding (Buckets 2\u20137)', font=F_DATA)
    wc(ws, r, 2, '=' + '+'.join(cf_inner(f'JB_B,{b}') for b in MRO_BKTS), font=F_DATA, fmt=NUM_FMT)
    r += 1; wc(ws, r, 1, '  Less: Unattributed (fleet-wide MRO items)', font=F_DATA)
    wc(ws, r, 2, f'=-({"+".join(cf_inner(f"JB_B,{b}", "JB_V," + chr(34) + chr(34)) for b in MRO_BKTS)})',
       font=F_DATA, fmt=NUM_FMT)
    for cat in TAM_EXCLUDED_CATEGORIES:
        r += 1; wc(ws, r, 1, f'  Less: {cat}', font=F_DATA)
        wc(ws, r, 2, f'=-({"+".join(cf_inner(f"JB_B,{b}", f"""JB_V,"{cat}" """.strip()) for b in MRO_BKTS)})',
           font=F_DATA, fmt=NUM_FMT)
    r += 1; bridge_tam_row = r
    total_label(ws, r, 1, 'MRO TAM')

    # (B) By Work Type
    r += 3; subsec_band(ws, r, '(B) MRO TAM by Work Type', 3)
    r += 1; hdr_row(ws, r, [('Work Type', 30), (f'{label} ($K)', 12), ('% of TAM', 10)])
    fb = r + 1
    for b in bkts:
        r += 1; wc(ws, r, 1, BUCKET_SHORT.get(b, BUCKET_NAMES[b]), font=F_DATA)
        parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES))
        wc(ws, r, 2, '=' + parts, font=F_DATA, fmt=NUM_FMT)
    lb = r; r += 1; btr = r
    total_label(ws, r, 1, 'MRO TAM')
    total_cell(ws, r, 2, f'=SUM(B{fb}:B{lb})')
    pct_total(ws, r, 3, 1.0)
    span_top_border(ws, r, 3)
    for rn in range(fb, lb + 1):
        wc(ws, rn, 3, f'=IFERROR(B{rn}/B${btr},0)', font=F_PCT, fmt=PCT_FMT)

    # Fill bridge total
    total_cell(ws, bridge_tam_row, 2, f'=B{btr}')
    span_top_border(ws, bridge_tam_row, 2)

    # (C) All vessel types
    r += 3; subsec_band(ws, r, '(C) MRO TAM by Vessel Type', 4)
    all_types = svc_order(d['all_tam_mro_sorted'], d['type_svc'])
    r += 1; hdr_row(ws, r, [('Vessel Type', 30), ('Service', 8), (f'{label} ($K)', 12), ('% of TAM', 10)])
    ft = r + 1
    for vt in all_types:
        r += 1; wc(ws, r, 1, vt, font=F_DATA)
        wc(ws, r, 2, d['type_svc'].get(vt, ''), font=F_DATA)
        mparts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_W,"{vt}"') for b in MRO_BKTS)
        wc(ws, r, 3, '=' + mparts, font=F_DATA, fmt=NUM_FMT)
    lt = r; r += 1; ttr = r
    total_label(ws, r, 1, 'MRO TAM'); total_label(ws, r, 2, '')
    total_cell(ws, r, 3, f'=SUM(C{ft}:C{lt})')
    pct_total(ws, r, 4, 1.0)
    span_top_border(ws, r, 4)
    for rn in range(ft, lt + 1):
        wc(ws, rn, 4, f'=IFERROR(C{rn}/C${ttr},0)', font=F_PCT, fmt=PCT_FMT)

    # (D) Mekko
    mk = svc_order(d['mro_nz'], d['type_svc'])
    r += 3; subsec_band(ws, r, '(D) MRO TAM \u2014 Work Type by Vessel Type (Mekko)', 2 + len(mk))
    r += 1; svc_r = r; mk, n_usn, n_cg = write_svc_header(ws, r, d['mro_nz'], d['type_svc'], col_start=2)
    r += 1; wc(ws, r, 1, 'Work Type', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
    for i, vt in enumerate(mk):
        c = 2 + i; wc(ws, r, c, vt, font=F_HDR, border=B_HDR); cw(ws, c, 12)
    tc = 2 + len(mk); wc(ws, r, tc, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc, 12)
    nwt = len(bkts); tkr = r + 1 + nwt

    wc(ws, tkr, 1, 'Total ($K)', font=F_TOTAL, border=B_HDR)
    for i, vt in enumerate(mk):
        c = 2 + i; mparts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_W,"{vt}"') for b in MRO_BKTS)
        total_cell(ws, tkr, c, '=' + mparts)
    total_cell(ws, tkr, tc, f'=SUM({get_column_letter(2)}{tkr}:{get_column_letter(tc-1)}{tkr})')

    tcl = get_column_letter(tc)
    for bi, b in enumerate(bkts):
        wr = r + 1 + bi
        wc(ws, wr, 1, BUCKET_SHORT.get(b, BUCKET_NAMES[b]), font=F_DATA)
        for i, vt in enumerate(mk):
            c = 2 + i; cl = get_column_letter(c)
            wc(ws, wr, c, f'=IFERROR(({cf_inner(f"JB_B,{b}", f"""JB_W,"{vt}" """.strip())})/{cl}${tkr},0)',
               font=F_PCT, fmt=PCT_FMT)
        bkt_parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES))
        wc(ws, wr, tc, f'=IFERROR(({bkt_parts})/{tcl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)

    span_top_border(ws, tkr, tc)
    apply_group_borders(ws, svc_r, tkr, 2, n_usn, n_cg)

    finish_sheet(ws)
    return {'bkt_total_row': btr}
