"""
newbuild_cost_detail.py — Supplemental P-5c/P-8a system-level detail sheet.
"""

from openpyxl.utils import get_column_letter
from collections import defaultdict

from build2.config import (
    SAM_EXCLUDED_TYPES, P5C_COST_CATEGORIES, P5C_VALID, P8A_COST_CATEGORIES,
)
from build2.styles import (
    F_DATA, F_HDR, F_PCT, F_TOTAL, F_GREEN, F_GRAY,
    B_HDR, NUM_FMT, PCT_FMT,
)
from build2.helpers import (
    cw, wc, title_band, subsec_band, subsubsec_band, purpose_row, hdr_row,
    total_label, total_cell, pct_total, span_top_border,
    apply_group_borders, write_svc_header, finish_sheet,
)


def create_newbuild_cost_detail(wb, d, label, prefix, acf_inner):
    """Supplemental sheet: direct P-5c costs + P-8a system detail.
    Organized by hull program. Only created when P-5c hull data is available."""
    p5c_hulls = d.get('p5c_hull_types', [])
    p8a_hulls = d.get('p8a_hull_types', [])
    if not p5c_hulls:
        return

    ws = wb.create_sheet(f'{prefix} Newbuild SAM Cost Detail')

    cats = [c for c in P5C_COST_CATEGORIES
            if any(d['p5c_hull'].get((h, c), 0) > 0 for h in p5c_hulls)]

    tc = 2 + len(p5c_hulls)
    mc = max(tc, 6)
    cw(ws, 1, 32)

    r = 1; title_band(ws, r, f'Newbuild SAM Cost Detail \u2014 {label} ($K)', mc)
    r = 2; purpose_row(ws, r,
        'Gross ship-cost breakdown from P-5c and P-8a exhibits by hull program. '
        'NOT net budget authority \u2014 see Newbuild SAM for reconciled values.')

    # (A) Direct P-5c Cost Category Breakdown by Hull
    r = 4; subsec_band(ws, r,
        f'(A) P-5c Total Ship Cost by Hull & Cost Category \u2014 {label} ($K)', tc)
    r += 1
    svc_r = r
    ordered, n_usn, n_cg = write_svc_header(ws, r, p5c_hulls, d['hull_svc'])

    r += 1
    wc(ws, r, 1, 'Cost Category', font=F_HDR, border=B_HDR)
    for i, hl in enumerate(ordered):
        c = 2 + i
        wc(ws, r, c, hl, font=F_HDR, border=B_HDR); cw(ws, c, 14)
    wc(ws, r, tc, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc, 14)

    first_a = r + 1
    for cat in cats:
        r += 1
        wc(ws, r, 1, cat, font=F_DATA)
        for i, hl in enumerate(ordered):
            c = 2 + i
            formula = acf_inner('JB_B,1', f'JB_EX,"P-5c"', f'JB_CC,"{cat}"', f'JB_H,"{hl}"')
            wc(ws, r, c, '=' + formula, font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, tc, f'=SUM({get_column_letter(2)}{r}:{get_column_letter(tc-1)}{r})',
           font=F_DATA, fmt=NUM_FMT)
    last_a = r

    r += 1; tkr_a = r
    wc(ws, r, 1, 'Gross Total ($K)', font=F_TOTAL, border=B_HDR)
    for c in range(2, tc + 1):
        total_cell(ws, r, c,
                   f'=SUM({get_column_letter(c)}{first_a}:{get_column_letter(c)}{last_a})')
    span_top_border(ws, r, tc)
    apply_group_borders(ws, svc_r, tkr_a, 2, n_usn, n_cg)

    # (B) P-8a System-Level Detail by Hull Program
    r += 3; _section_b_row = r; subsec_band(ws, r,
        f'(B) P-8a System-Level Detail by Hull Program \u2014 {label} ($K)', 4)
    r += 1
    purpose_row(ws, r, 'Individual GFE systems from Exhibit P-8a, grouped by cost category within each hull program.')

    for hl in p8a_hulls:
        systems = []
        for (hl2, cc, ce), val in d['p8a_hull'].items():
            if hl2 == hl and val > 0:
                systems.append((cc, ce, val))
        if not systems:
            continue

        r += 2
        vt_label = d['hull_type'].get(hl, '')
        subsubsec_band(ws, r, f'{hl} ({vt_label})' if vt_label else hl, 4)
        r += 1
        hdr_row(ws, r, [('System', 32), ('Cost Category', 22), (f'{label} ($K)', 14), ('% of Hull', 10)])
        ft = r + 1

        cat_order = list(P8A_COST_CATEGORIES)
        systems_sorted = sorted(systems, key=lambda x: (cat_order.index(x[0])
                                                         if x[0] in cat_order else 99, -x[2]))
        for cc, ce, val in systems_sorted:
            r += 1
            wc(ws, r, 1, ce, font=F_DATA)
            wc(ws, r, 2, cc, font=F_GRAY)
            ce_esc = ce.replace('"', '""')
            formula = acf_inner('JB_B,1', f'JB_EX,"P-8a"', f'JB_CC,"{cc}"',
                                f'JB_H,"{hl}"', f'JB_CE,"{ce_esc}"')
            wc(ws, r, 3, '=' + formula, font=F_GREEN, fmt=NUM_FMT)
        lt = r

        r += 1
        total_label(ws, r, 1, f'{hl} Total')
        total_cell(ws, r, 3, f'=SUM(C{ft}:C{lt})')
        span_top_border(ws, r, 4)

        for rn in range(ft, lt + 1):
            wc(ws, rn, 4, f'=IFERROR(C{rn}/C${r},0)', font=F_PCT, fmt=PCT_FMT)
        pct_total(ws, r, 4, 1.0)

    # (C) Top P-8a Systems — Cross-Program Summary
    r += 3; subsec_band(ws, r,
        f'(C) Top P-8a Systems \u2014 Cross-Program Summary \u2014 {label} ($K)', 5)
    r += 1
    purpose_row(ws, r, 'Largest GFE systems aggregated across all SAM hull programs.')

    sys_agg = defaultdict(lambda: {'total': 0, 'category': '', 'hulls': []})
    for (hl, cc, ce), val in d['p8a_hull'].items():
        if d['hull_type'].get(hl, '') in SAM_EXCLUDED_TYPES or val <= 0:
            continue
        sys_agg[ce]['total'] += val
        sys_agg[ce]['category'] = cc
        sys_agg[ce]['hulls'].append(hl)

    top_systems = sorted(sys_agg.items(), key=lambda x: -x[1]['total'])[:30]

    r += 1
    hdr_row(ws, r, [('System', 32), ('Cost Category', 22), (f'{label} ($K)', 14),
                     ('# Hulls', 10), ('Hull Programs', 24)])
    ft = r + 1
    for ce, info in top_systems:
        r += 1
        wc(ws, r, 1, ce, font=F_DATA)
        wc(ws, r, 2, info['category'], font=F_GRAY)
        ce_esc = ce.replace('"', '""')
        formula = acf_inner('JB_B,1', f'JB_EX,"P-8a"', f'JB_CE,"{ce_esc}"')
        wc(ws, r, 3, '=' + formula, font=F_GREEN, fmt=NUM_FMT)
        wc(ws, r, 4, len(set(info['hulls'])), font=F_DATA)
        wc(ws, r, 5, ', '.join(sorted(set(info['hulls']))), font=F_GRAY)
    lt = r

    r += 1
    total_label(ws, r, 1, 'Total (top systems)')
    total_cell(ws, r, 3, f'=SUM(C{ft}:C{lt})')
    span_top_border(ws, r, 5)

    for c in range(2, 11):
        cw(ws, c, 15)

    finish_sheet(ws)
    return {'section_b_row': _section_b_row}
