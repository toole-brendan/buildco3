"""
named_ranges.py — Helper column formulas and named range definitions.
"""

from openpyxl.workbook.defined_name import DefinedName

from build2.config import FY_CONFIGS


def add_helper_columns(wb):
    """Extend helper-column formulas to cover the full named-range span (rows 6-4000).

    Column layout (inserted between FY27 dollars and Reference):
      AC (29) — Row Class: "ADD" for additive, "ALT" for ALT_VIEW, blank otherwise
      AD (30) — FY26 Best ($K): cascade DAA Enacted > Total > Request
      AE (31) — FY27 Best ($K): cascade Total > Request
    """
    ws = wb['J Book Items Cons.']
    RC_COL, BV_START = 29, 30

    for r in range(6, 4001):
        if ws.cell(row=r, column=RC_COL).value is None:
            ws.cell(row=r, column=RC_COL,
                    value=f'=IF(OR(E{r}="",E{r}="[PARENT]"),"ADD",IF(E{r}="[ALT_VIEW]","ALT",""))')
        for i, fyc in enumerate(FY_CONFIGS):
            c = BV_START + i
            if ws.cell(row=r, column=c).value is None:
                ws.cell(row=r, column=c,
                        value=f'={fyc["best_value_formula"].format(r=r)}')


def create_named_ranges(wb):
    JB = "'J Book Items Cons.'"
    shared = [
        ('JB_A', 'A'),
        ('JB_B', 'I'),
        ('JB_F', 'J'),
        ('JB_H', 'N'),
        ('JB_S', 'E'),
        ('JB_U', 'K'),
        ('JB_V', 'L'),
        ('JB_W', 'M'),
        ('JB_CE', 'AJ'),
        ('JB_EX', 'G'),
        ('JB_CC', 'H'),
        ('JB_RC', 'AC'),
    ]
    bv_col_letters = ['AD', 'AE', 'AF', 'AG']
    for i, fyc in enumerate(FY_CONFIGS):
        shared.append((fyc['best_value_range'], bv_col_letters[i]))

    fy_ranges = []
    for fyc in FY_CONFIGS:
        fy_ranges.extend(fyc['named_ranges'])

    for nm, col in shared + fy_ranges:
        if nm in wb.defined_names:
            del wb.defined_names[nm]
        wb.defined_names.add(DefinedName(nm, attr_text=f"{JB}!${col}$6:${col}$4000"))
