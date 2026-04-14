"""
styles.py — Fonts, fills, borders, alignment, and number formats.
"""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ── Layout constants ─────────────────────────────────────────

FONT_NAME = 'Arial'
FONT_SIZE = 8
ROW_HEIGHT = 10.0
NOWRAP = Alignment(wrap_text=False, vertical='center')

# ── Text-color palette ───────────────────────────────────────

BLACK = '000000'
WHITE = 'FFFFFF'
SLATE = '5B7A99'
GREEN = '2E7D32'
NAVY = '1F314F'

# ── Fonts ────────────────────────────────────────────────────

F_TITLE = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=WHITE)
F_PURPOSE = Font(name=FONT_NAME, size=FONT_SIZE, italic=True, color='808080')
F_SUBSEC = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=WHITE)
F_SUBSUBSEC = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=BLACK)
F_HDR = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=BLACK)
F_DATA = Font(name=FONT_NAME, size=FONT_SIZE, color=BLACK)
F_GREEN = Font(name=FONT_NAME, size=FONT_SIZE, color=GREEN)
F_GRAY = Font(name=FONT_NAME, size=FONT_SIZE, italic=True, color='808080')
F_TOTAL = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=BLACK)
F_PCT = Font(name=FONT_NAME, size=FONT_SIZE, italic=True, color=BLACK)
F_PCT_T = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, italic=True, color=BLACK)
F_KPI = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=NAVY)
F_BLUE = Font(name=FONT_NAME, size=FONT_SIZE, color='0000FF')

# ── Fills ────────────────────────────────────────────────────

BG_BLACK = PatternFill(start_color=BLACK, end_color=BLACK, fill_type='solid')
BG_DGRAY = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
BG_GRAY = PatternFill(start_color='F2F4F7', end_color='F2F4F7', fill_type='solid')
BG_TEAL = PatternFill(start_color='E0F2F1', end_color='E0F2F1', fill_type='solid')

# ── Borders ──────────────────────────────────────────────────

B_HDR = Border(bottom=Side(style='thin'))
B_TOT = Border(top=Side(style='medium'))

# ── Number formats ───────────────────────────────────────────

NUM_FMT = '#,##0;[Red](#,##0);"-"'
PCT_FMT = '0.0%;[Red](0.0%);"-"'
