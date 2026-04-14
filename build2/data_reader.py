"""
data_reader.py — Pre-scan the data sheet to find which vessel types / buckets have data.
"""

import openpyxl
from collections import defaultdict

from build2.config import (
    TAM_CATEGORIES, SAM_EXCLUDED_TYPES, MRO_BKTS,
    P5C_COST_CATEGORIES, P5C_VALID, P8A_COST_CATEGORIES,
)


def parse_exhibit_level(title):
    """Extract exhibit level from an ALT_VIEW title string.
    Returns 'P-5c', 'P-8a', 'P-35', or None."""
    if not title:
        return None
    t = str(title).strip()
    if ' - P-5c - ' in t:
        return 'P-5c'
    if ' - P-8a ' in t:
        for cat in P8A_COST_CATEGORIES:
            if f' - P-8a {cat} - ' in t:
                return 'P-8a'
        return None
    if ' - P-35 ' in t:
        return 'P-35'
    return None


def parse_cost_category(title, exhibit_level):
    """Extract P-5c cost category from an ALT_VIEW title string.
    For P-5c rows: the category after ' - P-5c - '.
    For P-8a/P-35 rows: the P-5c parent category (Electronics, HM&E, Ordnance).
    Returns a category string or None."""
    if not title or not exhibit_level:
        return None
    t = str(title).strip()

    if exhibit_level == 'P-5c':
        cat = t.split(' - P-5c - ', 1)[1].strip()
        return cat if cat in P5C_VALID else None

    if exhibit_level == 'P-8a':
        for cat in P8A_COST_CATEGORIES:
            if f' - P-8a {cat} - ' in t:
                return cat
        return None

    if exhibit_level == 'P-35':
        return None

    return None


def read_data(src_path, col_indices):
    """
    Pre-scan Sheet 1 to find which vessel types / buckets have data.
    col_indices: 0-indexed column positions to try (cascade order).
    """
    def read_val(row):
        for i in col_indices:
            v = row[i].value
            if v is not None and isinstance(v, (int, float)): return float(v)
        return 0.0

    def additive(rt):
        if rt is None or str(rt).strip() == '': return True
        return str(rt).strip() == '[PARENT]'

    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = wb['J Book Items Cons.']

    nb = defaultdict(float); mro = defaultdict(float)
    mro_bkt = defaultdict(lambda: defaultdict(float)); mro_b = defaultdict(float)
    type_svc = {}; type_cat = {}
    all_tam_types = set(); tf_src = defaultdict(float)
    nb_excl = defaultdict(float); mro_excl = defaultdict(float)
    p5c_data = defaultdict(float)
    p5c_types = set()
    p8a_data = defaultdict(float)
    p8a_types = set()
    nb_hull = defaultdict(float)
    hull_svc = {}
    hull_type = {}
    p5c_hull = defaultdict(float)
    p5c_hull_set = set()
    p8a_hull = defaultdict(float)
    p8a_hull_set = set()

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        cat = str(row[11].value).strip() if row[11].value else ''
        vt  = str(row[12].value).strip() if row[12].value else ''
        sv  = str(row[10].value).strip() if row[10].value else ''
        hl  = str(row[13].value).strip() if row[13].value else ''

        if cat in TAM_CATEGORIES and vt:
            all_tam_types.add(vt)
            if sv: type_svc[vt] = sv
            type_cat[vt] = cat
            if hl:
                if sv: hull_svc[hl] = sv
                hull_type[hl] = vt

        rt = row[4].value
        bk = row[8].value
        bs = str(bk).strip() if bk is not None else ''

        # ALT_VIEW scan (P-5c / P-8a for cost category breakdown)
        if str(rt).strip() == '[ALT_VIEW]' and bk is not None and bs == '1':
            title = str(row[3].value).strip() if row[3].value else ''
            ex = parse_exhibit_level(title)
            cc_val = parse_cost_category(title, ex)
            if ex and cc_val and cat in TAM_CATEGORIES and vt and vt not in SAM_EXCLUDED_TYPES:
                val = read_val(row)
                if ex == 'P-5c' and cc_val in P5C_VALID:
                    p5c_data[(vt, cc_val)] += val
                    p5c_types.add(vt)
                    if hl:
                        p5c_hull[(hl, cc_val)] += val
                        p5c_hull_set.add(hl)
                elif ex == 'P-8a':
                    ce = str(row[35].value).strip() if row[35].value else ''
                    if ce:
                        p8a_data[(vt, cc_val, ce)] += val
                        p8a_types.add(vt)
                        if hl:
                            p8a_hull[(hl, cc_val, ce)] += val
                            p8a_hull_set.add(hl)

        if not additive(rt): continue
        if bk is None: continue
        if bs not in ('1', '2', '3', '4', '5', '6', '7'): continue

        v = read_val(row)
        src = str(row[0].value).strip() if row[0].value else '(blank)'
        tf_src[src] += v

        if bs == '1':
            if cat in TAM_CATEGORIES:
                nb[vt] += v
                if hl: nb_hull[hl] += v
            elif cat: nb_excl[cat] += v
        elif bs in MRO_BKTS:
            if cat in TAM_CATEGORIES:
                mro[vt] += v; mro_bkt[vt][bs] += v; mro_b[bs] += v
            elif cat: mro_excl[cat] += v

    wb.close()

    # ── Post-scan sorting and filtering ──────────────────────

    def sorted_svc(types, val_map):
        usn = sorted([t for t in types if type_svc.get(t, '') in ('USN', 'MSC', '')],
                      key=lambda t: -val_map.get(t, 0))
        cg  = sorted([t for t in types if type_svc.get(t, '') == 'USCG'],
                      key=lambda t: -val_map.get(t, 0))
        return usn + cg

    def sorted_svc_hull(hulls, val_map):
        usn = sorted([h for h in hulls if hull_svc.get(h, '') in ('USN', 'MSC', '')],
                      key=lambda h: -val_map.get(h, 0))
        cg  = sorted([h for h in hulls if hull_svc.get(h, '') == 'USCG'],
                      key=lambda h: -val_map.get(h, 0))
        return usn + cg

    all_sam_types = all_tam_types - SAM_EXCLUDED_TYPES
    nb_nz      = [t for t in sorted_svc(all_tam_types, nb)  if nb.get(t, 0) > 0]
    nb_sam_nz  = [t for t in nb_nz if t not in SAM_EXCLUDED_TYPES]
    mro_nz     = [t for t in sorted_svc(all_tam_types, mro) if mro.get(t, 0) > 0]
    mro_sam_nz = [t for t in mro_nz if t not in SAM_EXCLUDED_TYPES
                  and (mro_bkt[t].get('2', 0) + mro_bkt[t].get('4', 0)) > 0]

    return {
        'all_tam_types':       sorted_svc(all_tam_types, nb),
        'all_sam_types':       sorted_svc(all_sam_types, nb),
        'nb_nz': nb_nz, 'nb_sam_nz': nb_sam_nz,
        'mro_nz': mro_nz, 'mro_sam_nz': mro_sam_nz,
        'all_tam_mro_sorted':  sorted_svc(all_tam_types, mro),
        'all_sam_mro_sorted':  sorted_svc(all_sam_types, {
            t: mro_bkt[t].get('2', 0) + mro_bkt[t].get('4', 0) for t in all_sam_types
        }),
        'type_svc': type_svc,
        'mro_bkts_with_data':  [b for b in MRO_BKTS if mro_b.get(b, 0) > 0],
        'sources':             [s for s, _ in sorted(tf_src.items(), key=lambda x: -x[1])],
        'nb_excl':  [c for c, v in sorted(nb_excl.items(), key=lambda x: -x[1]) if v > 0],
        'mro_excl': [c for c, v in sorted(mro_excl.items(), key=lambda x: -x[1]) if v > 0],
        'p5c_data': dict(p5c_data),
        'p5c_types': sorted_svc([t for t in p5c_types
                                  if any(p5c_data.get((t, c), 0) > 0 for c in P5C_VALID)],
                                 nb),
        'p8a_data': dict(p8a_data),
        'p8a_types': sorted_svc(list(p8a_types), nb),
        'nb_hull': dict(nb_hull),
        'nb_hull_tam_nz': [h for h in sorted_svc_hull(
            {h for h in nb_hull}, nb_hull)
            if nb_hull.get(h, 0) > 0],
        'nb_hull_nz': [h for h in sorted_svc_hull(
            {h for h in nb_hull if hull_type.get(h, '') not in SAM_EXCLUDED_TYPES}, nb_hull)
            if nb_hull.get(h, 0) > 0],
        'hull_svc': hull_svc,
        'hull_type': hull_type,
        'p5c_hull': dict(p5c_hull),
        'p5c_hull_types': sorted_svc_hull(
            [h for h in p5c_hull_set if hull_type.get(h, '') not in SAM_EXCLUDED_TYPES
             and any(p5c_hull.get((h, c), 0) > 0 for c in P5C_VALID)],
            {h: sum(p5c_hull.get((h, c), 0) for c in P5C_VALID) for h in p5c_hull_set}),
        'p8a_hull': dict(p8a_hull),
        'p8a_hull_types': sorted_svc_hull(
            [h for h in p8a_hull_set if hull_type.get(h, '') not in SAM_EXCLUDED_TYPES],
            {h: sum(v for (h2, _, _), v in p8a_hull.items() if h2 == h) for h in p8a_hull_set}),
    }
