#!/usr/bin/env python3
"""
build_workbook.py — Master build script for the Newbuild & MRO Spend workbook.

Reads "J Book Items Cons." (Sheet 1) and generates all market-sizing sheets
for every configured fiscal year. Idempotent: strips and rebuilds calculated
sheets on each run.

Usage:
    python3 build/build_workbook.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from collections import defaultdict
import shutil
import glob
import re

# ── Configuration ─────────────────────────────────────────────

WORKBOOK_SRC = 'output/08APR2028_Newbuild_and_MRO_Spend_v2.3.xlsx'
OUTPUT_DIR   = 'output'
FILE_PREFIX  = '08APR2028_Newbuild_and_MRO_Spend'
VERSION_MAJOR = 3  # auto-increments minor: v3.0, v3.1, v3.2, ...


def next_output_path():
    """Scan output/ for existing v{MAJOR}.x files and return the next version."""
    pattern = re.compile(rf'{re.escape(FILE_PREFIX)}_v{VERSION_MAJOR}\.(\d+)\.xlsx$')
    max_minor = -1
    for path in glob.glob(f'{OUTPUT_DIR}/{FILE_PREFIX}_v{VERSION_MAJOR}.*.xlsx'):
        m = pattern.search(path)
        if m:
            max_minor = max(max_minor, int(m.group(1)))
    next_minor = max_minor + 1
    return f'{OUTPUT_DIR}/{FILE_PREFIX}_v{VERSION_MAJOR}.{next_minor}.xlsx'

# Sheets that are hand-built data — never touched by this script
BASE_SHEETS = {'J Book Items Cons.', 'Validation'}

# Each entry produces 5 sheets (Total Funding, Newbuild TAM/SAM, MRO TAM/SAM).
FY_CONFIGS = [
    {
        'label': 'FY2026',
        'prefix': 'FY26',
        'named_ranges': [('JB_R', 'R'), ('JB_N', 'N'), ('JB_L', 'L')],
        'value_ranges': ['JB_R', 'JB_N', 'JB_L'],
        'read_indices': [17, 13, 11],
        'tab_color': '4472C4',   # muted blue
    },
    {
        'label': 'FY2027',
        'prefix': 'FY27',
        'named_ranges': [('JB_27U', 'U'), ('JB_27S', 'S')],
        'value_ranges': ['JB_27U', 'JB_27S'],
        'read_indices': [20, 18],
        'tab_color': '548235',   # muted green
    },
]


# ── Taxonomy ──────────────────────────────────────────────────

TAM_CATEGORIES = {'Combatant Ships', 'Auxiliary Ships', 'Cutters', 'Unmanned Maritime Platforms'}
TAM_EXCLUDED_CATEGORIES = ['Combatant Crafts', 'Support Crafts']
SAM_EXCLUDED_TYPES = {'Submarines', 'Aircraft Carriers', 'Unmanned Undersea Vehicles'}

BUCKET_NAMES = {
    '1': 'New Construction', '2': 'Scheduled Depot Maintenance & Repair',
    '3': 'Continuous / Intermediate / Emergent Maintenance',
    '4': 'Modernization & Alteration Installation',
    '5': 'Major Life-Cycle Events / SLEP / MMA / RCOH',
    '6': 'Sustainment Engineering / Planning / Obsolescence Support',
    '7': 'Availability Support / Husbanding / Port Services',
}
BUCKET_SHORT = {
    '2': 'Scheduled Depot Maintenance', '3': 'Continuous / Emergent Maintenance',
    '4': 'Modernization', '5': 'Major Life-Cycle Events / RCOH',
    '6': 'Sustainment Engineering', '7': 'Availability Support',
}
MRO_BKTS = ['2', '3', '4', '5', '6', '7']


# ── Styles (per workbook styling guide) ──────────────────────

_F = 'Arial'
_S = 8
_RH = 10.0
_NOWRAP = Alignment(wrap_text=False, vertical='center')

# Text-color palette
_BLACK  = '000000'
_WHITE  = 'FFFFFF'
_SLATE  = '5B7A99'   # secondary slate (91,122,153)
_GREEN  = '2E7D32'   # cross-tab link
_NAVY   = '1F314F'   # primary navy (31,49,79)

# Fonts — all Arial 8, hierarchy via bold/italic/color only
F_TITLE    = Font(name=_F, size=_S, bold=True,  color=_WHITE)   # white on black band
F_PURPOSE  = Font(name=_F, size=_S, italic=True, color='808080') # gray italic purpose line
F_SUBSEC   = Font(name=_F, size=_S, bold=True,  color=_WHITE)   # subsection header (white on gray)
F_SUBSUBSEC = Font(name=_F, size=_S, bold=True, color=_BLACK)  # sub-subsection header (black on light gray)
F_HDR      = Font(name=_F, size=_S, bold=True,  color=_BLACK)   # column header
F_DATA     = Font(name=_F, size=_S, color=_BLACK)               # formula / label
F_GREEN    = Font(name=_F, size=_S, color=_GREEN)               # cross-tab link
F_GRAY     = Font(name=_F, size=_S, italic=True, color='808080') # note / helper text
F_TOTAL    = Font(name=_F, size=_S, bold=True,  color=_BLACK)   # total row
F_PCT      = Font(name=_F, size=_S, italic=True, color=_BLACK)  # percentage data
F_PCT_T    = Font(name=_F, size=_S, bold=True, italic=True, color=_BLACK)  # percentage total
F_KPI      = Font(name=_F, size=_S, bold=True,  color=_NAVY)    # KPI callout
F_BLUE     = Font(name=_F, size=_S, color='0000FF')             # hardcoded numeric input

# Fills
BG_BLACK = PatternFill(start_color=_BLACK,  end_color=_BLACK,  fill_type='solid')
BG_DGRAY = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # subsection header
BG_GRAY  = PatternFill(start_color='F2F4F7', end_color='F2F4F7', fill_type='solid')  # sub-subsection / col headers
BG_TEAL  = PatternFill(start_color='E0F2F1', end_color='E0F2F1', fill_type='solid')  # 224,242,241

# Borders
B_HDR     = Border(bottom=Side(style='thin'))     # column headers
B_TOT     = Border(top=Side(style='medium'))      # all total rows

# Number formats — zeros as "-", negatives in red parentheses
NUM_FMT = '#,##0;[Red](#,##0);"-"'
PCT_FMT = '0.0%;[Red](0.0%);"-"'


# ── Generic SUMIFS cascade ────────────────────────────────────

def make_cascade(value_ranges):
    """
    Build cf / cf_neg / cf_inner for a value-column cascade.

    value_ranges: ordered list of named-range names, e.g. ['JB_R','JB_N','JB_L'].
    The cascade tries each in priority order.  For each additive row type
    ("" and "[PARENT]"), it emits one SUMIFS per column:
      - prior columns must be empty
      - current column must be non-empty (except the last, which catches the rest)
    """
    def cf(*conditions):
        cond = (',' + ','.join(conditions)) if conditions else ''
        t = []
        for rt in ['""', '"[PARENT]"']:
            for i, vr in enumerate(value_ranges):
                prior = ''.join(f',{value_ranges[j]},""' for j in range(i))
                check = f',{vr},"<>"' if i < len(value_ranges) - 1 else ''
                t.append(f'SUMIFS({vr}{cond},JB_S,{rt}{prior}{check})')
        return '=' + '+'.join(t)

    def cf_neg(*conditions):
        return f'=-({cf(*conditions)[1:]})'

    def cf_inner(*conditions):
        return cf(*conditions)[1:]

    return cf, cf_neg, cf_inner


# ── Cell & layout helpers ─────────────────────────────────────

def cw(ws, c, w):
    ws.column_dimensions[get_column_letter(c)].width = w

def wc(ws, r, c, v, font=None, fill=None, fmt=None, border=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if fmt:    cell.number_format = fmt
    if border: cell.border = border
    cell.alignment = _NOWRAP

def band(ws, r, max_col, text, font, fill):
    """Fill a full-width band (title or section header)."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = fill
        cell.alignment = _NOWRAP
    wc(ws, r, 1, text, font=font, fill=fill)

def title_band(ws, r, text, mc):
    band(ws, r, mc, text, F_TITLE, BG_BLACK)

def subsec_band(ws, r, text, mc):
    band(ws, r, mc, text, F_SUBSEC, BG_DGRAY)

def subsubsec_band(ws, r, text, mc):
    band(ws, r, mc, text, F_SUBSUBSEC, BG_GRAY)

def purpose_row(ws, r, text):
    wc(ws, r, 1, text, font=F_PURPOSE)

def hdr_row(ws, r, headers, cs=1, fill=None):
    for i, (t, w) in enumerate(headers):
        c = cs + i
        wc(ws, r, c, t, font=F_HDR, fill=fill, border=B_HDR)
        cw(ws, c, w)

def total_cell(ws, r, c, v):
    wc(ws, r, c, v, font=F_TOTAL, fmt=NUM_FMT, border=B_TOT)

def total_label(ws, r, c, v):
    wc(ws, r, c, v, font=F_TOTAL, border=B_TOT)

def pct_total(ws, r, c, v):
    wc(ws, r, c, v, font=F_PCT_T, fmt=PCT_FMT, border=B_TOT)

def span_top_border(ws, r, mc):
    """Extend medium top border across all cells in row to last active column."""
    for c in range(1, mc + 1):
        cell = ws.cell(row=r, column=c)
        old = cell.border
        cell.border = Border(top=Side(style='medium'),
                             left=old.left, right=old.right, bottom=old.bottom)

def apply_group_borders(ws, r1, r2, col_start, n_usn, n_cg):
    """Add thin left borders at USN/USCG/Total group boundaries for mekko tables."""
    tc = col_start + n_usn + n_cg  # Total column
    boundaries = [col_start]
    if n_cg > 0:
        boundaries.append(col_start + n_usn)
    boundaries.append(tc)
    for c in boundaries:
        for r in range(r1, r2 + 1):
            cell = ws.cell(row=r, column=c)
            old = cell.border
            cell.border = Border(left=Side(style='thin'),
                                 top=old.top, right=old.right, bottom=old.bottom)

def svc_group(types, svc_map):
    usn = [t for t in types if svc_map.get(t, '') in ('USN', 'MSC', '')]
    cg  = [t for t in types if svc_map.get(t, '') == 'USCG']
    return usn, cg

def svc_order(types, svc_map):
    """Return types ordered USN/MSC first, USCG second (no cell output)."""
    usn, cg = svc_group(types, svc_map)
    return usn + cg

def write_svc_header(ws, r, types, svc_map, col_start=2):
    """Write USN / USCG label row for mekko sections with sub-subsection fill.
    Returns (ordered_types, n_usn, n_cg)."""
    usn, cg = svc_group(types, svc_map)
    ordered = usn + cg
    # Fill the row with sub-subsection bg
    span = col_start + len(ordered)  # through Total column
    for c in range(1, span + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = BG_GRAY
        cell.alignment = _NOWRAP
    if usn: wc(ws, r, col_start, 'USN', font=F_SUBSUBSEC, fill=BG_GRAY)
    if cg:  wc(ws, r, col_start + len(usn), 'USCG', font=F_SUBSUBSEC, fill=BG_GRAY)
    return ordered, len(usn), len(cg)

def finish_sheet(ws):
    """Apply sheet-level formatting: gridlines off, universal row height."""
    ws.sheet_view.showGridLines = False
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = _RH


# ── Data reading ──────────────────────────────────────────────

def read_data(src_path, col_indices):
    """
    Pre-scan Sheet 1 to find which vessel types / buckets have data.
    col_indices: 0-indexed column positions to try (cascade order).
    """
    def read_val(row):
        for i in col_indices:
            v = row[i].value
            if v is not None and isinstance(v, (int, float)): return float(v)
        return 0.0

    def additive(rt):
        if rt is None or str(rt).strip() == '': return True
        return str(rt).strip() == '[PARENT]'

    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = wb['J Book Items Cons.']

    nb = defaultdict(float); mro = defaultdict(float)
    mro_bkt = defaultdict(lambda: defaultdict(float)); mro_b = defaultdict(float)
    type_svc = {}; type_cat = {}
    all_tam_types = set(); tf_src = defaultdict(float)
    nb_excl = defaultdict(float); mro_excl = defaultdict(float)

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        cat = str(row[24].value).strip() if row[24].value else ''
        vt  = str(row[25].value).strip() if row[25].value else ''
        sv  = str(row[23].value).strip() if row[23].value else ''

        if cat in TAM_CATEGORIES and vt:
            all_tam_types.add(vt)
            if sv: type_svc[vt] = sv
            type_cat[vt] = cat

        rt = row[21].value
        if not additive(rt): continue
        bk = row[4].value
        if bk is None: continue
        bs = str(bk).strip()
        if bs not in ('1', '2', '3', '4', '5', '6', '7'): continue

        v = read_val(row)
        src = str(row[0].value).strip() if row[0].value else '(blank)'
        tf_src[src] += v

        if bs == '1':
            if cat in TAM_CATEGORIES: nb[vt] += v
            elif cat: nb_excl[cat] += v
        elif bs in MRO_BKTS:
            if cat in TAM_CATEGORIES:
                mro[vt] += v; mro_bkt[vt][bs] += v; mro_b[bs] += v
            elif cat: mro_excl[cat] += v

    wb.close()

    def sorted_svc(types, val_map):
        usn = sorted([t for t in types if type_svc.get(t, '') in ('USN', 'MSC', '')],
                      key=lambda t: -val_map.get(t, 0))
        cg  = sorted([t for t in types if type_svc.get(t, '') == 'USCG'],
                      key=lambda t: -val_map.get(t, 0))
        return usn + cg

    all_sam_types = all_tam_types - SAM_EXCLUDED_TYPES
    nb_nz      = [t for t in sorted_svc(all_tam_types, nb)  if nb.get(t, 0) > 0]
    nb_sam_nz  = [t for t in nb_nz if t not in SAM_EXCLUDED_TYPES]
    mro_nz     = [t for t in sorted_svc(all_tam_types, mro) if mro.get(t, 0) > 0]
    mro_sam_nz = [t for t in mro_nz if t not in SAM_EXCLUDED_TYPES
                  and (mro_bkt[t].get('2', 0) + mro_bkt[t].get('4', 0)) > 0]

    return {
        'all_tam_types':       sorted_svc(all_tam_types, nb),
        'all_sam_types':       sorted_svc(all_sam_types, nb),
        'nb_nz': nb_nz, 'nb_sam_nz': nb_sam_nz,
        'mro_nz': mro_nz, 'mro_sam_nz': mro_sam_nz,
        'all_tam_mro_sorted':  sorted_svc(all_tam_types, mro),
        'all_sam_mro_sorted':  sorted_svc(all_sam_types, {
            t: mro_bkt[t].get('2', 0) + mro_bkt[t].get('4', 0) for t in all_sam_types
        }),
        'type_svc': type_svc,
        'mro_bkts_with_data':  [b for b in MRO_BKTS if mro_b.get(b, 0) > 0],
        'sources':             [s for s, _ in sorted(tf_src.items(), key=lambda x: -x[1])],
        'nb_excl':  [c for c, v in sorted(nb_excl.items(), key=lambda x: -x[1]) if v > 0],
        'mro_excl': [c for c, v in sorted(mro_excl.items(), key=lambda x: -x[1]) if v > 0],
    }


# ── Sheet builders ────────────────────────────────────────────

def create_total_funding(wb, d, label, prefix, cf, cf_neg, cf_inner):
    ws = wb.create_sheet(f'{prefix} Total Funding')
    mc = 3  # max active columns: A, B, C

    r = 1; title_band(ws, r, 'Total Funding', mc)
    r = 2; purpose_row(ws, r, f'All Navy + USCG appropriations for vessel-related activity, {label} ($K)')
    # Row 3: spacer

    # (A) By Work Type
    r = 4; subsec_band(ws, r, f'(A) {label} Total Funding by Work Type', 2)
    r = 5; hdr_row(ws, r, [('Work Type', 30), (f'{label} ($K)', 12)])
    for b in ['1', '2', '3', '4', '5', '6', '7']:
        r += 1; wc(ws, r, 1, BUCKET_NAMES[b], font=F_DATA)
        wc(ws, r, 2, cf(f'JB_B,{b}'), font=F_DATA, fmt=NUM_FMT)
    r += 1; total_label(ws, r, 1, 'Total'); total_cell(ws, r, 2, f'=SUM(B6:B{r-1})')
    span_top_border(ws, r, 2)

    # (B) By Vessel Category
    r += 3; subsec_band(ws, r, f'(B) {label} Total Funding by Vessel Category', 3)
    r += 1; hdr_row(ws, r, [('Category', 30), (f'{label} ($K)', 12), ('TAM Scope', 18)])
    cats = [
        ('Combatant Ships',           'YES \u2192 TAM'),
        ('Auxiliary Ships',            'YES \u2192 TAM'),
        ('Cutters',                    'YES \u2192 TAM'),
        ('Unmanned Maritime Platforms', 'YES \u2192 TAM'),
        ('Combatant Crafts',           'No \u2014 small craft'),
        ('Support Crafts',             'No \u2014 non-oceangoing'),
        ('Boats',                      'No \u2014 under 65 ft'),
    ]
    tam_rows = []; first = r + 1
    for cat, scope in cats:
        r += 1; wc(ws, r, 1, cat, font=F_DATA)
        wc(ws, r, 2, cf(f'JB_V,"{cat}"'), font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, 3, scope, font=F_GRAY)
        if 'YES' in scope: tam_rows.append(r)
    r += 1; wc(ws, r, 1, 'Unattributed (fleet-wide)', font=F_DATA)
    wc(ws, r, 2, cf('JB_V,""'), font=F_DATA, fmt=NUM_FMT)
    wc(ws, r, 3, 'No \u2014 not vessel-specific', font=F_GRAY)
    r += 1; total_label(ws, r, 1, 'Total'); total_cell(ws, r, 2, f'=SUM(B{first}:B{r-1})')
    span_top_border(ws, r, 3)
    r += 1; wc(ws, r, 1, 'TAM-eligible', font=F_KPI, fill=BG_TEAL)
    wc(ws, r, 2, '=' + '+'.join(f'B{tr}' for tr in tam_rows), font=F_KPI, fill=BG_TEAL, fmt=NUM_FMT)

    # (C) By Source Book
    r += 3; subsec_band(ws, r, f'(C) {label} Total Funding by Source Book', 2)
    r += 1; hdr_row(ws, r, [('Source Book', 30), (f'{label} ($K)', 12)]); fs = r + 1
    for s in d['sources']:
        r += 1; wc(ws, r, 1, s, font=F_DATA)
        wc(ws, r, 2, cf(f'JB_A,"{s}"'), font=F_DATA, fmt=NUM_FMT)
    r += 1; total_label(ws, r, 1, 'Total'); total_cell(ws, r, 2, f'=SUM(B{fs}:B{r-1})')
    span_top_border(ws, r, 2)

    finish_sheet(ws)


def create_newbuild_tam(wb, d, label, prefix, cf, cf_neg, cf_inner):
    ws = wb.create_sheet(f'{prefix} Newbuild TAM')
    mc = max(4, 2 + len(d['nb_nz']) + 1)
    cw(ws, 1, 30); cw(ws, 2, 12)

    r = 1; title_band(ws, r, f'Newbuild TAM \u2014 {label} ($K)', mc)
    r = 2; purpose_row(ws, r, 'New Construction scoped to Combatant Ships, Auxiliary Ships, USCG Cutters & Unmanned Platforms')

    # (A) Bridge
    r = 4; subsec_band(ws, r, '(A) Bridge: Total NC Funding \u2192 Newbuild TAM', 2)
    r = 5; wc(ws, r, 1, 'Total New Construction Funding', font=F_DATA)
    wc(ws, r, 2, cf('JB_B,1'), font=F_DATA, fmt=NUM_FMT)
    r += 1; wc(ws, r, 1, '  Less: Unattributed (fleet-wide NC items)', font=F_DATA)
    wc(ws, r, 2, cf_neg('JB_B,1', 'JB_V,""'), font=F_DATA, fmt=NUM_FMT)
    for cat in TAM_EXCLUDED_CATEGORIES:
        r += 1; wc(ws, r, 1, f'  Less: {cat}', font=F_DATA)
        wc(ws, r, 2, cf_neg('JB_B,1', f'JB_V,"{cat}"'), font=F_DATA, fmt=NUM_FMT)
    r += 1; total_label(ws, r, 1, 'Newbuild TAM')
    tam_f = '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES))
    total_cell(ws, r, 2, '=' + tam_f)
    span_top_border(ws, r, 2)

    # (B) All vessel types
    r += 3; subsec_band(ws, r, '(B) Newbuild TAM by Vessel Type', 4)
    all_types = svc_order(d['all_tam_types'], d['type_svc'])
    r += 1; hdr_row(ws, r, [('Vessel Type', 30), ('Service', 8), (f'{label} ($K)', 12), ('% of TAM', 10)])
    ft = r + 1
    for vt in all_types:
        r += 1; wc(ws, r, 1, vt, font=F_DATA)
        wc(ws, r, 2, d['type_svc'].get(vt, ''), font=F_DATA)
        wc(ws, r, 3, cf('JB_B,1', f'JB_W,"{vt}"'), font=F_DATA, fmt=NUM_FMT)
    lt = r; r += 1; tr = r
    total_label(ws, r, 1, 'Newbuild TAM'); total_label(ws, r, 2, '')
    total_cell(ws, r, 3, f'=SUM(C{ft}:C{lt})')
    pct_total(ws, r, 4, 1.0)
    span_top_border(ws, r, 4)
    for rn in range(ft, lt + 1):
        wc(ws, rn, 4, f'=IFERROR(C{rn}/C${tr},0)', font=F_PCT, fmt=PCT_FMT)

    # (C) Mekko
    mk = svc_order(d['nb_nz'], d['type_svc'])
    r += 3; subsec_band(ws, r, '(C) Newbuild TAM \u2014 DD&C vs AP/LLTM by Vessel Type (Mekko)', 2 + len(mk))
    r += 1; svc_r = r; mk, n_usn, n_cg = write_svc_header(ws, r, d['nb_nz'], d['type_svc'])
    r += 1; wc(ws, r, 1, 'Sub-Category', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
    for i, vt in enumerate(mk):
        c = 2 + i; wc(ws, r, c, vt, font=F_HDR, border=B_HDR); cw(ws, c, 12)
    tc = 2 + len(mk); wc(ws, r, tc, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc, 12)
    dr = r + 1; ar = r + 2; orr = r + 3; tkr = r + 4

    wc(ws, tkr, 1, 'Total ($K)', font=F_TOTAL, border=B_TOT)
    for i, vt in enumerate(mk):
        total_cell(ws, tkr, 2 + i, cf('JB_B,1', f'JB_W,"{vt}"'))
    total_cell(ws, tkr, tc, f'=SUM({get_column_letter(2)}{tkr}:{get_column_letter(tc-1)}{tkr})')

    ddc = ['"Full ship DD&C"', '"Full ship DD&C / LLTM"', '"Full vessel procurement"']
    wc(ws, dr, 1, 'Full Ship DD&C', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        parts = '+'.join(cf_inner('JB_B,1', f'JB_W,"{vt}"', f'JB_F,{s}') for s in ddc)
        wc(ws, dr, c, f'=IFERROR(({parts})/{cl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)
    tcl = get_column_letter(tc)
    wc(ws, dr, tc, f'=IFERROR(1-{tcl}{ar},0)', font=F_PCT, fmt=PCT_FMT)

    wc(ws, ar, 1, 'Advance Procurement / LLTM', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        ap = cf_inner('JB_B,1', f'JB_W,"{vt}"', 'JB_F,"Advance procurement / LLTM"')
        wc(ws, ar, c, f'=IFERROR(({ap})/{cl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)
    ap_cat = '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"', 'JB_F,"Advance procurement / LLTM"') for c in sorted(TAM_CATEGORIES))
    wc(ws, ar, tc, f'=IFERROR(({ap_cat})/{tcl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)

    wc(ws, orr, 1, 'Other NC Sub-Categories', font=F_DATA)
    for i in range(len(mk)):
        c = 2 + i; cl = get_column_letter(c)
        wc(ws, orr, c, f'=IFERROR(1-{cl}{dr}-{cl}{ar},0)', font=F_PCT, fmt=PCT_FMT)
    wc(ws, orr, tc, f'=IFERROR(1-{tcl}{dr}-{tcl}{ar},0)', font=F_PCT, fmt=PCT_FMT)

    span_top_border(ws, tkr, tc)
    apply_group_borders(ws, svc_r, tkr, 2, n_usn, n_cg)

    finish_sheet(ws)


def create_newbuild_sam(wb, d, label, prefix, cf, cf_neg, cf_inner):
    ws = wb.create_sheet(f'{prefix} Newbuild SAM')
    mc = max(4, 2 + len(d['nb_sam_nz']) + 1)
    cw(ws, 1, 30); cw(ws, 2, 12)

    r = 1; title_band(ws, r, f'Newbuild SAM \u2014 {label} ($K)', mc)
    r = 2; purpose_row(ws, r, 'Newbuild TAM excluding vessel types not addressable by PA')

    # (A) Bridge
    r = 4; subsec_band(ws, r, '(A) Bridge: Newbuild TAM \u2192 Newbuild SAM', 2)
    r = 5; wc(ws, r, 1, 'Newbuild TAM', font=F_DATA)
    wc(ws, r, 2, '=' + '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES)), font=F_DATA, fmt=NUM_FMT)
    for vt in ['Submarines', 'Aircraft Carriers', 'Unmanned Undersea Vehicles']:
        r += 1; wc(ws, r, 1, f'  Less: {vt}', font=F_DATA)
        wc(ws, r, 2, cf_neg('JB_B,1', f'JB_W,"{vt}"'), font=F_DATA, fmt=NUM_FMT)
    r += 1; total_label(ws, r, 1, 'Newbuild SAM')
    total_cell(ws, r, 2, f'=SUM(B5:B{r-1})')
    span_top_border(ws, r, 2)

    # (B) All SAM vessel types
    r += 3; subsec_band(ws, r, '(B) Newbuild SAM by Vessel Type', 4)
    all_types = svc_order(d['all_sam_types'], d['type_svc'])
    r += 1; hdr_row(ws, r, [('Vessel Type', 30), ('Service', 8), (f'{label} ($K)', 12), ('% of SAM', 10)])
    ft = r + 1
    for vt in all_types:
        r += 1; wc(ws, r, 1, vt, font=F_DATA)
        wc(ws, r, 2, d['type_svc'].get(vt, ''), font=F_DATA)
        wc(ws, r, 3, cf('JB_B,1', f'JB_W,"{vt}"'), font=F_DATA, fmt=NUM_FMT)
    lt = r; r += 1; tr = r
    total_label(ws, r, 1, 'Newbuild SAM'); total_label(ws, r, 2, '')
    total_cell(ws, r, 3, f'=SUM(C{ft}:C{lt})')
    pct_total(ws, r, 4, 1.0)
    span_top_border(ws, r, 4)
    for rn in range(ft, lt + 1):
        wc(ws, rn, 4, f'=IFERROR(C{rn}/C${tr},0)', font=F_PCT, fmt=PCT_FMT)

    # (C) Mekko
    mk = svc_order(d['nb_sam_nz'], d['type_svc'])
    r += 3; subsec_band(ws, r, '(C) Newbuild SAM \u2014 DD&C vs AP/LLTM by Vessel Type (Mekko)', 2 + len(mk))
    r += 1; svc_r = r; mk, n_usn, n_cg = write_svc_header(ws, r, d['nb_sam_nz'], d['type_svc'])
    r += 1; wc(ws, r, 1, 'Sub-Category', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
    for i, vt in enumerate(mk):
        c = 2 + i; wc(ws, r, c, vt, font=F_HDR, border=B_HDR); cw(ws, c, 12)
    tc = 2 + len(mk); wc(ws, r, tc, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc, 12)
    dr = r + 1; ar = r + 2; tkr = r + 3

    wc(ws, tkr, 1, 'Total ($K)', font=F_TOTAL, border=B_TOT)
    for i, vt in enumerate(mk):
        total_cell(ws, tkr, 2 + i, cf('JB_B,1', f'JB_W,"{vt}"'))
    total_cell(ws, tkr, tc, f'=SUM({get_column_letter(2)}{tkr}:{get_column_letter(tc-1)}{tkr})')

    ddc = ['"Full ship DD&C"', '"Full ship DD&C / LLTM"', '"Full vessel procurement"']
    wc(ws, dr, 1, 'Full Ship DD&C', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        parts = '+'.join(cf_inner('JB_B,1', f'JB_W,"{vt}"', f'JB_F,{s}') for s in ddc)
        wc(ws, dr, c, f'=IFERROR(({parts})/{cl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)
    tcl = get_column_letter(tc)
    wc(ws, dr, tc, f'=IFERROR(1-{tcl}{ar},0)', font=F_PCT, fmt=PCT_FMT)

    wc(ws, ar, 1, 'Advance Procurement / LLTM', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        ap = cf_inner('JB_B,1', f'JB_W,"{vt}"', 'JB_F,"Advance procurement / LLTM"')
        wc(ws, ar, c, f'=IFERROR(({ap})/{cl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)
    ap_cat = '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"', 'JB_F,"Advance procurement / LLTM"') for c in sorted(TAM_CATEGORIES))
    wc(ws, ar, tc, f'=IFERROR(({ap_cat})/{tcl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)

    span_top_border(ws, tkr, tc)
    apply_group_borders(ws, svc_r, tkr, 2, n_usn, n_cg)

    finish_sheet(ws)


def create_mro_tam(wb, d, label, prefix, cf, cf_neg, cf_inner):
    ws = wb.create_sheet(f'{prefix} MRO TAM')
    mc = max(4, 2 + len(d['mro_nz']) + 1)
    cw(ws, 1, 30); cw(ws, 2, 12)

    bkts = d['mro_bkts_with_data']

    r = 1; title_band(ws, r, f'MRO TAM \u2014 {label} ($K)', mc)
    r = 2; purpose_row(ws, r, 'MRO & sustainment scoped to Combatant Ships, Auxiliary Ships, USCG Cutters & Unmanned Platforms')

    # (A) Bridge
    r = 4; subsec_band(ws, r, '(A) Bridge: Total MRO Funding \u2192 MRO TAM', 2)
    r = 5; wc(ws, r, 1, 'Total MRO Funding (Buckets 2\u20137)', font=F_DATA)
    wc(ws, r, 2, '=' + '+'.join(cf_inner(f'JB_B,{b}') for b in MRO_BKTS), font=F_DATA, fmt=NUM_FMT)
    r += 1; wc(ws, r, 1, '  Less: Unattributed (fleet-wide MRO items)', font=F_DATA)
    wc(ws, r, 2, f'=-({"+".join(cf_inner(f"JB_B,{b}", "JB_V," + chr(34) + chr(34)) for b in MRO_BKTS)})',
       font=F_DATA, fmt=NUM_FMT)
    for cat in TAM_EXCLUDED_CATEGORIES:
        r += 1; wc(ws, r, 1, f'  Less: {cat}', font=F_DATA)
        wc(ws, r, 2, f'=-({"+".join(cf_inner(f"JB_B,{b}", f"""JB_V,"{cat}" """.strip()) for b in MRO_BKTS)})',
           font=F_DATA, fmt=NUM_FMT)
    r += 1; bridge_tam_row = r
    total_label(ws, r, 1, 'MRO TAM')

    # (B) By Work Type
    r += 3; subsec_band(ws, r, '(B) MRO TAM by Work Type', 3)
    r += 1; hdr_row(ws, r, [('Work Type', 30), (f'{label} ($K)', 12), ('% of TAM', 10)])
    fb = r + 1
    for b in bkts:
        r += 1; wc(ws, r, 1, BUCKET_SHORT.get(b, BUCKET_NAMES[b]), font=F_DATA)
        parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES))
        wc(ws, r, 2, '=' + parts, font=F_DATA, fmt=NUM_FMT)
    lb = r; r += 1; btr = r
    total_label(ws, r, 1, 'MRO TAM')
    total_cell(ws, r, 2, f'=SUM(B{fb}:B{lb})')
    pct_total(ws, r, 3, 1.0)
    span_top_border(ws, r, 3)
    for rn in range(fb, lb + 1):
        wc(ws, rn, 3, f'=IFERROR(B{rn}/B${btr},0)', font=F_PCT, fmt=PCT_FMT)

    # Fill bridge total
    total_cell(ws, bridge_tam_row, 2, f'=B{btr}')
    span_top_border(ws, bridge_tam_row, 2)

    # (C) All vessel types
    r += 3; subsec_band(ws, r, '(C) MRO TAM by Vessel Type', 4)
    all_types = svc_order(d['all_tam_mro_sorted'], d['type_svc'])
    r += 1; hdr_row(ws, r, [('Vessel Type', 30), ('Service', 8), (f'{label} ($K)', 12), ('% of TAM', 10)])
    ft = r + 1
    for vt in all_types:
        r += 1; wc(ws, r, 1, vt, font=F_DATA)
        wc(ws, r, 2, d['type_svc'].get(vt, ''), font=F_DATA)
        mparts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_W,"{vt}"') for b in MRO_BKTS)
        wc(ws, r, 3, '=' + mparts, font=F_DATA, fmt=NUM_FMT)
    lt = r; r += 1; ttr = r
    total_label(ws, r, 1, 'MRO TAM'); total_label(ws, r, 2, '')
    total_cell(ws, r, 3, f'=SUM(C{ft}:C{lt})')
    pct_total(ws, r, 4, 1.0)
    span_top_border(ws, r, 4)
    for rn in range(ft, lt + 1):
        wc(ws, rn, 4, f'=IFERROR(C{rn}/C${ttr},0)', font=F_PCT, fmt=PCT_FMT)

    # (D) Mekko
    mk = svc_order(d['mro_nz'], d['type_svc'])
    r += 3; subsec_band(ws, r, '(D) MRO TAM \u2014 Work Type by Vessel Type (Mekko)', 2 + len(mk))
    r += 1; svc_r = r; mk, n_usn, n_cg = write_svc_header(ws, r, d['mro_nz'], d['type_svc'], col_start=2)
    r += 1; wc(ws, r, 1, 'Work Type', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
    for i, vt in enumerate(mk):
        c = 2 + i; wc(ws, r, c, vt, font=F_HDR, border=B_HDR); cw(ws, c, 12)
    tc = 2 + len(mk); wc(ws, r, tc, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc, 12)
    nwt = len(bkts); tkr = r + 1 + nwt

    wc(ws, tkr, 1, 'Total ($K)', font=F_TOTAL, border=B_TOT)
    for i, vt in enumerate(mk):
        c = 2 + i; mparts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_W,"{vt}"') for b in MRO_BKTS)
        total_cell(ws, tkr, c, '=' + mparts)
    total_cell(ws, tkr, tc, f'=SUM({get_column_letter(2)}{tkr}:{get_column_letter(tc-1)}{tkr})')

    tcl = get_column_letter(tc)
    for bi, b in enumerate(bkts):
        wr = r + 1 + bi
        wc(ws, wr, 1, BUCKET_SHORT.get(b, BUCKET_NAMES[b]), font=F_DATA)
        for i, vt in enumerate(mk):
            c = 2 + i; cl = get_column_letter(c)
            wc(ws, wr, c, f'=IFERROR(({cf_inner(f"JB_B,{b}", f"""JB_W,"{vt}" """.strip())})/{cl}${tkr},0)',
               font=F_PCT, fmt=PCT_FMT)
        bkt_parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES))
        wc(ws, wr, tc, f'=IFERROR(({bkt_parts})/{tcl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)

    span_top_border(ws, tkr, tc)
    apply_group_borders(ws, svc_r, tkr, 2, n_usn, n_cg)

    finish_sheet(ws)
    return {'bkt_total_row': btr}


def create_mro_sam(wb, d, label, prefix, cf, cf_neg, cf_inner, mro_tam_info=None):
    ws = wb.create_sheet(f'{prefix} MRO SAM')
    mc = max(6, 2 + len(d['mro_sam_nz']) + 1)
    cw(ws, 1, 30); cw(ws, 2, 12)

    tam_sheet = f'{prefix} MRO TAM'

    r = 1; title_band(ws, r, f'MRO SAM \u2014 {label} ($K)', mc)
    r = 2; purpose_row(ws, r, 'Outsourceable MRO: SDM + Modernization, excluding non-addressable vessel types')

    # (A) Bridge
    r = 4; subsec_band(ws, r, '(A) Bridge: MRO TAM \u2192 MRO SAM', 2)
    r = 5; wc(ws, r, 1, 'MRO TAM', font=F_DATA)
    tam_ref_row = mro_tam_info['bkt_total_row'] if mro_tam_info else 18
    wc(ws, r, 2, f"='{tam_sheet}'!B{tam_ref_row}", font=F_GREEN, fmt=NUM_FMT)

    excl_rows = []
    for vt in ['Submarines', 'Aircraft Carriers', 'Unmanned Undersea Vehicles']:
        r += 1; excl_rows.append(r)
        wc(ws, r, 1, f'  Less: {vt}', font=F_DATA)
        parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_W,"{vt}"') for b in MRO_BKTS)
        wc(ws, r, 2, f'=-({parts})', font=F_DATA, fmt=NUM_FMT)

    r += 1; nowork_row = r
    wc(ws, r, 1, '  Less: Non-outsourceable work types', font=F_DATA)

    r += 1; sam_result_row = r
    total_label(ws, r, 1, 'MRO SAM')
    sam_parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_V,"{c}"') for b in ['2', '4'] for c in sorted(TAM_CATEGORIES))
    excl_parts = '+'.join(
        f'-({cf_inner(f"JB_B,{b}", f"""JB_W,"{vt}" """.strip())})'
        for b in ['2', '4'] for vt in ['Submarines', 'Aircraft Carriers', 'Unmanned Undersea Vehicles']
    )
    total_cell(ws, r, 2, f'={sam_parts}{excl_parts}')
    span_top_border(ws, r, 2)

    excl_sum = '+'.join(f'B{er}' for er in excl_rows)
    wc(ws, nowork_row, 2, f'=-(B5+{excl_sum}-B{sam_result_row})', font=F_DATA, fmt=NUM_FMT)

    # (B) All SAM vessel types
    r += 3; subsec_band(ws, r, '(B) MRO SAM by Vessel Type', 6)
    all_types = svc_order(d['all_sam_mro_sorted'], d['type_svc'])
    r += 1; hdr_row(ws, r, [('Vessel Type', 30), ('Service', 8), ('SDM ($K)', 12), ('Mod ($K)', 12), ('Total ($K)', 12), ('% of SAM', 10)])
    ft = r + 1
    for vt in all_types:
        r += 1; wc(ws, r, 1, vt, font=F_DATA)
        wc(ws, r, 2, d['type_svc'].get(vt, ''), font=F_DATA)
        wc(ws, r, 3, cf('JB_B,2', f'JB_W,"{vt}"'), font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, 4, cf('JB_B,4', f'JB_W,"{vt}"'), font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, 5, f'=C{r}+D{r}', font=F_DATA, fmt=NUM_FMT)
    lt = r; r += 1; tr = r
    total_label(ws, r, 1, 'MRO SAM'); total_label(ws, r, 2, '')
    for c in [3, 4, 5]:
        total_cell(ws, r, c, f'=SUM({get_column_letter(c)}{ft}:{get_column_letter(c)}{lt})')
    pct_total(ws, r, 6, 1.0)
    span_top_border(ws, r, 6)
    for rn in range(ft, lt + 1):
        wc(ws, rn, 6, f'=IFERROR(E{rn}/E${tr},0)', font=F_PCT, fmt=PCT_FMT)

    # (C) Mekko
    mk = svc_order(d['mro_sam_nz'], d['type_svc'])
    r += 3; subsec_band(ws, r, '(C) MRO SAM \u2014 SDM vs Modernization by Vessel Type (Mekko)', 2 + len(mk))
    r += 1; svc_r = r; mk, n_usn, n_cg = write_svc_header(ws, r, d['mro_sam_nz'], d['type_svc'], col_start=2)
    r += 1; wc(ws, r, 1, 'Work Type', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
    for i, vt in enumerate(mk):
        c = 2 + i; wc(ws, r, c, vt, font=F_HDR, border=B_HDR); cw(ws, c, 12)
    tc = 2 + len(mk); wc(ws, r, tc, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc, 12)
    sr = r + 1; mr = r + 2; tkr = r + 3

    wc(ws, tkr, 1, 'Total ($K)', font=F_TOTAL, border=B_TOT)
    for i, vt in enumerate(mk):
        c = 2 + i
        parts = cf_inner('JB_B,2', f'JB_W,"{vt}"') + '+' + cf_inner('JB_B,4', f'JB_W,"{vt}"')
        total_cell(ws, tkr, c, '=' + parts)
    total_cell(ws, tkr, tc, f'=SUM({get_column_letter(2)}{tkr}:{get_column_letter(tc-1)}{tkr})')

    tcl = get_column_letter(tc)
    wc(ws, sr, 1, 'Scheduled Depot Maintenance', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        wc(ws, sr, c, f'=IFERROR(({cf_inner("JB_B,2", f"""JB_W,"{vt}" """.strip())})/{cl}${tkr},0)',
           font=F_PCT, fmt=PCT_FMT)
    wc(ws, sr, tc, f'=IFERROR(1-{tcl}{mr},0)', font=F_PCT, fmt=PCT_FMT)

    wc(ws, mr, 1, 'Modernization', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        wc(ws, mr, c, f'=IFERROR(({cf_inner("JB_B,4", f"""JB_W,"{vt}" """.strip())})/{cl}${tkr},0)',
           font=F_PCT, fmt=PCT_FMT)
    mod_parts = '+'.join(cf_inner('JB_B,4', f'JB_W,"{vt}"') for vt in mk)
    wc(ws, mr, tc, f'=IFERROR(({mod_parts})/{tcl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)

    span_top_border(ws, tkr, tc)
    apply_group_borders(ws, svc_r, tkr, 2, n_usn, n_cg)

    finish_sheet(ws)


# ── Competitive Dynamics ──────────────────────────────────────

# Data extracted from analysis/FY2026_Key_Programs_Contract_Awards.md
# Values in $M unless noted otherwise.

_PRIMES = [
    (1, 'Huntington Ingalls Industries', 'HII', 22000000, 'DDG-51 $11.4B, LPD $4.7B, LHA $4.5B, DDG-1000 $0.3B, depot'),
    (2, 'Raytheon / RTX', 'RTX Corp', 13700000, 'Standard Missile $5.7B, SPY-6 $3.5B, CIWS/RAM $2.0B, DDG-1000 $0.9B, CEC $0.6B'),
    (3, 'Bath Iron Works', 'General Dynamics', 8800000, 'DDG-51 FY23-27 $5.0B, FY18-22 $3.3B, FY13-17 $0.5B, DDG-1000 $0.2B'),
    (4, 'BAE Systems', 'BAE Systems plc', 5000000, 'DDG/LHD/LPD/LSD/CG depot maintenance — Norfolk, San Diego, Jacksonville, Hawaii'),
    (5, 'Lockheed Martin', 'LM', 1500000, 'SLQ-32(V)6 $0.8B, SSDS $0.4B, HELIOS $0.1B, LCS Freedom $0.3B'),
    (6, 'Metro Machine Corp', '--', 700000, 'LHD/DDG/LSD depot maintenance — Norfolk'),
    (7, 'NASSCO', 'General Dynamics', 600000, 'LHA/LHD/LSD/DDG/CG depot maintenance — San Diego'),
    (8, 'Vigor Marine LLC', '--', 500000, 'DDG/CG depot maintenance — Portland, OR'),
    (9, 'Northrop Grumman Systems', 'NG', 487000, 'Knifefish/LCS $0.35B, SEWIP Block 3 $0.14B, ALMDS $0.07B'),
    (10, 'QED Systems', '--', 300000, 'Third-party planning DDG/CG/LPD/LHD'),
    (11, 'Continental Maritime', '--', 300000, 'DDG depot — San Diego'),
    (12, 'Textron Systems', 'Textron', 242000, 'UISS $0.24B'),
    (13, 'Advanced Technology International', '--', 235000, 'OTA consortium: MK-54 MOD 2, SPY-6 prototype, HEL, LHD midlife'),
    (14, 'General Dynamics Mission Systems', 'General Dynamics', 210000, 'MK-54 MOD 1 $0.12B, Knifefish Block 1 $0.08B'),
    (15, 'SAIC', 'SAIC', 200000, 'DDG-1000 engineering, SSTD Nixie, torpedo defense'),
    (16, 'Marine Hydraulics International', '--', 200000, 'DDG/LSD depot + sub-prime on LHD primes'),
    (17, 'Bollinger Shipyards', 'Bollinger', 116000, 'MCM USV production'),
    (18, 'Ultra Electronics Ocean Systems', 'Ultra', 68000, 'MK 54 MOD 0 array kits + Nixie production'),
    (19, 'Penn State Applied Research Lab', 'Penn State', 60000, 'MK-54 MOD 2 engineering, HEL development'),
]

_SUBS = [
    (1, 'Fincantieri (Marinette Marine)', 2487000, 2, 1, 'LCS Freedom hull builder — actual constructor under LM prime'),
    (2, 'Timken Gears & Services', 516000, 22, 20, 'DDG/LHD/CG propulsion gearing'),
    (3, 'General Dynamics (rolled up)', 432000, 23, 14, 'GD Mission Systems $359M on SPY-6, GD-OTS, GDIT'),
    (4, 'L3Harris (Aerojet Rocketdyne)', 378000, 4, 4, 'Solid rocket motors for Standard Missile family'),
    (5, 'Rolls-Royce', 313000, 25, 25, 'Gas turbines, waterjets, gearing — DDG/LCS/LPD/LHA + depot'),
    (6, 'CAES (Cobham)', 217000, 10, 8, 'RF assemblies on SPY-6 $170M, SEWIP $110M, SLQ-32 $95M'),
    (7, 'Northrop Grumman (as sub)', 215000, 10, 9, 'SPY-6 RF $114M, CIWS, LPD/LHA construction'),
    (8, 'IMIA', 204000, 28, 28, 'Hull coatings / paint across 28 depot availabilities'),
    (9, 'L3Harris (excl. Aerojet)', 203000, 53, 45, 'Maritime Power, Cincinnati Electronics, Fuzing — broadest footprint'),
    (10, 'Johnson Controls (incl. York)', 172000, 31, 24, 'HVAC / chilled water on DDG/LPD/LHA + depot'),
    (11, 'GE Aerospace', 166000, 16, 14, 'LM2500 gas turbines on DDG-51 + GE Power Conversion LHA 9'),
    (12, 'Fairbanks Morse Defense', 161000, 25, 18, 'Diesel generators / propulsion on LPD/LHA + depot rebuilds'),
    (13, 'Leonardo DRS', 134000, 18, 15, 'Naval Power Systems, Network & Imaging, Laurel Technologies'),
    (14, 'Sparton DeLeon Springs', 108000, 2, 2, 'ASW sensors / payload on Raytheon Barracuda'),
    (15, 'RAM-System GmbH', 98000, 1, 1, 'RAM missile co-development with Raytheon (Germany)'),
    (16, 'Bay Metals & Fabrication', 93000, 9, 9, 'Norfolk metal fab serving 9 depot availabilities'),
    (17, 'US Joiner LLC', 88000, 12, 12, 'Interior outfitting — LPD construction + depot'),
    (18, 'United Rentals', 77000, 29, 29, 'Equipment rental across 29 depot availabilities'),
    (19, 'Honeywell', 71000, 12, 12, 'IMUs / IRUs on Standard Missile + CIWS'),
    (20, 'Anaren', 64000, 7, 7, 'RF combiners/dividers on SPY-6'),
]

_PROGRAMS = [
    ('DDG-51 New Construction', 530000, 'GE Aerospace $88M', 'Johnson Controls $76M', 'Timken Gears $41M'),
    ('DDG Mod — SPY-6', 1220000, 'GD Mission Systems $359M', 'CAES $170M', 'Northrop Grumman $114M'),
    ('DDG Mod — CIWS/RAM/SeaRAM', 560000, 'RAM-System GmbH $98M', 'L3Harris $39M', 'GD $35M'),
    ('LCS Freedom Construction', 2910000, 'Marinette Marine $2,487M', 'Rolls-Royce $196M', 'Airbus US $56M'),
    ('LCS MCM Mission Modules', 1230000, 'Trident Sensors $215M', 'Teledyne $55M', 'GD $13M'),
    ('LHA New Construction', 340000, 'Fairbanks Morse $40M', 'L3Harris $39M', 'Leonardo DRS $34M'),
    ('LPD Flight II New Construction', 340000, 'Fairbanks Morse $65M', 'US Joiner $36M', 'Caterpillar $20M'),
    ('Standard Missile Family', 930000, 'L3Harris (Aerojet) $390M', 'Honeywell $43M', 'Goodrich $32M'),
    ('DDG-1000', 130000, 'Red River Technology $61M', 'Air Masters $6M', '--'),
    ('DDG Mod — CEC', 30000, 'Sechan Electronics $7M', 'Action Electronics $5M', '--'),
    ('DDG Mod — SSDS', 40000, 'Mission Solutions $14M', 'PMAT $6M', '--'),
    ('DDG Mod — HELIOS', 60000, 'MZA Associates $25M', 'L3 Technologies $22M', '--'),
    ('Airborne MCM', 150000, 'Sparton $108M', 'Teledyne $17M', '--'),
    ('MK-54 Torpedo', 30000, 'L3Harris (Aerojet) $25M', 'J&E Precision Tool $3M', '--'),
    ('Depot Maintenance', 1500000, 'IMIA $204M', 'Bay Metals $93M', 'US Joiner $88M'),
]

_CROSSWALK = {
    'LHD': [
        ('OMN_Vol2 LHD/AMPHIBS depot', 346443, 'BAE Norfolk/SD, Metro Machine, NASSCO'),
        ('OMN_Vol2 LHD/AMPHIBS modernization', 631859, 'BAE Norfolk/SD, Metro Machine'),
        ('OPN BA1 Line 8 LHA/LHD Midlife', 123384, 'BAE / Metro / Gibbs & Cox'),
    ],
    'LPD': [
        ('SCN 3010 LPD Flight II procurement', 835037, 'HII — N0002424C2473 ($5.80B ceiling)'),
        ('SCN 3010 LPD Flight II AP', 275000, 'HII — N0002424C2473 AP scope'),
        ('OMN_Vol2 LPD depot', 149595, 'BAE Norfolk/SD'),
        ('OPN BA1 Line 15 LPD Class Support', 125542, 'HII Class Eng ($225M) + Raytheon LCE&S ($485M)'),
    ],
    'LSD': [
        ('OMN_Vol2 LSD depot', 94421, 'NASSCO, BAE, Metro Machine'),
        ('OMN_Vol2 LSD modernization', 25931, 'Alteration installation (small)'),
    ],
    'LHA': [
        ('SCN 3041 LHA Replacement', 634963, 'HII — N0002420C2437 LHA 9 ($3.14B ceiling)'),
        ('OMN_Vol2 LHA depot', 45012, 'NASSCO — LHA 6 DSRA ($198M ceiling)'),
    ],
    'DDG': [
        ('SCN 2122 DDG-51 procurement', 5410773, 'HII FY23-27 MYP + BIW FY23-27 MYP'),
        ('SCN 2122 DDG-51 AP', 1750000, 'HII + BIW — AP against FY23-27 MYPs'),
        ('SCN 2119 DDG-1000', 52358, 'HII Mod Planning + BIW Planning Yard'),
        ('OMN_Vol2 DDG depot', 952000, 'BAE / NASSCO / Continental / Vigor / MHI / Metro'),
        ('OMN_Vol2 DDG modernization', 1282612, 'Cross-vehicle alteration installation'),
        ('WPN 2234 Standard Missile', 1008875, 'Raytheon SM production vehicles'),
        ('OPN BA1 Line 5 DDG Mod', 878787, 'SPY-6, SEWIP, SSDS, CEC, CIWS, HELIOS, Aegis kits'),
        ('WPN 3215 MK-54 Torpedo Mods', 128513, 'GDMS MK 54 MOD 1 + Ultra MK 54 MOD 0 + ATI OTAs'),
        ('OPN BA1 Line 16 DDG-1000 Support', 115267, 'Raytheon DDG-1000 mission sys + BIW Planning Yard'),
    ],
    'LCS': [
        ('OMN_Vol2 LCS depot', 576389, 'LM Freedom ISEA + GDMS Independence ISEA + avails'),
        ('OMN_Vol2 LCS modernization', 383361, 'LM + GDMS combat system ISEA + ITT-2A'),
        ('OPN BA1 Line 37 LCS In-Service Mod', 189458, 'LM, GDMS, Textron, BAH PEO LCS support'),
        ('OPN BA1 Line 34 LCS MCM Modules', 91372, 'UISS (Textron), Knifefish (GDMS/NG), MCM USV (Bollinger)'),
        ('WPN 2292 Naval Strike Missile', 32238, 'Raytheon NSM integration + Kongsberg OEM via FMS'),
    ],
    'CG': [
        ('OMN_Vol2 CG depot', 115501, 'Vigor, BAE Norfolk, NASSCO, HII CG planning'),
        ('OMN_Vol2 CG modernization', 44940, 'Same vendor pool — alteration installation'),
    ],
}


def create_competitive_dynamics(wb):
    ws = wb.create_sheet('Competitive Dynamics')
    ws.sheet_properties.tabColor = '7B5B3A'  # muted brown — distinct from FY26 blue / FY27 green

    # Column widths
    cw(ws, 1, 30)  # A: Contractor / Program / SAM Line
    cw(ws, 2, 16)  # B: Parent / Total Subs / FY26 $K
    cw(ws, 3, 12)  # C: Window $M / Sub Total / Top Sub 1
    cw(ws, 4, 26)  # D: Key Programs / # Primes / Top Sub 2 / Vehicle
    cw(ws, 5, 26)  # E: # PIIDs / Top Sub 3

    r = 1; title_band(ws, r, 'Competitive Dynamics', 5)
    r = 2; purpose_row(ws, r, 'FY20-26 contract obligation analysis — primes, subcontractors, and FY26 vehicle crosswalk ($K)')

    # ── Section A: Top Prime Contractors ──
    r = 4; subsec_band(ws, r, '(A) Top Prime Contractors — FY20-26 Window Obligation', 4)
    r += 1; hdr_row(ws, r, [('Contractor', 30), ('Parent', 16), ('Window ($K)', 12), ('Key Programs', 44)])
    ft = r + 1
    for rank, name, parent, val, programs in _PRIMES:
        r += 1
        wc(ws, r, 1, name, font=F_DATA)
        wc(ws, r, 2, parent, font=F_GRAY)
        wc(ws, r, 3, val, font=F_BLUE, fmt=NUM_FMT)
        wc(ws, r, 4, programs, font=F_GRAY)
    r += 1; total_label(ws, r, 1, 'Total (top 19)')
    total_cell(ws, r, 3, f'=SUM(C{ft}:C{r-1})')
    span_top_border(ws, r, 4)

    # ── Section B: Top Subcontractors ──
    r += 3; subsec_band(ws, r, '(B) Top Subcontractors — Hidden Supply Chain', 5)
    r += 1; hdr_row(ws, r, [('Parent Company', 30), ('Sub Total ($K)', 16), ('# Primes', 12), ('# PIIDs', 10), ('Primary Programs', 44)])
    ft = r + 1
    for rank, name, val, pairs, piids, programs in _SUBS:
        r += 1
        wc(ws, r, 1, name, font=F_DATA)
        wc(ws, r, 2, val, font=F_BLUE, fmt=NUM_FMT)
        wc(ws, r, 3, pairs, font=F_BLUE, fmt=NUM_FMT)
        wc(ws, r, 4, piids, font=F_BLUE, fmt=NUM_FMT)
        wc(ws, r, 5, programs, font=F_GRAY)
    r += 1; total_label(ws, r, 1, 'Total (top 20)')
    total_cell(ws, r, 2, f'=SUM(B{ft}:B{r-1})')
    span_top_border(ws, r, 5)

    # ── Section C: Per-Program Subaward Summary ──
    r += 3; subsec_band(ws, r, '(C) Per-Program Subaward Summary', 5)
    r += 1; hdr_row(ws, r, [('Program', 30), ('Total Subs ($K)', 16), ('Top Sub 1', 26), ('Top Sub 2', 26), ('Top Sub 3', 26)])
    ft = r + 1
    for prog, val, s1, s2, s3 in _PROGRAMS:
        r += 1
        wc(ws, r, 1, prog, font=F_DATA)
        wc(ws, r, 2, val, font=F_BLUE, fmt=NUM_FMT)
        wc(ws, r, 3, s1, font=F_GRAY)
        wc(ws, r, 4, s2, font=F_GRAY)
        wc(ws, r, 5, s3, font=F_GRAY)
    r += 1; total_label(ws, r, 1, 'Total')
    total_cell(ws, r, 2, f'=SUM(B{ft}:B{r-1})')
    span_top_border(ws, r, 5)

    # ── Section D: FY26 Contract Vehicle Crosswalk ──
    r += 3; subsec_band(ws, r, '(D) FY2026 SAM Line Item to Contract Vehicle Crosswalk', 3)

    for hull, lines in _CROSSWALK.items():
        r += 2
        # Hull sub-header (sub-sub-section style) — 3 cols only
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.fill = BG_GRAY; cell.alignment = _NOWRAP
        wc(ws, r, 1, hull, font=F_SUBSUBSEC, fill=BG_GRAY)

        r += 1; hdr_row(ws, r, [('FY26 SAM Line', 30), ('FY26 ($K)', 16), ('Contract Vehicle / Prime', 44)])
        ft = r + 1
        for line, val, vehicle in lines:
            r += 1
            wc(ws, r, 1, line, font=F_DATA)
            wc(ws, r, 2, val, font=F_BLUE, fmt=NUM_FMT)
            wc(ws, r, 3, vehicle, font=F_GRAY)
        r += 1; total_label(ws, r, 1, f'{hull} Total')
        total_cell(ws, r, 2, f'=SUM(B{ft}:B{r-1})')
        span_top_border(ws, r, 3)

    finish_sheet(ws)


# ── Named ranges ──────────────────────────────────────────────

def create_named_ranges(wb):
    JB = "'J Book Items Cons.'"
    shared = [
        ('JB_A', 'A'), ('JB_B', 'E'), ('JB_F', 'F'),
        ('JB_S', 'V'), ('JB_U', 'X'), ('JB_V', 'Y'), ('JB_W', 'Z'),
    ]
    fy_ranges = []
    for fyc in FY_CONFIGS:
        fy_ranges.extend(fyc['named_ranges'])

    for nm, col in shared + fy_ranges:
        if nm in wb.defined_names:
            del wb.defined_names[nm]
        wb.defined_names.add(DefinedName(nm, attr_text=f"{JB}!${col}$5:${col}$3381"))


# ── Main ──────────────────────────────────────────────────────

def main():
    out_path = next_output_path()
    print(f'Source: {WORKBOOK_SRC}')
    print(f'Output: {out_path}')
    shutil.copy2(WORKBOOK_SRC, out_path)

    wb = openpyxl.load_workbook(out_path)

    # Strip all non-base sheets
    for nm in list(wb.sheetnames):
        if nm not in BASE_SHEETS:
            del wb[nm]
    print(f'Base sheets kept: {wb.sheetnames}')

    # Named ranges
    create_named_ranges(wb)

    # Build each fiscal year
    generated = []
    for fyc in FY_CONFIGS:
        label  = fyc['label']
        prefix = fyc['prefix']
        tab_color = fyc['tab_color']
        print(f'\n--- {label} ---')

        d = read_data(WORKBOOK_SRC, fyc['read_indices'])
        print(f'  TAM types: {len(d["all_tam_types"])}, SAM types: {len(d["all_sam_types"])}')
        print(f'  NB non-zero: {len(d["nb_nz"])}, NB SAM non-zero: {len(d["nb_sam_nz"])}')
        print(f'  MRO non-zero: {len(d["mro_nz"])}, MRO SAM non-zero: {len(d["mro_sam_nz"])}')
        print(f'  MRO buckets with data: {d["mro_bkts_with_data"]}')

        # Divider tab
        divider_name = f'{label} >>'
        div_ws = wb.create_sheet(divider_name)
        div_ws.sheet_properties.tabColor = tab_color
        band(div_ws, 1, 6, label, F_TITLE, BG_BLACK)
        div_ws.sheet_view.showGridLines = False
        div_ws.row_dimensions[1].height = _RH

        cf, cf_neg, cf_inner = make_cascade(fyc['value_ranges'])

        args = (wb, d, label, prefix, cf, cf_neg, cf_inner)
        create_total_funding(*args)
        create_newbuild_tam(*args)
        create_newbuild_sam(*args)
        mro_tam_info = create_mro_tam(*args)
        create_mro_sam(*args, mro_tam_info=mro_tam_info)

        # Apply tab colors to all sheets in this FY group
        fy_sheets = [
            f'{prefix} Total Funding', f'{prefix} Newbuild TAM',
            f'{prefix} Newbuild SAM',  f'{prefix} MRO TAM',
            f'{prefix} MRO SAM',
        ]
        for sn in fy_sheets:
            wb[sn].sheet_properties.tabColor = tab_color

        generated.append(divider_name)
        generated.extend([
            f'{prefix} Total Funding', f'{prefix} Newbuild TAM',
            f'{prefix} Newbuild SAM',  f'{prefix} MRO TAM',
            f'{prefix} MRO SAM',
        ])

    # Competitive Dynamics
    print('\n--- Competitive Dynamics ---')
    create_competitive_dynamics(wb)
    generated.append('Competitive Dynamics')

    # Order: Sheet 1, then generated sheets in FY order, then remaining base sheets
    base_first = 'J Book Items Cons.'
    base_rest  = [s for s in wb.sheetnames if s in BASE_SHEETS and s != base_first]
    ordered    = [base_first] + generated + base_rest
    wb._sheets = [wb[s] for s in ordered]

    wb.save(out_path)
    print(f'\nSaved {out_path}')
    print(f'Sheet order: {[s for s in wb.sheetnames]}')

    # Verify formula lengths
    wb2 = openpyxl.load_workbook(out_path)
    max_len = 0; max_cell = ''
    for nm in generated:
        for row in wb2[nm].iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith('=') and len(v) > max_len:
                    max_len = len(v); max_cell = f'{nm}!R{cell.row}C{cell.column}'
    wb2.close()
    status = 'OK' if max_len < 8192 else 'OVER LIMIT!'
    print(f'Max formula: {max_len} chars at {max_cell} — {status}')


if __name__ == '__main__':
    main()
