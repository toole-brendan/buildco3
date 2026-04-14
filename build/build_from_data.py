#!/usr/bin/env python3
"""
build_from_data.py — Build market-sizing sheets from an assembled data workbook.

Reads "J Book Items Cons." (Sheet 1) from a pre-assembled workbook (with derived
columns already in place) and generates all market-sizing sheets for every
configured fiscal year. Idempotent: strips and rebuilds calculated sheets on
each run.

Unlike build_workbook.py (which reads the v2.3 source and inserts derived
columns at build time), this script expects the data sheet to already contain
Exhibit Level (col X) and Cost Category (col Y) between Parent Item and
Vessel Service.

Usage:
    python3 build/build_from_data.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from collections import defaultdict
from copy import copy
import shutil
import glob
import re
import os
from prototype_annotations import add_sticky_notes

# ── Configuration ─────────────────────────────────────────────

WORKBOOK_SRC = 'build/data_v2.xlsx'
OUTPUT_DIR   = 'output'
FILE_PREFIX  = '08APR2028_Newbuild_and_MRO_Spend'
VERSION_MAJOR = 5  # auto-increments minor: v5.0, v5.1, v5.2, ...


ARCHIVE_DIR = f'{OUTPUT_DIR}/archive'


def next_output_path():
    """Scan output/ for existing v{MAJOR}.x files, archive the previous minor
    version if present, and return the next version path."""
    import os
    pattern = re.compile(rf'{re.escape(FILE_PREFIX)}_v{VERSION_MAJOR}\.(\d+)\.xlsx$')
    max_minor = -1
    for path in glob.glob(f'{OUTPUT_DIR}/{FILE_PREFIX}_v{VERSION_MAJOR}.*.xlsx'):
        m = pattern.search(path)
        if m:
            max_minor = max(max_minor, int(m.group(1)))
    next_minor = max_minor + 1

    # Archive the previous minor version (n-1) if it exists in output/
    if next_minor >= 1:
        prev = f'{OUTPUT_DIR}/{FILE_PREFIX}_v{VERSION_MAJOR}.{next_minor - 1}.xlsx'
        if os.path.exists(prev):
            os.makedirs(ARCHIVE_DIR, exist_ok=True)
            dest = f'{ARCHIVE_DIR}/{os.path.basename(prev)}'
            shutil.move(prev, dest)
            print(f'Archived: {prev} → {dest}')

    return f'{OUTPUT_DIR}/{FILE_PREFIX}_v{VERSION_MAJOR}.{next_minor}.xlsx'

# Sheets that are hand-built data — never touched by this script
BASE_SHEETS = {'J Book Items Cons.', 'Notes'}

# Each entry produces 5 sheets (Total Funding, Newbuild TAM/SAM, MRO TAM/SAM).
FY_CONFIGS = [
    {
        'label': 'FY2026',
        'prefix': 'FY26',
        'named_ranges': [('JB_R', 'Y'), ('JB_N', 'U'), ('JB_L', 'S')],
        'best_value_range': 'JB_26BV',
        'read_indices': [24, 20, 18],  # Y=FY26 DAA Enacted, U=FY26 Total, S=FY26 Request
        'best_value_formula': 'IF(Y{r}<>"",Y{r},IF(U{r}<>"",U{r},S{r}))',
        'tab_color': '4472C4',   # muted blue
    },
    {
        'label': 'FY2027',
        'prefix': 'FY27',
        'named_ranges': [('JB_27U', 'AB'), ('JB_27S', 'Z')],
        'best_value_range': 'JB_27BV',
        'read_indices': [27, 25],  # AB=FY27 Total, Z=FY27 Request
        'best_value_formula': 'IF(AB{r}<>"",AB{r},Z{r})',
        'tab_color': '548235',   # muted green
    },
]


# ── Taxonomy ──────────────────────────────────────────────────

TAM_CATEGORIES = {'Combatant Ships', 'Auxiliary Ships', 'Cutters', 'Unmanned Maritime Platforms'}
TAM_EXCLUDED_CATEGORIES = ['Combatant Crafts', 'Support Crafts', 'Boats']
SAM_EXCLUDED_TYPES = {'Submarines', 'Aircraft Carriers', 'Unmanned Undersea Vehicles'}

BUCKET_NAMES = {
    '1': 'New Construction', '2': 'Scheduled Depot Maintenance & Repair',
    '3': 'Continuous / Intermediate / Emergent Maintenance',
    '4': 'Modernization & Alteration Installation',
    '5': 'Major Life-Cycle Events / SLEP / MMA / RCOH',
    '6': 'Sustainment Engineering / Planning / Obsolescence Support',
    '7': 'Availability Support / Husbanding / Port Services',
}
BUCKET_SHORT = {
    '2': 'Scheduled Depot Maintenance', '3': 'Continuous / Emergent Maintenance',
    '4': 'Modernization', '5': 'Major Life-Cycle Events / RCOH',
    '6': 'Sustainment Engineering', '7': 'Availability Support',
}
MRO_BKTS = ['2', '3', '4', '5', '6', '7']
MRO_SAM_BKTS = ['2', '3', '4', '5']
MRO_SAM_BKT_LABELS = {
    '2': 'SDM', '3': 'Cont/Emerg', '4': 'Mod', '5': 'Major LC',
}
MRO_NON_SAM_BKTS = ['6', '7']


# ── P-5c / P-8a cost breakdown ───────────────────────────────

P5C_COST_CATEGORIES = [
    'Plan Costs',
    'Basic Construction',
    'Change Orders',
    'Electronics',
    'Propulsion Equipment',
    'Hull, Mechanical, and Electrical',
    'Ordnance',
    'Other Cost',
    'Technology Insertion',
]
P5C_VALID = set(P5C_COST_CATEGORIES)

P8A_COST_CATEGORIES = {'Electronics', 'Hull, Mechanical, and Electrical', 'Ordnance'}


def parse_exhibit_level(title):
    """Extract exhibit level from an ALT_VIEW title string.
    Returns 'P-5c', 'P-8a', 'P-35', or None."""
    if not title:
        return None
    t = str(title).strip()
    if ' - P-5c - ' in t:
        return 'P-5c'
    if ' - P-8a ' in t:
        for cat in P8A_COST_CATEGORIES:
            if f' - P-8a {cat} - ' in t:
                return 'P-8a'
        return None
    if ' - P-35 ' in t:
        return 'P-35'
    return None


def parse_cost_category(title, exhibit_level):
    """Extract P-5c cost category from an ALT_VIEW title string.
    For P-5c rows: the category after ' - P-5c - '.
    For P-8a/P-35 rows: the P-5c parent category (Electronics, HM&E, Ordnance).
    Returns a category string or None."""
    if not title or not exhibit_level:
        return None
    t = str(title).strip()

    if exhibit_level == 'P-5c':
        cat = t.split(' - P-5c - ', 1)[1].strip()
        return cat if cat in P5C_VALID else None

    if exhibit_level == 'P-8a':
        for cat in P8A_COST_CATEGORIES:
            if f' - P-8a {cat} - ' in t:
                return cat
        return None

    if exhibit_level == 'P-35':
        return None

    return None


# ── Styles (per workbook styling guide) ──────────────────────

_F = 'Arial'
_S = 8
_RH = 10.0
_NOWRAP = Alignment(wrap_text=False, vertical='center')

# Text-color palette
_BLACK  = '000000'
_WHITE  = 'FFFFFF'
_SLATE  = '5B7A99'   # secondary slate (91,122,153)
_GREEN  = '2E7D32'   # cross-tab link
_NAVY   = '1F314F'   # primary navy (31,49,79)

# Fonts — all Arial 8, hierarchy via bold/italic/color only
F_TITLE    = Font(name=_F, size=_S, bold=True,  color=_WHITE)   # white on black band
F_PURPOSE  = Font(name=_F, size=_S, italic=True, color='808080') # gray italic purpose line
F_SUBSEC   = Font(name=_F, size=_S, bold=True,  color=_WHITE)   # subsection header (white on gray)
F_SUBSUBSEC = Font(name=_F, size=_S, bold=True, color=_BLACK)  # sub-subsection header (black on light gray)
F_HDR      = Font(name=_F, size=_S, bold=True,  color=_BLACK)   # column header
F_DATA     = Font(name=_F, size=_S, color=_BLACK)               # formula / label
F_GREEN    = Font(name=_F, size=_S, color=_GREEN)               # cross-tab link
F_GRAY     = Font(name=_F, size=_S, italic=True, color='808080') # note / helper text
F_TOTAL    = Font(name=_F, size=_S, bold=True,  color=_BLACK)   # total row
F_PCT      = Font(name=_F, size=_S, italic=True, color=_BLACK)  # percentage data
F_PCT_T    = Font(name=_F, size=_S, bold=True, italic=True, color=_BLACK)  # percentage total
F_KPI      = Font(name=_F, size=_S, bold=True,  color=_NAVY)    # KPI callout
F_BLUE     = Font(name=_F, size=_S, color='0000FF')             # hardcoded numeric input

# Fills
BG_BLACK = PatternFill(start_color=_BLACK,  end_color=_BLACK,  fill_type='solid')
BG_DGRAY = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # subsection header
BG_GRAY  = PatternFill(start_color='F2F4F7', end_color='F2F4F7', fill_type='solid')  # sub-subsection / col headers
BG_TEAL  = PatternFill(start_color='E0F2F1', end_color='E0F2F1', fill_type='solid')  # 224,242,241

# Borders
B_HDR     = Border(bottom=Side(style='thin'))     # column headers
B_TOT     = Border(top=Side(style='medium'))      # all total rows

# Number formats — zeros as "-", negatives in red parentheses
NUM_FMT = '#,##0;[Red](#,##0);"-"'
PCT_FMT = '0.0%;[Red](0.0%);"-"'


# ── Generic SUMIFS cascade ────────────────────────────────────

def make_cascade(best_value_range):
    """
    Build cf / cf_neg / cf_inner using a pre-computed best-value column
    and the JB_RC (Row Class) helper column.

    The helper columns on the data sheet handle the old cascade logic:
    - FY26 Best: picks Y (DAA Enacted) > U (Total) > S (Request)
    - FY27 Best: picks AB (Total) > Z (Request)
    - Row Class: "ADD" for additive rows (blank/[PARENT]), "ALT" for [ALT_VIEW]
    """
    bv = best_value_range

    def cf(*conditions):
        cond = (',' + ','.join(conditions)) if conditions else ''
        return f'=SUMIFS({bv}{cond},JB_RC,"ADD")'

    def cf_neg(*conditions):
        cond = (',' + ','.join(conditions)) if conditions else ''
        return f'=-SUMIFS({bv}{cond},JB_RC,"ADD")'

    def cf_inner(*conditions):
        cond = (',' + ','.join(conditions)) if conditions else ''
        return f'SUMIFS({bv}{cond},JB_RC,"ADD")'

    return cf, cf_neg, cf_inner


def make_altview_cascade(best_value_range):
    """
    Build cf / cf_inner for [ALT_VIEW] rows using pre-computed best-value column.
    """
    bv = best_value_range

    def cf(*conditions):
        cond = (',' + ','.join(conditions)) if conditions else ''
        return f'=SUMIFS({bv}{cond},JB_RC,"ALT")'

    def cf_inner(*conditions):
        cond = (',' + ','.join(conditions)) if conditions else ''
        return f'SUMIFS({bv}{cond},JB_RC,"ALT")'

    return cf, cf_inner


# ── Cell & layout helpers ─────────────────────────────────────

def cw(ws, c, w):
    ws.column_dimensions[get_column_letter(c)].width = w

def wc(ws, r, c, v, font=None, fill=None, fmt=None, border=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if fmt:    cell.number_format = fmt
    if border: cell.border = border
    cell.alignment = _NOWRAP

def band(ws, r, max_col, text, font, fill):
    """Fill a full-width band (title or section header)."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = fill
        cell.alignment = _NOWRAP
    wc(ws, r, 1, text, font=font, fill=fill)

def title_band(ws, r, text, mc):
    band(ws, r, mc, text, F_TITLE, BG_BLACK)

def subsec_band(ws, r, text, mc):
    band(ws, r, mc, text, F_SUBSEC, BG_DGRAY)

def subsubsec_band(ws, r, text, mc):
    band(ws, r, mc, text, F_SUBSUBSEC, BG_GRAY)

def purpose_row(ws, r, text):
    wc(ws, r, 1, text, font=F_PURPOSE)

def hdr_row(ws, r, headers, cs=1, fill=None):
    for i, (t, w) in enumerate(headers):
        c = cs + i
        wc(ws, r, c, t, font=F_HDR, fill=fill, border=B_HDR)
        cw(ws, c, w)

def total_cell(ws, r, c, v):
    wc(ws, r, c, v, font=F_TOTAL, fmt=NUM_FMT, border=B_TOT)

def total_label(ws, r, c, v):
    wc(ws, r, c, v, font=F_TOTAL, border=B_TOT)

def pct_total(ws, r, c, v):
    wc(ws, r, c, v, font=F_PCT_T, fmt=PCT_FMT, border=B_TOT)

def span_top_border(ws, r, mc):
    """Extend medium top border across all cells in row to last active column."""
    for c in range(1, mc + 1):
        cell = ws.cell(row=r, column=c)
        old = cell.border
        cell.border = Border(top=Side(style='medium'),
                             left=old.left, right=old.right, bottom=old.bottom)

def apply_group_borders(ws, r1, r2, col_start, n_usn, n_cg):
    """Add thin left borders at USN/USCG/Total group boundaries for mekko tables."""
    tc = col_start + n_usn + n_cg  # Total column
    boundaries = [col_start]
    if n_cg > 0:
        boundaries.append(col_start + n_usn)
    boundaries.append(tc)
    for c in boundaries:
        for r in range(r1, r2 + 1):
            cell = ws.cell(row=r, column=c)
            old = cell.border
            cell.border = Border(left=Side(style='thin'),
                                 top=old.top, right=old.right, bottom=old.bottom)

def svc_group(types, svc_map):
    usn = [t for t in types if svc_map.get(t, '') in ('USN', 'MSC', '')]
    cg  = [t for t in types if svc_map.get(t, '') == 'USCG']
    return usn, cg

def svc_order(types, svc_map):
    """Return types ordered USN/MSC first, USCG second (no cell output)."""
    usn, cg = svc_group(types, svc_map)
    return usn + cg

def write_svc_header(ws, r, types, svc_map, col_start=2):
    """Write USN / USCG label row for mekko sections with sub-subsection fill.
    Returns (ordered_types, n_usn, n_cg)."""
    usn, cg = svc_group(types, svc_map)
    ordered = usn + cg
    # Fill the row with sub-subsection bg
    span = col_start + len(ordered)  # through Total column
    for c in range(1, span + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = BG_GRAY
        cell.alignment = _NOWRAP
    if usn: wc(ws, r, col_start, 'USN', font=F_SUBSUBSEC, fill=BG_GRAY)
    if cg:  wc(ws, r, col_start + len(usn), 'USCG', font=F_SUBSUBSEC, fill=BG_GRAY)
    return ordered, len(usn), len(cg)

def finish_sheet(ws):
    """Apply sheet-level formatting: gridlines off, universal row height."""
    ws.sheet_view.showGridLines = False
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = _RH


# ── Data reading ──────────────────────────────────────────────

def read_data(src_path, col_indices):
    """
    Pre-scan Sheet 1 to find which vessel types / buckets have data.
    col_indices: 0-indexed column positions to try (cascade order).
    """
    def read_val(row):
        for i in col_indices:
            v = row[i].value
            if v is not None and isinstance(v, (int, float)): return float(v)
        return 0.0

    def additive(rt):
        if rt is None or str(rt).strip() == '': return True
        return str(rt).strip() == '[PARENT]'

    wb = openpyxl.load_workbook(src_path, data_only=True)
    ws = wb['J Book Items Cons.']

    nb = defaultdict(float); mro = defaultdict(float)
    mro_bkt = defaultdict(lambda: defaultdict(float)); mro_b = defaultdict(float)
    type_svc = {}; type_cat = {}
    all_tam_types = set(); tf_src = defaultdict(float)
    nb_excl = defaultdict(float); mro_excl = defaultdict(float)
    p5c_data = defaultdict(float)   # (vessel_type, cost_category) -> $K
    p5c_types = set()
    p8a_data = defaultdict(float)   # (vessel_type, cost_category, system_name) -> $K
    p8a_types = set()
    # Hull-level accumulators
    nb_hull = defaultdict(float)     # hull -> $K (bucket 1 additive)
    hull_svc = {}                     # hull -> service (USN/MSC/USCG)
    hull_type = {}                    # hull -> vessel type
    p5c_hull = defaultdict(float)    # (hull, cost_category) -> $K
    p5c_hull_set = set()
    p8a_hull = defaultdict(float)    # (hull, cost_category, system_name) -> $K
    p8a_hull_set = set()
    mro_hull = defaultdict(lambda: defaultdict(float))  # hull -> bucket -> $K
    sdm_cost_mix = defaultdict(float)  # SDM cost category -> $K
    sdm_excluded = []  # (title, $K) — hull-specific items excluded from cost mix
    _hull_re = re.compile(r'(USS |DDG |LSD |LCS |CG |CVN |SSN |LPD |LHD |LHA |LCC |MCM |ESB |SSBN |SSGN )')

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        cat = str(row[11].value).strip() if row[11].value else ''   # Vessel Category (col L)
        vt  = str(row[12].value).strip() if row[12].value else ''   # Vessel Type (col M)
        sv  = str(row[10].value).strip() if row[10].value else ''   # Vessel Service (col K)
        hl  = str(row[13].value).strip() if row[13].value else ''   # Vessel Hull (col N)

        if cat in TAM_CATEGORIES and vt:
            all_tam_types.add(vt)
            if sv: type_svc[vt] = sv
            type_cat[vt] = cat
            if hl:
                if sv: hull_svc[hl] = sv
                hull_type[hl] = vt

        rt = row[4].value    # Row Type (col E)
        bk = row[8].value    # Bucket (col I)
        bs = str(bk).strip() if bk is not None else ''

        # ── ALT_VIEW scan (P-5c / P-8a for cost category breakdown) ──
        if str(rt).strip() == '[ALT_VIEW]' and bk is not None and bs == '1':
            title = str(row[3].value).strip() if row[3].value else ''
            ex = parse_exhibit_level(title)
            cc_val = parse_cost_category(title, ex)
            if ex and cc_val and cat in TAM_CATEGORIES and vt and vt not in SAM_EXCLUDED_TYPES:
                val = read_val(row)
                if ex == 'P-5c' and cc_val in P5C_VALID:
                    p5c_data[(vt, cc_val)] += val
                    p5c_types.add(vt)
                    if hl:
                        p5c_hull[(hl, cc_val)] += val
                        p5c_hull_set.add(hl)
                elif ex == 'P-8a':
                    ce = str(row[35].value).strip() if row[35].value else ''  # Cost Element (col AJ — shifted +3 by helper insert)
                    if ce:
                        p8a_data[(vt, cc_val, ce)] += val
                        p8a_types.add(vt)
                        if hl:
                            p8a_hull[(hl, cc_val, ce)] += val
                            p8a_hull_set.add(hl)

        # Collect aggregate SDM cost breakdown from OMN [SUB] rows (no hull)
        if str(rt).strip() == '[SUB]' and bs == '2' and not hl:
            _src = str(row[0].value).strip() if row[0].value else ''
            if _src == 'OMN':
                _title = str(row[3].value).strip() if row[3].value else ''
                if _title:
                    _v = read_val(row)
                    if _hull_re.search(_title):
                        sdm_excluded.append((_title, _v))
                    else:
                        sdm_cost_mix[_title] += _v

        if not additive(rt): continue
        if bk is None: continue
        if bs not in ('1', '2', '3', '4', '5', '6', '7'): continue

        v = read_val(row)
        src = str(row[0].value).strip() if row[0].value else '(blank)'
        tf_src[src] += v

        if bs == '1':
            if cat in TAM_CATEGORIES:
                nb[vt] += v
                if hl: nb_hull[hl] += v
            elif cat: nb_excl[cat] += v
        elif bs in MRO_BKTS:
            if cat in TAM_CATEGORIES:
                mro[vt] += v; mro_bkt[vt][bs] += v; mro_b[bs] += v
                if hl:
                    mro_hull[hl][bs] += v
            elif cat: mro_excl[cat] += v

    wb.close()

    def sorted_svc(types, val_map):
        usn = sorted([t for t in types if type_svc.get(t, '') in ('USN', 'MSC', '')],
                      key=lambda t: -val_map.get(t, 0))
        cg  = sorted([t for t in types if type_svc.get(t, '') == 'USCG'],
                      key=lambda t: -val_map.get(t, 0))
        return usn + cg

    def sorted_svc_hull(hulls, val_map):
        usn = sorted([h for h in hulls if hull_svc.get(h, '') in ('USN', 'MSC', '')],
                      key=lambda h: -val_map.get(h, 0))
        cg  = sorted([h for h in hulls if hull_svc.get(h, '') == 'USCG'],
                      key=lambda h: -val_map.get(h, 0))
        return usn + cg

    all_sam_types = all_tam_types - SAM_EXCLUDED_TYPES
    nb_nz      = [t for t in sorted_svc(all_tam_types, nb)  if nb.get(t, 0) > 0]
    nb_sam_nz  = [t for t in nb_nz if t not in SAM_EXCLUDED_TYPES]
    mro_nz     = [t for t in sorted_svc(all_tam_types, mro) if mro.get(t, 0) > 0]
    def _mro_sam_total(bkt_map):
        return sum(bkt_map.get(b, 0) for b in MRO_SAM_BKTS)

    mro_sam_nz = [t for t in mro_nz if t not in SAM_EXCLUDED_TYPES
                  and _mro_sam_total(mro_bkt[t]) > 0]
    mro_hull_sam_nz = [h for h in sorted_svc_hull(
        {h for h in mro_hull
         if hull_type.get(h, '') not in SAM_EXCLUDED_TYPES
         and _mro_sam_total(mro_hull[h]) > 0},
        {h: _mro_sam_total(mro_hull[h]) for h in mro_hull})
        if _mro_sam_total(mro_hull.get(h, {})) > 0]

    return {
        'all_tam_types':       sorted_svc(all_tam_types, nb),
        'all_sam_types':       sorted_svc(all_sam_types, nb),
        'nb_nz': nb_nz, 'nb_sam_nz': nb_sam_nz,
        'mro_nz': mro_nz, 'mro_sam_nz': mro_sam_nz,
        'all_tam_mro_sorted':  sorted_svc(all_tam_types, mro),
        'all_sam_mro_sorted':  sorted_svc(all_sam_types, {
            t: _mro_sam_total(mro_bkt[t]) for t in all_sam_types
        }),
        'type_svc': type_svc,
        'mro_bkts_with_data':  [b for b in MRO_BKTS if mro_b.get(b, 0) > 0],
        'sources':             [s for s, _ in sorted(tf_src.items(), key=lambda x: -x[1])],
        'nb_excl':  [c for c, v in sorted(nb_excl.items(), key=lambda x: -x[1]) if v > 0],
        'mro_excl': [c for c, v in sorted(mro_excl.items(), key=lambda x: -x[1]) if v > 0],
        # P-5c / P-8a breakdown data
        'p5c_data': dict(p5c_data),
        'p5c_types': sorted_svc([t for t in p5c_types
                                  if any(p5c_data.get((t, c), 0) > 0 for c in P5C_VALID)],
                                 nb),
        'p8a_data': dict(p8a_data),
        'p8a_types': sorted_svc(list(p8a_types), nb),
        # Hull-level data
        'nb_hull': dict(nb_hull),
        'nb_hull_tam_nz': [h for h in sorted_svc_hull(
            {h for h in nb_hull}, nb_hull)
            if nb_hull.get(h, 0) > 0],
        'nb_hull_nz': [h for h in sorted_svc_hull(
            {h for h in nb_hull if hull_type.get(h, '') not in SAM_EXCLUDED_TYPES}, nb_hull)
            if nb_hull.get(h, 0) > 0],
        'hull_svc': hull_svc,
        'hull_type': hull_type,
        'mro_hull': dict(mro_hull),
        'mro_hull_sam_nz': mro_hull_sam_nz,
        'sdm_cost_mix': sorted(sdm_cost_mix.items(), key=lambda x: -x[1]),
        'sdm_excluded': sdm_excluded,
        'p5c_hull': dict(p5c_hull),
        'p5c_hull_types': sorted_svc_hull(
            [h for h in p5c_hull_set if hull_type.get(h, '') not in SAM_EXCLUDED_TYPES
             and any(p5c_hull.get((h, c), 0) > 0 for c in P5C_VALID)],
            {h: sum(p5c_hull.get((h, c), 0) for c in P5C_VALID) for h in p5c_hull_set}),
        'p8a_hull': dict(p8a_hull),
        'p8a_hull_types': sorted_svc_hull(
            [h for h in p8a_hull_set if hull_type.get(h, '') not in SAM_EXCLUDED_TYPES],
            {h: sum(v for (h2, _, _), v in p8a_hull.items() if h2 == h) for h in p8a_hull_set}),
    }


# ── Sheet builders ────────────────────────────────────────────

def create_total_funding(wb, d, label, prefix, cf, cf_neg, cf_inner):
    ws = wb.create_sheet(f'{prefix} Total Funding')
    mc = 3  # max active columns: A, B, C

    r = 1; title_band(ws, r, 'Total Funding', mc)
    if prefix == 'FY27':
        r = 2; purpose_row(ws, r,
            f'All USN, MSC & USCG budget authority for any type of work on any type of vessel — '
            f'from new construction through depot maintenance, modernization, sustainment, and end-of-life. '
            f'Includes line items that lack vessel type or work type attribution. '
            f'{label} budget request — detailed justification books not yet released, program-level visibility only. ($K)')
    else:
        r = 2; purpose_row(ws, r,
            f'All USN, MSC & USCG budget authority for any type of work on any type of vessel — '
            f'from new construction through depot maintenance, modernization, sustainment, and end-of-life. '
            f'Includes line items that lack vessel type or work type attribution. {label} ($K)')

    # (A) By Work Type
    r = 4; subsec_band(ws, r, f'(A) {label} Total Funding by Work Type', 2)
    r += 1; hdr_row(ws, r, [('Work Type', 30), (f'{label} ($K)', 12)])
    first_data = r + 1
    for b in ['1', '2', '3', '4', '5', '6', '7']:
        r += 1; wc(ws, r, 1, BUCKET_NAMES[b], font=F_DATA)
        wc(ws, r, 2, cf(f'JB_B,{b}'), font=F_DATA, fmt=NUM_FMT)
    r += 1; total_label(ws, r, 1, 'Total'); total_cell(ws, r, 2, f'=SUM(B{first_data}:B{r-1})')
    span_top_border(ws, r, 2)

    # (B) By Vessel Category
    r += 3; subsec_band(ws, r, f'(B) {label} Total Funding by Vessel Category', 3)
    r += 1; hdr_row(ws, r, [('Category', 30), (f'{label} ($K)', 12), ('TAM Scope', 18)])
    cats = [
        ('Combatant Ships',           'YES \u2192 TAM'),
        ('Auxiliary Ships',            'YES \u2192 TAM'),
        ('Cutters',                    'YES \u2192 TAM'),
        ('Unmanned Maritime Platforms', 'YES \u2192 TAM'),
        ('Combatant Crafts',           'No \u2014 small craft'),
        ('Support Crafts',             'No \u2014 non-oceangoing'),
        ('Boats',                      'No \u2014 under 65 ft'),
    ]
    tam_rows = []; first = r + 1
    for cat, scope in cats:
        r += 1; wc(ws, r, 1, cat, font=F_DATA)
        wc(ws, r, 2, cf(f'JB_V,"{cat}"'), font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, 3, scope, font=F_GRAY)
        if 'YES' in scope: tam_rows.append(r)
    r += 1; wc(ws, r, 1, 'Unattributed (fleet-wide)', font=F_DATA)
    wc(ws, r, 2, cf('JB_V,""'), font=F_DATA, fmt=NUM_FMT)
    wc(ws, r, 3, 'No \u2014 not vessel-specific', font=F_GRAY)
    r += 1; total_label(ws, r, 1, 'Total'); total_cell(ws, r, 2, f'=SUM(B{first}:B{r-1})')
    span_top_border(ws, r, 3)
    r += 1; wc(ws, r, 1, 'TAM-eligible', font=F_KPI, fill=BG_TEAL)
    wc(ws, r, 2, '=' + '+'.join(f'B{tr}' for tr in tam_rows), font=F_KPI, fill=BG_TEAL, fmt=NUM_FMT)

    # (C) By Source Book
    r += 3; subsec_band(ws, r, f'(C) {label} Total Funding by Source Book', 2)
    r += 1; hdr_row(ws, r, [('Source Book', 30), (f'{label} ($K)', 12)]); fs = r + 1
    for s in d['sources']:
        r += 1; wc(ws, r, 1, s, font=F_DATA)
        wc(ws, r, 2, cf(f'JB_A,"{s}"'), font=F_DATA, fmt=NUM_FMT)
    r += 1; total_label(ws, r, 1, 'Total'); total_cell(ws, r, 2, f'=SUM(B{fs}:B{r-1})')
    span_top_border(ws, r, 2)

    finish_sheet(ws)


def create_newbuild_tam(wb, d, label, prefix, cf, cf_neg, cf_inner):
    ws = wb.create_sheet(f'{prefix} Newbuild TAM')
    hull_tam_nz = d.get('nb_hull_tam_nz', [])
    mc = max(5, 2 + len(d['nb_nz']) + 1, 2 + len(hull_tam_nz) + 1)
    cw(ws, 1, 30); cw(ws, 2, 12)

    r = 1; title_band(ws, r, f'Newbuild TAM \u2014 {label} ($K)', mc)
    r = 2; purpose_row(ws, r,
        'Total Funding narrowed to new construction only and only oceangoing vessels \u2014 '
        'excludes small craft (Combatant Crafts, Support Crafts) and USCG Boats. '
        'Limited to line items with clear vessel type designation; unattributed fleet-wide items are excluded.')

    # (A) Bridge
    r = 4; subsec_band(ws, r, '(A) Bridge: Total NC Funding \u2192 Newbuild TAM', 2)
    bridge_start = r + 1
    r += 1; wc(ws, r, 1, 'Total New Construction Funding', font=F_DATA)
    wc(ws, r, 2, cf('JB_B,1'), font=F_DATA, fmt=NUM_FMT)
    r += 1; wc(ws, r, 1, '  Less: Unattributed (fleet-wide NC items)', font=F_DATA)
    wc(ws, r, 2, cf_neg('JB_B,1', 'JB_V,""'), font=F_DATA, fmt=NUM_FMT)
    for cat in TAM_EXCLUDED_CATEGORIES:
        r += 1; wc(ws, r, 1, f'  Less: {cat}', font=F_DATA)
        wc(ws, r, 2, cf_neg('JB_B,1', f'JB_V,"{cat}"'), font=F_DATA, fmt=NUM_FMT)
    r += 1; wc(ws, r, 1, '  Less: No vessel type designation', font=F_DATA)
    novt_f = '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"', 'JB_W,""') for c in sorted(TAM_CATEGORIES))
    wc(ws, r, 2, f'=-({novt_f})', font=F_DATA, fmt=NUM_FMT)
    bridge_end = r
    r += 1; total_label(ws, r, 1, 'Newbuild TAM')
    total_cell(ws, r, 2, f'=SUM(B{bridge_start}:B{bridge_end})')
    span_top_border(ws, r, 2)

    # (B) Newbuild TAM by Vessel Type
    r += 3; subsec_band(ws, r, '(B) Newbuild TAM by Vessel Type', 4)
    all_types = svc_order(d['all_tam_types'], d['type_svc'])
    r += 1; hdr_row(ws, r, [('Vessel Type', 30), ('Service', 8), (f'{label} ($K)', 12), ('% of TAM', 10)])
    ft = r + 1
    for vt in all_types:
        r += 1; wc(ws, r, 1, vt, font=F_DATA)
        wc(ws, r, 2, d['type_svc'].get(vt, ''), font=F_DATA)
        wc(ws, r, 3, cf('JB_B,1', f'JB_W,"{vt}"'), font=F_DATA, fmt=NUM_FMT)
    lt = r; r += 1; tr = r
    total_label(ws, r, 1, 'Newbuild TAM'); total_label(ws, r, 2, '')
    total_cell(ws, r, 3, f'=SUM(C{ft}:C{lt})')
    pct_total(ws, r, 4, 1.0)
    span_top_border(ws, r, 4)
    for rn in range(ft, lt + 1):
        wc(ws, rn, 4, f'=IFERROR(C{rn}/C${tr},0)', font=F_PCT, fmt=PCT_FMT)

    # Dynamic section lettering (C/D/E skip gracefully if no hull data)
    _sec = ord('C')
    def next_sec():
        nonlocal _sec; ch = chr(_sec); _sec += 1; return ch

    # (C) Newbuild TAM by Hull Program
    if hull_tam_nz:
        sl = next_sec()
        r += 3
        subsec_band(ws, r, f'({sl}) Newbuild TAM by Hull Program \u2014 {label} ($K)', 5)
        r += 1
        purpose_row(ws, r, 'Same funding as (B) broken down to individual hull programs.')
        r += 1
        hdr_row(ws, r, [('Hull Program', 16), ('Vessel Type', 24), ('Service', 8),
                         (f'{label} ($K)', 12), ('% of TAM', 10)])
        ft_h = r + 1
        for hl in hull_tam_nz:
            r += 1
            wc(ws, r, 1, hl, font=F_DATA)
            wc(ws, r, 2, d['hull_type'].get(hl, ''), font=F_DATA)
            wc(ws, r, 3, d['hull_svc'].get(hl, ''), font=F_DATA)
            wc(ws, r, 4, cf('JB_B,1', f'JB_H,"{hl}"'), font=F_DATA, fmt=NUM_FMT)
        lt_h = r; r += 1; tr_h = r
        total_label(ws, r, 1, 'Newbuild TAM'); total_label(ws, r, 2, '')
        total_label(ws, r, 3, '')
        total_cell(ws, r, 4, f'=SUM(D{ft_h}:D{lt_h})')
        pct_total(ws, r, 5, 1.0)
        span_top_border(ws, r, 5)
        for rn in range(ft_h, lt_h + 1):
            wc(ws, rn, 5, f'=IFERROR(D{rn}/D${tr_h},0)', font=F_PCT, fmt=PCT_FMT)

    # (D) Mekko — Sub-Category Mix by Vessel Type
    sl = next_sec()
    mk = svc_order(d['nb_nz'], d['type_svc'])
    r += 3; subsec_band(ws, r, f'({sl}) Newbuild TAM \u2014 Sub-Category Mix by Vessel Type (Mekko)', 2 + len(mk))
    r += 1; svc_r = r; mk, n_usn, n_cg = write_svc_header(ws, r, d['nb_nz'], d['type_svc'])
    r += 1; wc(ws, r, 1, 'Sub-Category', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
    for i, vt in enumerate(mk):
        c = 2 + i; wc(ws, r, c, vt, font=F_HDR, border=B_HDR); cw(ws, c, 12)
    tc = 2 + len(mk); wc(ws, r, tc, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc, 12)
    dr = r + 1; ar = r + 2; rcoh_r = r + 3; slep_r = r + 4; tkr = r + 5

    ddc = ['"Full ship DD&C"', '"Full ship DD&C / LLTM"', '"Full vessel procurement"']

    wc(ws, tkr, 1, 'Total ($K)', font=F_TOTAL, border=B_TOT)
    for i, vt in enumerate(mk):
        total_cell(ws, tkr, 2 + i, cf('JB_B,1', f'JB_W,"{vt}"'))
    total_cell(ws, tkr, tc, f'=SUM({get_column_letter(2)}{tkr}:{get_column_letter(tc-1)}{tkr})')

    tcl = get_column_letter(tc)
    fcl = get_column_letter(2); lcl = get_column_letter(tc - 1)

    wc(ws, dr, 1, 'Full Ship DD&C', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        parts = '+'.join(cf_inner('JB_B,1', f'JB_W,"{vt}"', f'JB_F,{s}') for s in ddc)
        wc(ws, dr, c, f'=IFERROR(({parts})/{cl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)
    wc(ws, dr, tc, f'=IFERROR(SUMPRODUCT({fcl}{dr}:{lcl}{dr},{fcl}${tkr}:{lcl}${tkr})/{tcl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)

    wc(ws, ar, 1, 'Advance Procurement / LLTM', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        ap = cf_inner('JB_B,1', f'JB_W,"{vt}"', 'JB_F,"Advance procurement / LLTM"')
        wc(ws, ar, c, f'=IFERROR(({ap})/{cl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)
    wc(ws, ar, tc, f'=IFERROR(SUMPRODUCT({fcl}{ar}:{lcl}{ar},{fcl}${tkr}:{lcl}${tkr})/{tcl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)

    wc(ws, rcoh_r, 1, 'RCOH', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        rcoh = cf_inner('JB_B,1', f'JB_W,"{vt}"', 'JB_F,"RCOH"')
        wc(ws, rcoh_r, c, f'=IFERROR(({rcoh})/{cl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)
    wc(ws, rcoh_r, tc, f'=IFERROR(SUMPRODUCT({fcl}{rcoh_r}:{lcl}{rcoh_r},{fcl}${tkr}:{lcl}${tkr})/{tcl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)

    wc(ws, slep_r, 1, 'SLEP', font=F_DATA)
    for i, vt in enumerate(mk):
        c = 2 + i; cl = get_column_letter(c)
        slep = cf_inner('JB_B,1', f'JB_W,"{vt}"', 'JB_F,"SLEP"')
        wc(ws, slep_r, c, f'=IFERROR(({slep})/{cl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)
    wc(ws, slep_r, tc, f'=IFERROR(SUMPRODUCT({fcl}{slep_r}:{lcl}{slep_r},{fcl}${tkr}:{lcl}${tkr})/{tcl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)

    span_top_border(ws, tkr, tc)
    apply_group_borders(ws, svc_r, tkr, 2, n_usn, n_cg)
    r = tkr

    # (E) Mekko — Sub-Category Mix by Hull Program
    if hull_tam_nz:
        sl = next_sec()
        r += 3; subsec_band(ws, r, f'({sl}) Newbuild TAM \u2014 Sub-Category Mix by Hull Program (Mekko)',
                            2 + len(hull_tam_nz))
        r += 1; svc_r_h = r
        hull_mk, n_usn_h, n_cg_h = write_svc_header(ws, r, hull_tam_nz, d['hull_svc'])
        r += 1; wc(ws, r, 1, 'Sub-Category', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
        for i, hl in enumerate(hull_mk):
            c = 2 + i; wc(ws, r, c, hl, font=F_HDR, border=B_HDR); cw(ws, c, 12)
        tc_h = 2 + len(hull_mk)
        wc(ws, r, tc_h, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc_h, 12)
        dr_h = r + 1; ar_h = r + 2; rcoh_h = r + 3; slep_h = r + 4; tkr_h = r + 5

        wc(ws, tkr_h, 1, 'Total ($K)', font=F_TOTAL, border=B_TOT)
        for i, hl in enumerate(hull_mk):
            total_cell(ws, tkr_h, 2 + i, cf('JB_B,1', f'JB_H,"{hl}"'))
        total_cell(ws, tkr_h, tc_h,
                   f'=SUM({get_column_letter(2)}{tkr_h}:{get_column_letter(tc_h-1)}{tkr_h})')

        tcl_h = get_column_letter(tc_h)

        wc(ws, dr_h, 1, 'Full Ship DD&C', font=F_DATA)
        for i, hl in enumerate(hull_mk):
            c = 2 + i; cl = get_column_letter(c)
            parts = '+'.join(cf_inner('JB_B,1', f'JB_H,"{hl}"', f'JB_F,{s}') for s in ddc)
            wc(ws, dr_h, c, f'=IFERROR(({parts})/{cl}${tkr_h},0)', font=F_PCT, fmt=PCT_FMT)
        ddc_cat_h = '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"', f'JB_F,{s}') for c in sorted(TAM_CATEGORIES) for s in ddc)
        wc(ws, dr_h, tc_h, f'=IFERROR(({ddc_cat_h})/{tcl_h}${tkr_h},0)', font=F_PCT, fmt=PCT_FMT)

        wc(ws, ar_h, 1, 'Advance Procurement / LLTM', font=F_DATA)
        for i, hl in enumerate(hull_mk):
            c = 2 + i; cl = get_column_letter(c)
            ap = cf_inner('JB_B,1', f'JB_H,"{hl}"', 'JB_F,"Advance procurement / LLTM"')
            wc(ws, ar_h, c, f'=IFERROR(({ap})/{cl}${tkr_h},0)', font=F_PCT, fmt=PCT_FMT)
        ap_cat_h = '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"',
                   'JB_F,"Advance procurement / LLTM"') for c in sorted(TAM_CATEGORIES))
        wc(ws, ar_h, tc_h, f'=IFERROR(({ap_cat_h})/{tcl_h}${tkr_h},0)', font=F_PCT, fmt=PCT_FMT)

        wc(ws, rcoh_h, 1, 'RCOH', font=F_DATA)
        for i, hl in enumerate(hull_mk):
            c = 2 + i; cl = get_column_letter(c)
            rcoh = cf_inner('JB_B,1', f'JB_H,"{hl}"', 'JB_F,"RCOH"')
            wc(ws, rcoh_h, c, f'=IFERROR(({rcoh})/{cl}${tkr_h},0)', font=F_PCT, fmt=PCT_FMT)
        rcoh_cat_h = '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"', 'JB_F,"RCOH"') for c in sorted(TAM_CATEGORIES))
        wc(ws, rcoh_h, tc_h, f'=IFERROR(({rcoh_cat_h})/{tcl_h}${tkr_h},0)', font=F_PCT, fmt=PCT_FMT)

        wc(ws, slep_h, 1, 'SLEP', font=F_DATA)
        for i, hl in enumerate(hull_mk):
            c = 2 + i; cl = get_column_letter(c)
            slep = cf_inner('JB_B,1', f'JB_H,"{hl}"', 'JB_F,"SLEP"')
            wc(ws, slep_h, c, f'=IFERROR(({slep})/{cl}${tkr_h},0)', font=F_PCT, fmt=PCT_FMT)
        slep_cat_h = '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"', 'JB_F,"SLEP"') for c in sorted(TAM_CATEGORIES))
        wc(ws, slep_h, tc_h, f'=IFERROR(({slep_cat_h})/{tcl_h}${tkr_h},0)', font=F_PCT, fmt=PCT_FMT)

        span_top_border(ws, tkr_h, tc_h)
        apply_group_borders(ws, svc_r_h, tkr_h, 2, n_usn_h, n_cg_h)

    finish_sheet(ws)


def create_newbuild_sam(wb, d, label, prefix, cf, cf_neg, cf_inner, altview_cf=None, altview_inner=None):
    ws = wb.create_sheet(f'{prefix} Newbuild SAM')

    ddc_subs = ['"Full ship DD&C"', '"Full ship DD&C / LLTM"', '"Full vessel procurement"']
    p5c_hulls = d.get('p5c_hull_types', [])
    hull_nz = d.get('nb_hull_nz', [])
    has_p5c = bool(p5c_hulls) and altview_inner is not None

    # Max columns across all sections
    mc_subcat = 8  # VT, Svc, DD&C, AP, RCOH, SLEP, Total, %
    mc_hull = 9    # Hull, VT, Svc, DD&C, AP, RCOH, SLEP, Total, %
    tc_p5c = 2 + len(p5c_hulls) if has_p5c else 4
    mc = max(mc_hull, tc_p5c)
    cw(ws, 1, 30)

    # ── Title & Purpose ──
    r = 1; title_band(ws, r, f'Newbuild SAM \u2014 {label} ($K)', mc)
    if prefix == 'FY27':
        r = 2; purpose_row(ws, r,
            'Newbuild TAM narrowed further: excludes Submarines and Aircraft Carriers '
            '(single-yard / nuclear-restricted programs) and Unmanned Undersea Vehicles. '
            'Focused on outsourceable new construction where a company could compete as subprime '
            'on module fabrication or systems integration. Includes all new-construction funding '
            '(DD&C and advance procurement). Justification books (P-5c, P-8a) not yet released for FY27 — '
            'visibility is at program level only, not cost element level.')
    else:
        r = 2; purpose_row(ws, r,
            'Outsourceable new construction sized at cost-element level. Starts from Newbuild TAM, '
            'excludes single-yard/nuclear programs (Submarines, Aircraft Carriers, UUVs), then narrows '
            'to hull programs with P-5c exhibit data. P-5c percentage mix applied to net budget authority '
            'gives the Newbuild SAM — the market a company could compete in, broken by component.')

    # ── (A) Bridge: Newbuild TAM → Newbuild SAM (all programs) ──
    r = 4; subsec_band(ws, r, '(A) Bridge: Newbuild TAM \u2192 Newbuild SAM (all programs)', 2)
    r += 1; bridge_first = r
    wc(ws, r, 1, 'Newbuild TAM', font=F_DATA)
    wc(ws, r, 2, '=' + '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES)), font=F_DATA, fmt=NUM_FMT)
    for vt in ['Submarines', 'Aircraft Carriers', 'Unmanned Undersea Vehicles']:
        r += 1; wc(ws, r, 1, f'  Less: {vt}', font=F_DATA)
        wc(ws, r, 2, cf_neg('JB_B,1', f'JB_W,"{vt}"'), font=F_DATA, fmt=NUM_FMT)
    r += 1; total_label(ws, r, 1, 'Newbuild SAM (all programs)')
    total_cell(ws, r, 2, f'=SUM(B{bridge_first}:B{r-1})')
    span_top_border(ws, r, 2)

    # ── (B) SAM (all programs) by Vessel Type & Sub-Category ──
    vt_nz = svc_order(d['nb_sam_nz'], d['type_svc'])
    r += 3; subsec_band(ws, r, f'(B) Newbuild SAM (all programs) by Vessel Type & Sub-Category \u2014 {label} ($K)', mc_subcat)
    r += 1; hdr_row(ws, r, [
        ('Vessel Type', 30), ('Service', 8),
        ('DD&C', 12), ('AP/LLTM', 12), ('RCOH', 12), ('SLEP', 12),
        ('Total', 12), ('% of Total', 10),
    ])
    ft_b = r + 1
    for vt in vt_nz:
        r += 1
        wc(ws, r, 1, vt, font=F_DATA)
        wc(ws, r, 2, d['type_svc'].get(vt, ''), font=F_DATA)
        ddc_parts = '+'.join(cf_inner('JB_B,1', f'JB_W,"{vt}"', f'JB_F,{s}') for s in ddc_subs)
        wc(ws, r, 3, f'={ddc_parts}', font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, 4, cf('JB_B,1', f'JB_W,"{vt}"', 'JB_F,"Advance procurement / LLTM"'), font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, 5, cf('JB_B,1', f'JB_W,"{vt}"', 'JB_F,"RCOH"'), font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, 6, cf('JB_B,1', f'JB_W,"{vt}"', 'JB_F,"SLEP"'), font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, 7, cf('JB_B,1', f'JB_W,"{vt}"'), font=F_DATA, fmt=NUM_FMT)
    lt_b = r; r += 1; tr_b = r
    total_label(ws, r, 1, 'Newbuild SAM (all programs)'); total_label(ws, r, 2, '')
    for c in range(3, 8):
        total_cell(ws, r, c, f'=SUM({get_column_letter(c)}{ft_b}:{get_column_letter(c)}{lt_b})')
    pct_total(ws, r, 8, 1.0)
    span_top_border(ws, r, 8)
    for rn in range(ft_b, lt_b + 1):
        wc(ws, rn, 8, f'=IFERROR(G{rn}/G${tr_b},0)', font=F_PCT, fmt=PCT_FMT)

    # ── (C) SAM (all programs) by Hull Program & Sub-Category ──
    mc_hull = 9  # Hull, VT, Svc, DD&C, AP, RCOH, SLEP, Total, %
    tr_c = None
    if hull_nz:
        r += 3
        subsec_band(ws, r, f'(C) Newbuild SAM (all programs) by Hull Program \u2014 {label} ($K)', mc_hull)
        r += 1
        purpose_row(ws, r, 'Same funding as (B) broken down to individual hull programs — the level at which budget line items are traceable.')
        r += 1
        hdr_row(ws, r, [
            ('Hull Program', 16), ('Vessel Type', 24), ('Service', 8),
            ('DD&C', 12), ('AP/LLTM', 12), ('RCOH', 12), ('SLEP', 12),
            ('Total', 12), ('% of Total', 10),
        ])
        ft_c = r + 1
        for hl in hull_nz:
            r += 1
            wc(ws, r, 1, hl, font=F_DATA)
            wc(ws, r, 2, d['hull_type'].get(hl, ''), font=F_DATA)
            wc(ws, r, 3, d['hull_svc'].get(hl, ''), font=F_DATA)
            ddc_parts = '+'.join(cf_inner('JB_B,1', f'JB_H,"{hl}"', f'JB_F,{s}') for s in ddc_subs)
            wc(ws, r, 4, f'={ddc_parts}', font=F_DATA, fmt=NUM_FMT)
            wc(ws, r, 5, cf('JB_B,1', f'JB_H,"{hl}"', 'JB_F,"Advance procurement / LLTM"'), font=F_DATA, fmt=NUM_FMT)
            wc(ws, r, 6, cf('JB_B,1', f'JB_H,"{hl}"', 'JB_F,"RCOH"'), font=F_DATA, fmt=NUM_FMT)
            wc(ws, r, 7, cf('JB_B,1', f'JB_H,"{hl}"', 'JB_F,"SLEP"'), font=F_DATA, fmt=NUM_FMT)
            wc(ws, r, 8, cf('JB_B,1', f'JB_H,"{hl}"'), font=F_DATA, fmt=NUM_FMT)
        lt_c = r; r += 1; tr_c = r
        total_label(ws, r, 1, 'Newbuild SAM (all programs)')
        total_label(ws, r, 2, ''); total_label(ws, r, 3, '')
        for c in range(4, 9):
            total_cell(ws, r, c, f'=SUM({get_column_letter(c)}{ft_c}:{get_column_letter(c)}{lt_c})')
        pct_total(ws, r, 9, 1.0)
        span_top_border(ws, r, 9)
        for rn in range(ft_c, lt_c + 1):
            wc(ws, rn, 9, f'=IFERROR(H{rn}/H${tr_c},0)', font=F_PCT, fmt=PCT_FMT)

    # ── FY27 / no P-5c: stop here ──
    if not has_p5c:
        if prefix == 'FY27':
            r += 2
            purpose_row(ws, r, 'Justification books (P-5c, P-8a) not yet released for FY27 — '
                'cost-element breakdown and Newbuild SAM (cost-element coverage) are not available.')
        finish_sheet(ws)
        return {'section_d_row': None, 'tc_e': None}

    # ── Bridge: All Programs → Cost-Element Coverage ──
    p5c_set = set(p5c_hulls)
    no_p5c = [hl for hl in hull_nz if hl not in p5c_set]
    no_p5c_uscg = [hl for hl in no_p5c if d['hull_svc'].get(hl, '') == 'USCG']
    no_p5c_other = [hl for hl in no_p5c if d['hull_svc'].get(hl, '') != 'USCG']

    r += 3
    subsec_band(ws, r, '(C\u2192D) Bridge: All Programs \u2192 Cost-Element Coverage', 2)
    r += 1; bridge2_first = r
    wc(ws, r, 1, 'Newbuild SAM (all programs)', font=F_DATA)
    wc(ws, r, 2, f'=H{tr_c}', font=F_DATA, fmt=NUM_FMT)
    for hl in no_p5c_uscg:
        r += 1
        vt_label = d['hull_type'].get(hl, '')
        wc(ws, r, 1, f'  Less: {hl} ({vt_label}, USCG — no P-5c)', font=F_DATA)
        wc(ws, r, 2, cf_neg('JB_B,1', f'JB_H,"{hl}"'), font=F_DATA, fmt=NUM_FMT)
    for hl in no_p5c_other:
        r += 1
        vt_label = d['hull_type'].get(hl, '')
        wc(ws, r, 1, f'  Less: {hl} ({vt_label} — no P-5c)', font=F_DATA)
        wc(ws, r, 2, cf_neg('JB_B,1', f'JB_H,"{hl}"'), font=F_DATA, fmt=NUM_FMT)
    r += 1
    total_label(ws, r, 1, 'Newbuild SAM (cost-element coverage)')
    total_cell(ws, r, 2, f'=SUM(B{bridge2_first}:B{r-1})')
    span_top_border(ws, r, 2)

    # ── (D) P-5c Gross Cost Category Breakdown by Hull ──
    # Only hulls with both P-5c data AND nonzero net BA (same scope as E/F)
    p5c_funded = [h for h in p5c_hulls if d['nb_hull'].get(h, 0) > 0]
    cats = [c for c in P5C_COST_CATEGORIES
            if any(d['p5c_hull'].get((h, c), 0) > 0 for h in p5c_funded)]

    n_hulls = len(p5c_funded)
    tc_e = 2 + n_hulls

    r += 3; _section_d_row = r
    subsec_band(ws, r, f'(D) P-5c Gross Cost Category Breakdown by Hull \u2014 {label} ($K)', tc_e)
    r += 1
    purpose_row(ws, r, 'Total estimated cost of each ship by component, from Exhibit P-5c. '
        'These are full ship-cost estimates \u2014 not single-year budget authority. '
        'Section (E) applies this cost mix to FY2026 net funding.')
    r += 1
    svc_r_d = r
    p5c_ordered, n_usn_d, n_cg_d = write_svc_header(ws, r, p5c_funded, d['hull_svc'])

    r += 1
    wc(ws, r, 1, 'Cost Category', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
    for i, hl in enumerate(p5c_ordered):
        c = 2 + i
        wc(ws, r, c, hl, font=F_HDR, border=B_HDR); cw(ws, c, 14)
    wc(ws, r, tc_e, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc_e, 14)

    first_d = r + 1
    for cat in cats:
        r += 1
        wc(ws, r, 1, cat, font=F_DATA)
        for i, hl in enumerate(p5c_ordered):
            c = 2 + i
            formula = altview_inner('JB_B,1', f'JB_EX,"P-5c"', f'JB_CC,"{cat}"', f'JB_H,"{hl}"')
            wc(ws, r, c, '=' + formula, font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, tc_e, f'=SUM({get_column_letter(2)}{r}:{get_column_letter(tc_e-1)}{r})',
           font=F_DATA, fmt=NUM_FMT)
    last_d = r

    r += 1; tkr_d = r
    wc(ws, r, 1, 'Gross Total ($K)', font=F_TOTAL, border=B_TOT)
    for c in range(2, tc_e + 1):
        total_cell(ws, r, c,
                   f'=SUM({get_column_letter(c)}{first_d}:{get_column_letter(c)}{last_d})')
    span_top_border(ws, r, tc_e)
    apply_group_borders(ws, svc_r_d, tkr_d, 2, n_usn_d, n_cg_d)

    # Percentage sub-table
    r += 2
    subsubsec_band(ws, r, 'Cost category mix (% of gross total per hull program)', tc_e)
    r += 1
    wc(ws, r, 1, 'Cost Category', font=F_HDR, border=B_HDR)
    for i, hl in enumerate(p5c_ordered):
        wc(ws, r, 2 + i, hl, font=F_HDR, border=B_HDR)
    wc(ws, r, tc_e, 'Total', font=F_HDR, border=B_HDR)

    pct_first = r + 1
    for ci, cat in enumerate(cats):
        r += 1
        data_row = first_d + ci
        wc(ws, r, 1, cat, font=F_DATA)
        for c in range(2, tc_e + 1):
            cl = get_column_letter(c)
            wc(ws, r, c, f'=IFERROR({cl}{data_row}/{cl}${tkr_d},0)',
               font=F_PCT, fmt=PCT_FMT)
    r += 1
    wc(ws, r, 1, 'Total', font=F_TOTAL, border=B_TOT)
    for c in range(2, tc_e + 1):
        pct_total(ws, r, c, 1.0)
    span_top_border(ws, r, tc_e)
    apply_group_borders(ws, pct_first - 2, r, 2, n_usn_d, n_cg_d)

    # ── (E) Newbuild SAM (cost-element coverage) ──
    # Only include P-5c hulls with nonzero net BA; sort descending
    p5c_with_ba = [h for h in p5c_hulls if d['nb_hull'].get(h, 0) > 0]
    p5c_by_net = svc_order(
        sorted(p5c_with_ba, key=lambda h: -d['nb_hull'].get(h, 0)),
        d['hull_svc'])
    tc_ef = 2 + len(p5c_by_net)  # may differ from tc_e if some P-5c hulls have $0 net BA

    r += 3
    subsec_band(ws, r, f'(E) Newbuild SAM (cost-element coverage) \u2014 {label} ($K)', tc_ef)
    r += 1
    purpose_row(ws, r, 'P-5c percentage mix from (D) applied to SAM net budget-authority totals. '
        'This is the Newbuild SAM — the outsourceable new-construction market sized at component level.')
    r += 1
    svc_r_e = r
    p5c_ordered_e, n_usn_e, n_cg_e = write_svc_header(ws, r, p5c_by_net, d['hull_svc'])

    r += 1
    wc(ws, r, 1, 'Cost Category', font=F_HDR, border=B_HDR)
    for i, hl in enumerate(p5c_ordered_e):
        c = 2 + i
        wc(ws, r, c, hl, font=F_HDR, border=B_HDR); cw(ws, c, 14)
    wc(ws, r, tc_ef, 'Total', font=F_HDR, border=B_HDR)

    # SAM net total reference row
    r += 1; sam_ref_r = r
    wc(ws, r, 1, 'SAM Net Total ($K)', font=F_KPI, fill=BG_TEAL)
    for i, hl in enumerate(p5c_ordered_e):
        c = 2 + i
        wc(ws, r, c, cf('JB_B,1', f'JB_H,"{hl}"'), font=F_KPI, fill=BG_TEAL, fmt=NUM_FMT)
    wc(ws, r, tc_ef, f'=SUM({get_column_letter(2)}{r}:{get_column_letter(tc_ef-1)}{r})',
       font=F_KPI, fill=BG_TEAL, fmt=NUM_FMT)

    # Allocated rows — SAM net * (P-5c category / P-5c total) per hull
    first_sam = r + 1
    for cat in cats:
        r += 1
        wc(ws, r, 1, cat, font=F_DATA)
        for i, hl in enumerate(p5c_ordered_e):
            c = 2 + i; cl = get_column_letter(c)
            p5c_cat = altview_inner('JB_B,1', f'JB_EX,"P-5c"', f'JB_CC,"{cat}"', f'JB_H,"{hl}"')
            p5c_tot = altview_inner('JB_B,1', f'JB_EX,"P-5c"', f'JB_H,"{hl}"')
            wc(ws, r, c, f'=IFERROR({cl}${sam_ref_r}*({p5c_cat})/({p5c_tot}),0)',
               font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, tc_ef, f'=SUM({get_column_letter(2)}{r}:{get_column_letter(tc_ef-1)}{r})',
           font=F_DATA, fmt=NUM_FMT)
    last_sam = r

    # Newbuild SAM total row
    r += 1
    wc(ws, r, 1, 'Newbuild SAM ($K)', font=F_TOTAL, border=B_TOT)
    for c in range(2, tc_ef + 1):
        total_cell(ws, r, c,
                   f'=SUM({get_column_letter(c)}{first_sam}:{get_column_letter(c)}{last_sam})')
    span_top_border(ws, r, tc_ef)
    apply_group_borders(ws, svc_r_e, r, 2, n_usn_e, n_cg_e)

    # ── (F) Mekko: Cost Category Mix by Hull ──
    r += 3
    subsec_band(ws, r, f'(F) Newbuild SAM \u2014 Cost Category Mix by Hull (Mekko)', tc_ef)
    r += 1
    svc_r_f = r
    p5c_ordered_f, n_usn_f, n_cg_f = write_svc_header(ws, r, p5c_by_net, d['hull_svc'])

    r += 1
    wc(ws, r, 1, 'Cost Category', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
    for i, hl in enumerate(p5c_ordered_f):
        c = 2 + i
        wc(ws, r, c, hl, font=F_HDR, border=B_HDR); cw(ws, c, 14)
    wc(ws, r, tc_ef, 'Total', font=F_HDR, border=B_HDR)

    n_cats = len(cats)
    tkr_f = r + 1 + n_cats
    tcl_f = get_column_letter(tc_ef)
    fcl_f = get_column_letter(2); lcl_f = get_column_letter(tc_ef - 1)

    # Percentage rows — self-contained P-5c percentage per hull
    for ci, cat in enumerate(cats):
        cr = r + 1 + ci
        wc(ws, cr, 1, cat, font=F_DATA)
        for i, hl in enumerate(p5c_ordered_f):
            c = 2 + i
            p5c_cat = altview_inner('JB_B,1', f'JB_EX,"P-5c"', f'JB_CC,"{cat}"', f'JB_H,"{hl}"')
            p5c_tot = altview_inner('JB_B,1', f'JB_EX,"P-5c"', f'JB_H,"{hl}"')
            wc(ws, cr, c, f'=IFERROR(({p5c_cat})/({p5c_tot}),0)', font=F_PCT, fmt=PCT_FMT)
        wc(ws, cr, tc_ef,
           f'=IFERROR(SUMPRODUCT({fcl_f}{cr}:{lcl_f}{cr},{fcl_f}${tkr_f}:{lcl_f}${tkr_f})/{tcl_f}${tkr_f},0)',
           font=F_PCT, fmt=PCT_FMT)

    # Total ($K) row
    wc(ws, tkr_f, 1, 'Total ($K)', font=F_TOTAL, border=B_TOT)
    for i, hl in enumerate(p5c_ordered_f):
        c = 2 + i
        total_cell(ws, tkr_f, c, cf('JB_B,1', f'JB_H,"{hl}"'))
    total_cell(ws, tkr_f, tc_ef,
               f'=SUM({get_column_letter(2)}{tkr_f}:{get_column_letter(tc_ef-1)}{tkr_f})')
    span_top_border(ws, tkr_f, tc_ef)
    apply_group_borders(ws, svc_r_f, tkr_f, 2, n_usn_f, n_cg_f)

    # Advance r past section (F) data rows
    r = tkr_f

    # ── (G) P-8a System-Level Detail by Hull Program (grid) ──
    p5c_funded_set = set(p5c_funded)
    p8a_sam_hulls = [hl for hl in d.get('p8a_hull_types', []) if hl in p5c_funded_set]

    if p8a_sam_hulls:
        COLS_PER_TABLE = 4
        GAP = 1
        TABLES_PER_ROW = 6
        STRIDE = COLS_PER_TABLE + GAP  # 5
        P8A_CAT_ORDER = [c for c in P5C_COST_CATEGORIES if c in P8A_COST_CATEGORIES]

        # Set column widths for all table positions
        for ti in range(TABLES_PER_ROW):
            c1 = 1 + ti * STRIDE
            cw(ws, c1, 32); cw(ws, c1 + 1, 22); cw(ws, c1 + 2, 14); cw(ws, c1 + 3, 10)

        # Section header (spans full width of first row of tables)
        sec_mc = TABLES_PER_ROW * STRIDE - GAP
        r += 3
        subsec_band(ws, r, f'(G) P-8a System-Level Detail by Hull Program \u2014 {label} ($K)', sec_mc)
        r += 1
        purpose_row(ws, r,
            'Individual GFE systems from Exhibit P-8a (radars, weapons, electronics, propulsion), '
            'grouped by cost category within each hull program. Gross values from SCN justification books.')

        # Pre-compute systems per hull, grouped by cost category
        hull_grouped = {}
        for hl in p8a_sam_hulls:
            by_cat = defaultdict(list)
            for (h, cc, ce), val in d['p8a_hull'].items():
                if h == hl and val > 0:
                    by_cat[cc].append((ce, val))
            groups = []
            for cat in P8A_CAT_ORDER:
                if cat in by_cat:
                    systems = sorted(by_cat[cat], key=lambda x: -x[1])
                    groups.append((cat, systems))
            hull_grouped[hl] = groups

        def _table_height(hl):
            groups = hull_grouped[hl]
            n_cats = len(groups)
            n_sys = sum(len(slist) for _, slist in groups)
            return 2 + n_cats + n_sys + 1  # subsubsec + hdr + cats + systems + total

        # Sort hulls by number of systems descending (longest tables first / leftmost)
        p8a_sam_hulls = sorted(p8a_sam_hulls, key=lambda hl: -_table_height(hl))

        # Lay out in grid: chunks of TABLES_PER_ROW
        for chunk_start in range(0, len(p8a_sam_hulls), TABLES_PER_ROW):
            chunk = p8a_sam_hulls[chunk_start:chunk_start + TABLES_PER_ROW]
            max_h = max(_table_height(hl) for hl in chunk)

            r += 2
            row_start = r

            for ti, hl in enumerate(chunk):
                c1 = 1 + ti * STRIDE

                # Sub-header band (replicate subsubsec_band at offset)
                vt_label = d['hull_type'].get(hl, '')
                hl_text = f'{hl} ({vt_label})' if vt_label else hl
                for c in range(c1, c1 + COLS_PER_TABLE):
                    cell = ws.cell(row=row_start, column=c)
                    cell.fill = BG_GRAY; cell.alignment = _NOWRAP
                wc(ws, row_start, c1, hl_text, font=F_SUBSUBSEC, fill=BG_GRAY)

                # Header row
                hdr_row(ws, row_start + 1, [
                    ('System', 32), ('Cost Category', 22),
                    (f'{label} ($K)', 14), ('% of Hull', 10),
                ], cs=c1)

                # Category groups + system rows
                cr = row_start + 2
                ft = cr  # first data row for % calculation
                cat_subtotal_rows = []  # track category subtotal rows

                for cat, systems in hull_grouped[hl]:
                    # Bold category subtotal row
                    cat_ft = cr + 1  # first system row in this category
                    cat_lt = cr + len(systems)  # last system row
                    cl3 = get_column_letter(c1 + 2)
                    wc(ws, cr, c1, cat, font=F_HDR)
                    wc(ws, cr, c1 + 2,
                       f'=SUM({cl3}{cat_ft}:{cl3}{cat_lt})',
                       font=F_HDR, fmt=NUM_FMT)
                    cat_subtotal_rows.append(cr)
                    cr += 1

                    # Italic system rows
                    for ce, _ in systems:
                        ce_esc = ce.replace('"', '""')
                        formula = altview_inner('JB_B,1', f'JB_EX,"P-8a"', f'JB_CC,"{cat}"',
                                                f'JB_H,"{hl}"', f'JB_CE,"{ce_esc}"')
                        wc(ws, cr, c1, f'  {ce}', font=F_PCT)
                        wc(ws, cr, c1 + 1, cat, font=F_PCT)
                        wc(ws, cr, c1 + 2, '=' + formula, font=F_PCT, fmt=NUM_FMT)
                        cr += 1

                # Total row — immediately after last data row
                total_r = cr
                cl3 = get_column_letter(c1 + 2)
                total_formula = '=' + '+'.join(f'{cl3}{sr}' for sr in cat_subtotal_rows)
                wc(ws, total_r, c1, f'{hl} Total', font=F_TOTAL, border=B_TOT)
                wc(ws, total_r, c1 + 2, total_formula, font=F_TOTAL, fmt=NUM_FMT, border=B_TOT)
                # Top border across all columns
                for c in range(c1, c1 + COLS_PER_TABLE):
                    ws.cell(row=total_r, column=c).border = B_TOT

                # Fill % column for category rows and system rows
                for rn in range(ft, cr):
                    cl_val = get_column_letter(c1 + 2)
                    cl_pct = get_column_letter(c1 + 3)
                    font = F_HDR if rn in cat_subtotal_rows else F_PCT
                    fmt = PCT_FMT
                    wc(ws, rn, c1 + 3,
                       f'=IFERROR({cl_val}{rn}/{cl_val}${total_r},0)',
                       font=font, fmt=fmt)
                pct_total(ws, total_r, c1 + 3, 1.0)

            r = row_start + max_h - 1

    finish_sheet(ws)
    return {'section_d_row': _section_d_row, 'tc_e': tc_e}


def create_mro_tam(wb, d, label, prefix, cf, cf_neg, cf_inner):
    ws = wb.create_sheet(f'{prefix} MRO TAM')
    mc = max(4, 2 + len(d['mro_nz']) + 1)
    cw(ws, 1, 30); cw(ws, 2, 12)

    bkts = d['mro_bkts_with_data']

    r = 1; title_band(ws, r, f'MRO TAM \u2014 {label} ($K)', mc)
    r = 2; purpose_row(ws, r,
        'Total Funding narrowed to all post-construction work — depot maintenance, continuous & emergent repair, '
        'modernization & alterations, major life-cycle events, sustainment engineering, and availability support — '
        'and only oceangoing vessels. Same vessel scope as Newbuild TAM; limited to line items with clear vessel type designation.')

    # (A) Bridge
    r = 4; subsec_band(ws, r, '(A) Bridge: Total MRO Funding \u2192 MRO TAM', 2)
    bridge_start = r + 1
    r += 1; wc(ws, r, 1, 'Total MRO Funding (Buckets 2\u20137)', font=F_DATA)
    wc(ws, r, 2, '=' + '+'.join(cf_inner(f'JB_B,{b}') for b in MRO_BKTS), font=F_DATA, fmt=NUM_FMT)
    r += 1; wc(ws, r, 1, '  Less: Unattributed (fleet-wide MRO items)', font=F_DATA)
    wc(ws, r, 2, f'=-({"+".join(cf_inner(f"JB_B,{b}", "JB_V," + chr(34) + chr(34)) for b in MRO_BKTS)})',
       font=F_DATA, fmt=NUM_FMT)
    for cat in TAM_EXCLUDED_CATEGORIES:
        r += 1; wc(ws, r, 1, f'  Less: {cat}', font=F_DATA)
        wc(ws, r, 2, f'=-({"+".join(cf_inner(f"JB_B,{b}", f"""JB_V,"{cat}" """.strip()) for b in MRO_BKTS)})',
           font=F_DATA, fmt=NUM_FMT)
    r += 1; wc(ws, r, 1, '  Less: No vessel type designation', font=F_DATA)
    novt_f = '+'.join(cf_inner(f'JB_B,{b}', f'JB_V,"{c}"', 'JB_W,""')
                      for b in MRO_BKTS for c in sorted(TAM_CATEGORIES))
    wc(ws, r, 2, f'=-({novt_f})', font=F_DATA, fmt=NUM_FMT)
    bridge_end = r
    r += 1; bridge_tam_row = r
    total_label(ws, r, 1, 'MRO TAM')

    # (B) By Work Type
    r += 3; subsec_band(ws, r, '(B) MRO TAM by Work Type', 3)
    r += 1; hdr_row(ws, r, [('Work Type', 30), (f'{label} ($K)', 12), ('% of TAM', 10)])
    fb = r + 1
    for b in bkts:
        r += 1; wc(ws, r, 1, BUCKET_SHORT.get(b, BUCKET_NAMES[b]), font=F_DATA)
        parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES))
        wc(ws, r, 2, '=' + parts, font=F_DATA, fmt=NUM_FMT)
    lb = r; r += 1; btr = r
    total_label(ws, r, 1, 'MRO TAM')
    total_cell(ws, r, 2, f'=SUM(B{fb}:B{lb})')
    pct_total(ws, r, 3, 1.0)
    span_top_border(ws, r, 3)
    for rn in range(fb, lb + 1):
        wc(ws, rn, 3, f'=IFERROR(B{rn}/B${btr},0)', font=F_PCT, fmt=PCT_FMT)

    # Fill bridge total (self-summing, like Newbuild TAM)
    total_cell(ws, bridge_tam_row, 2, f'=SUM(B{bridge_start}:B{bridge_end})')
    span_top_border(ws, bridge_tam_row, 2)

    # (C) All vessel types
    r += 3; subsec_band(ws, r, '(C) MRO TAM by Vessel Type', 4)
    all_types = svc_order(d['all_tam_mro_sorted'], d['type_svc'])
    r += 1; hdr_row(ws, r, [('Vessel Type', 30), ('Service', 8), (f'{label} ($K)', 12), ('% of TAM', 10)])
    ft = r + 1
    for vt in all_types:
        r += 1; wc(ws, r, 1, vt, font=F_DATA)
        wc(ws, r, 2, d['type_svc'].get(vt, ''), font=F_DATA)
        mparts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_W,"{vt}"') for b in MRO_BKTS)
        wc(ws, r, 3, '=' + mparts, font=F_DATA, fmt=NUM_FMT)
    lt = r; r += 1; ttr = r
    total_label(ws, r, 1, 'MRO TAM'); total_label(ws, r, 2, '')
    total_cell(ws, r, 3, f'=SUM(C{ft}:C{lt})')
    pct_total(ws, r, 4, 1.0)
    span_top_border(ws, r, 4)
    for rn in range(ft, lt + 1):
        wc(ws, rn, 4, f'=IFERROR(C{rn}/C${ttr},0)', font=F_PCT, fmt=PCT_FMT)

    # (D) Mekko
    mk = svc_order(d['mro_nz'], d['type_svc'])
    r += 3; subsec_band(ws, r, '(D) MRO TAM \u2014 Work Type by Vessel Type (Mekko)', 2 + len(mk))
    r += 1; svc_r = r; mk, n_usn, n_cg = write_svc_header(ws, r, d['mro_nz'], d['type_svc'], col_start=2)
    r += 1; wc(ws, r, 1, 'Work Type', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
    for i, vt in enumerate(mk):
        c = 2 + i; wc(ws, r, c, vt, font=F_HDR, border=B_HDR); cw(ws, c, 12)
    tc = 2 + len(mk); wc(ws, r, tc, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc, 12)
    nwt = len(bkts); tkr = r + 1 + nwt

    wc(ws, tkr, 1, 'Total ($K)', font=F_TOTAL, border=B_TOT)
    for i, vt in enumerate(mk):
        c = 2 + i; mparts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_W,"{vt}"') for b in MRO_BKTS)
        total_cell(ws, tkr, c, '=' + mparts)
    total_cell(ws, tkr, tc, f'=SUM({get_column_letter(2)}{tkr}:{get_column_letter(tc-1)}{tkr})')

    tcl = get_column_letter(tc)
    for bi, b in enumerate(bkts):
        wr = r + 1 + bi
        wc(ws, wr, 1, BUCKET_SHORT.get(b, BUCKET_NAMES[b]), font=F_DATA)
        for i, vt in enumerate(mk):
            c = 2 + i; cl = get_column_letter(c)
            wc(ws, wr, c, f'=IFERROR(({cf_inner(f"JB_B,{b}", f"""JB_W,"{vt}" """.strip())})/{cl}${tkr},0)',
               font=F_PCT, fmt=PCT_FMT)
        bkt_parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES))
        wc(ws, wr, tc, f'=IFERROR(({bkt_parts})/{tcl}${tkr},0)', font=F_PCT, fmt=PCT_FMT)

    span_top_border(ws, tkr, tc)
    apply_group_borders(ws, svc_r, tkr, 2, n_usn, n_cg)

    finish_sheet(ws)
    return {'bkt_total_row': btr}


def create_mro_sam(wb, d, label, prefix, cf, cf_neg, cf_inner, mro_tam_info=None):
    ws = wb.create_sheet(f'{prefix} MRO SAM')

    hull_nz = d.get('mro_hull_sam_nz', [])
    sdm_mix = d.get('sdm_cost_mix', [])  # [(title, $K), ...] sorted desc
    mc = max(8, 2 + len(hull_nz) + 1)
    cw(ws, 1, 30); cw(ws, 2, 12)

    tam_sheet = f'{prefix} MRO TAM'
    n_wt = len(MRO_SAM_BKTS)  # 4 work-type columns

    # ── Title & Purpose ──
    r = 1; title_band(ws, r, f'MRO SAM \u2014 {label} ($K)', mc)
    r = 2; purpose_row(ws, r,
        'Outsourceable MRO sized at hull-program level. Starts from MRO TAM, '
        'excludes single-yard/nuclear programs (Submarines, Aircraft Carriers, UUVs) '
        'and non-outsourceable work types (sustainment engineering, availability support). '
        'Retains scheduled depot maintenance, continuous/emergent maintenance, modernization, '
        'and major life-cycle events \u2014 the work a company could compete for. '
        'SDM cost-element estimation applies aggregate OMN Ship Maintenance cost mix '
        'to per-hull SDM totals as a directional approximation.')

    # ── (A) Bridge: MRO TAM → MRO SAM ──
    r = 4; subsec_band(ws, r, '(A) Bridge: MRO TAM \u2192 MRO SAM', 2)
    r += 1; bridge_first = r
    tam_ref_row = mro_tam_info['bkt_total_row'] if mro_tam_info else 18
    wc(ws, r, 1, 'MRO TAM', font=F_DATA)
    wc(ws, r, 2, f"='{tam_sheet}'!B{tam_ref_row}", font=F_GREEN, fmt=NUM_FMT)

    excl_rows = []
    for vt in ['Submarines', 'Aircraft Carriers', 'Unmanned Undersea Vehicles']:
        r += 1; excl_rows.append(r)
        wc(ws, r, 1, f'  Less: {vt}', font=F_DATA)
        parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_W,"{vt}"') for b in MRO_BKTS)
        wc(ws, r, 2, f'=-({parts})', font=F_DATA, fmt=NUM_FMT)

    r += 1; nowork_row = r
    wc(ws, r, 1, '  Less: Non-outsourceable work types (Buckets 6, 7)', font=F_DATA)

    r += 1; sam_result_row = r
    total_label(ws, r, 1, 'MRO SAM (all programs)')
    sam_parts = '+'.join(
        cf_inner(f'JB_B,{b}', f'JB_V,"{c}"')
        for b in MRO_SAM_BKTS for c in sorted(TAM_CATEGORIES))
    excl_parts = '+'.join(
        f'-({cf_inner(f"JB_B,{b}", f"""JB_W,"{vt}" """.strip())})'
        for b in MRO_SAM_BKTS
        for vt in ['Submarines', 'Aircraft Carriers', 'Unmanned Undersea Vehicles'])
    total_cell(ws, r, 2, f'={sam_parts}{excl_parts}')
    span_top_border(ws, r, 2)

    excl_sum = '+'.join(f'B{er}' for er in excl_rows)
    wc(ws, nowork_row, 2, f'=-(B{bridge_first}+{excl_sum}-B{sam_result_row})',
       font=F_DATA, fmt=NUM_FMT)

    # ── (B) MRO SAM by Vessel Type ──
    all_types = svc_order(d['all_sam_mro_sorted'], d['type_svc'])
    mc_b = 2 + n_wt + 2  # VT, Svc, 4 work types, Total, %
    r += 3; subsec_band(ws, r, '(B) MRO SAM (all programs) by Vessel Type', mc_b)
    r += 1; hdr_row(ws, r, [
        ('Vessel Type', 30), ('Service', 8),
        ('SDM ($K)', 12), ('Cont/Emerg ($K)', 12), ('Mod ($K)', 12),
        ('Major LC ($K)', 12), ('Total ($K)', 12), ('% of SAM', 10),
    ])
    tot_col_b = 7  # G = Total
    pct_col_b = 8  # H = %
    ft = r + 1
    for vt in all_types:
        r += 1
        wc(ws, r, 1, vt, font=F_DATA)
        wc(ws, r, 2, d['type_svc'].get(vt, ''), font=F_DATA)
        for bi, b in enumerate(MRO_SAM_BKTS):
            wc(ws, r, 3 + bi, cf(f'JB_B,{b}', f'JB_W,"{vt}"'), font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, tot_col_b, f'=SUM(C{r}:F{r})', font=F_DATA, fmt=NUM_FMT)
    lt = r; r += 1; tr = r
    total_label(ws, r, 1, 'MRO SAM (all programs)'); total_label(ws, r, 2, '')
    for c in range(3, tot_col_b + 1):
        total_cell(ws, r, c,
                   f'=SUM({get_column_letter(c)}{ft}:{get_column_letter(c)}{lt})')
    pct_total(ws, r, pct_col_b, 1.0)
    span_top_border(ws, r, pct_col_b)
    for rn in range(ft, lt + 1):
        wc(ws, rn, pct_col_b, f'=IFERROR({get_column_letter(tot_col_b)}{rn}/{get_column_letter(tot_col_b)}${tr},0)',
           font=F_PCT, fmt=PCT_FMT)

    # ── (C) MRO SAM — Work Type by Hull (Mekko) ──
    mk_nz = [h for h in svc_order(hull_nz, d['hull_svc'])
             if sum(d['mro_hull'].get(h, {}).get(b, 0) for b in MRO_SAM_BKTS) > 0]
    mekko_tc = None; mekko_row = None
    if mk_nz:
        tc = 2 + len(mk_nz)
        r += 3
        subsec_band(ws, r,
                    '(C) MRO SAM \u2014 Work Type by Hull (Mekko)',
                    tc)
        r += 1; svc_r = r; mekko_row = svc_r - 1
        mk_nz, n_usn, n_cg = write_svc_header(ws, r, mk_nz, d['hull_svc'],
                                                col_start=2)
        r += 1
        wc(ws, r, 1, 'Work Type', font=F_HDR, border=B_HDR); cw(ws, 1, 30)
        for i, hl in enumerate(mk_nz):
            c = 2 + i
            wc(ws, r, c, hl, font=F_HDR, border=B_HDR); cw(ws, c, 12)
        wc(ws, r, tc, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc, 12)

        # Work-type percentage rows + total $K row
        wt_rows = {}
        for bi, b in enumerate(MRO_SAM_BKTS):
            wt_rows[b] = r + 1 + bi
        tkr = r + 1 + n_wt

        wc(ws, tkr, 1, 'Total ($K)', font=F_TOTAL, border=B_TOT)
        for i, hl in enumerate(mk_nz):
            c = 2 + i
            parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_H,"{hl}"')
                             for b in MRO_SAM_BKTS)
            total_cell(ws, tkr, c, '=' + parts)
        total_cell(ws, tkr, tc,
                   f'=SUM({get_column_letter(2)}{tkr}:'
                   f'{get_column_letter(tc-1)}{tkr})')

        tcl = get_column_letter(tc)
        for bi, b in enumerate(MRO_SAM_BKTS):
            wr = wt_rows[b]
            wc(ws, wr, 1, BUCKET_SHORT[b], font=F_DATA)
            for i, hl in enumerate(mk_nz):
                c = 2 + i; cl = get_column_letter(c)
                wc(ws, wr, c,
                   f'=IFERROR(({cf_inner(f"JB_B,{b}", f"""JB_H,"{hl}" """.strip())})'
                   f'/{cl}${tkr},0)',
                   font=F_PCT, fmt=PCT_FMT)
            # Total column: weighted average
            bkt_parts = '+'.join(
                cf_inner(f'JB_B,{b}', f'JB_H,"{hl}"') for hl in mk_nz)
            wc(ws, wr, tc,
               f'=IFERROR(({bkt_parts})/{tcl}${tkr},0)',
               font=F_PCT, fmt=PCT_FMT)

        span_top_border(ws, tkr, tc)
        apply_group_borders(ws, svc_r, tkr, 2, n_usn, n_cg)
        mekko_tc = tc
        r = tkr  # advance past mekko data rows

    # ── (D) MRO SAM by Hull Program ──
    mc_hull = 3 + n_wt + 2  # Hull, VT, Svc, 4 work types, Total, %
    tr_d = None
    hull_sdm_col = 4  # col D = SDM in section D
    hull_tot_col = 3 + n_wt + 1  # col H = Total
    hull_pct_col = hull_tot_col + 1  # col I = %
    hull_sdm_cells = {}  # hull -> cell ref for SDM column
    if hull_nz:
        r += 3
        subsec_band(ws, r, f'(D) MRO SAM (all programs) by Hull Program \u2014 {label} ($K)', mc_hull)
        r += 1
        purpose_row(ws, r,
            'Same funding as (B) broken down to individual hull programs \u2014 '
            'the level at which depot maintenance, modernization, and life-cycle '
            'line items are traceable.')
        r += 1
        hdr_row(ws, r, [
            ('Hull Program', 16), ('Vessel Type', 24), ('Service', 8),
            ('SDM ($K)', 12), ('Cont/Emerg ($K)', 12), ('Mod ($K)', 12),
            ('Major LC ($K)', 12), ('Total ($K)', 12), ('% of Total', 10),
        ])
        ft_d = r + 1
        for hl in hull_nz:
            r += 1
            wc(ws, r, 1, hl, font=F_DATA)
            wc(ws, r, 2, d['hull_type'].get(hl, ''), font=F_DATA)
            wc(ws, r, 3, d['hull_svc'].get(hl, ''), font=F_DATA)
            for bi, b in enumerate(MRO_SAM_BKTS):
                wc(ws, r, hull_sdm_col + bi,
                   cf(f'JB_B,{b}', f'JB_H,"{hl}"'), font=F_DATA, fmt=NUM_FMT)
            fc = get_column_letter(hull_sdm_col)
            lc = get_column_letter(hull_sdm_col + n_wt - 1)
            wc(ws, r, hull_tot_col, f'=SUM({fc}{r}:{lc}{r})', font=F_DATA, fmt=NUM_FMT)
            hull_sdm_cells[hl] = f'{get_column_letter(hull_sdm_col)}{r}'
        lt_d = r; r += 1; tr_d = r
        total_label(ws, r, 1, 'MRO SAM (all programs)')
        total_label(ws, r, 2, ''); total_label(ws, r, 3, '')
        for c in range(hull_sdm_col, hull_tot_col + 1):
            total_cell(ws, r, c,
                       f'=SUM({get_column_letter(c)}{ft_d}:{get_column_letter(c)}{lt_d})')
        pct_total(ws, r, hull_pct_col, 1.0)
        span_top_border(ws, r, hull_pct_col)
        tcl = get_column_letter(hull_tot_col)
        for rn in range(ft_d, lt_d + 1):
            wc(ws, rn, hull_pct_col,
               f'=IFERROR({tcl}{rn}/{tcl}${tr_d},0)', font=F_PCT, fmt=PCT_FMT)

    # ── (E) Aggregate SDM Cost Breakdown ──
    e_cat_cells = {}  # title -> cell ref for $K value
    e_total_cell = None
    e_section_row = None; f_section_row = None; f_tc = None
    sdm_titles = [(t, v) for t, v in sdm_mix if v != 0]  # keep raw book titles, drop zeros
    if sdm_titles:
        r += 3; e_section_row = r
        subsec_band(ws, r, '(E) Aggregate SDM Cost Breakdown', 3)
        r += 1
        purpose_row(ws, r,
            'Aggregate cost breakdown of OMN Ship Maintenance (Bucket 2) spending by cost type. '
            'Titles are verbatim from OMN justification book sub-components \u2014 not hull-specific. '
            'Used in section (F) to estimate cost-element coverage per hull program.')
        r += 1
        hdr_row(ws, r, [('OMN Sub-Component', 45), ('$K', 14), ('% of SDM', 10)])
        cw(ws, 1, 45)
        ft_e = r + 1
        for title, val in sdm_titles:
            r += 1
            wc(ws, r, 1, title, font=F_DATA)
            wc(ws, r, 2, val, font=F_DATA, fmt=NUM_FMT)
            e_cat_cells[title] = f'B{r}'
        lt_e = r; r += 1
        total_label(ws, r, 1, 'Total')
        total_cell(ws, r, 2, f'=SUM(B{ft_e}:B{lt_e})')
        e_total_cell = f'B{r}'
        pct_total(ws, r, 3, 1.0)
        span_top_border(ws, r, 3)
        # Percentage column
        for rn in range(ft_e, lt_e + 1):
            wc(ws, rn, 3, f'=IFERROR(B{rn}/{e_total_cell},0)',
               font=F_PCT, fmt=PCT_FMT)

    # ── (F) MRO SAM (SDM cost-element estimate) ──
    sdm_hull_nz = [h for h in hull_nz
                   if d['mro_hull'].get(h, {}).get('2', 0) > 0]
    if sdm_hull_nz and e_total_cell and hull_sdm_cells:
        sdm_ordered = svc_order(
            sorted(sdm_hull_nz, key=lambda h: -d['mro_hull'].get(h, {}).get('2', 0)),
            d['hull_svc'])
        tc_f = 2 + len(sdm_ordered)
        r += 3; f_section_row = r; f_tc = tc_f
        subsec_band(ws, r,
                    f'(F) MRO SAM (SDM cost-element estimate) \u2014 {label} ($K)', tc_f)
        r += 1
        purpose_row(ws, r,
            'Estimated SDM cost breakdown by hull program. Applies the aggregate OMN '
            'cost mix from section (E) to each hull\u2019s SDM funding. This is an '
            'approximation \u2014 actual cost mix varies by hull class. Analogous to '
            'Newbuild SAM section (E) which applies P-5c cost mix to net budget authority.')
        r += 1; svc_r_f = r
        sdm_ordered, n_usn_f, n_cg_f = write_svc_header(ws, r, sdm_ordered, d['hull_svc'])

        r += 1
        wc(ws, r, 1, 'OMN Sub-Component', font=F_HDR, border=B_HDR); cw(ws, 1, 45)
        for i, hl in enumerate(sdm_ordered):
            c = 2 + i
            wc(ws, r, c, hl, font=F_HDR, border=B_HDR); cw(ws, c, 14)
        wc(ws, r, tc_f, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc_f, 14)

        # SAM SDM total reference row
        r += 1; sdm_ref_r = r
        wc(ws, r, 1, 'SAM SDM Total ($K)', font=F_KPI, fill=BG_TEAL)
        for i, hl in enumerate(sdm_ordered):
            c = 2 + i
            ref = hull_sdm_cells.get(hl, '0')
            wc(ws, r, c, f'={ref}', font=F_KPI, fill=BG_TEAL, fmt=NUM_FMT)
        wc(ws, r, tc_f,
           f'=SUM({get_column_letter(2)}{r}:{get_column_letter(tc_f-1)}{r})',
           font=F_KPI, fill=BG_TEAL, fmt=NUM_FMT)

        # Allocated rows: hull SDM × (title $K / total $K)
        first_f = r + 1
        for title, _ in sdm_titles:
            r += 1
            wc(ws, r, 1, title, font=F_DATA)
            cat_ref = e_cat_cells[title]
            for i, hl in enumerate(sdm_ordered):
                c = 2 + i; cl = get_column_letter(c)
                wc(ws, r, c,
                   f'=IFERROR({cl}${sdm_ref_r}*{cat_ref}/{e_total_cell},0)',
                   font=F_DATA, fmt=NUM_FMT)
            wc(ws, r, tc_f,
               f'=SUM({get_column_letter(2)}{r}:{get_column_letter(tc_f-1)}{r})',
               font=F_DATA, fmt=NUM_FMT)
        last_f = r

        # Estimated SDM total row
        r += 1
        wc(ws, r, 1, 'Estimated SDM Total ($K)', font=F_TOTAL, border=B_TOT)
        for c in range(2, tc_f + 1):
            total_cell(ws, r, c,
                       f'=SUM({get_column_letter(c)}{first_f}:{get_column_letter(c)}{last_f})')
        span_top_border(ws, r, tc_f)
        apply_group_borders(ws, svc_r_f, r, 2, n_usn_f, n_cg_f)

    # ── (G) MRO SAM (contract-addressable) ──
    CONTRACT_TITLES = ['Ship Maintenance By Contract', 'Equipment Maintenance By Contract']
    g_section_row = None; g_tc = None
    contract_refs = [e_cat_cells[t] for t in CONTRACT_TITLES if t in e_cat_cells]
    # Need: hulls with any SAM bucket data, contract refs from (E), hull SDM cells from (D)
    g_hull_nz = [h for h in svc_order(
        sorted(hull_nz, key=lambda h: -sum(d['mro_hull'].get(h, {}).get(b, 0) for b in MRO_SAM_BKTS)),
        d['hull_svc'])
        if sum(d['mro_hull'].get(h, {}).get(b, 0) for b in MRO_SAM_BKTS) > 0]
    if g_hull_nz and contract_refs and e_total_cell and hull_sdm_cells:
        tc_g = 2 + len(g_hull_nz)
        r += 3; g_section_row = r; g_tc = tc_g
        subsec_band(ws, r,
                    f'(G) MRO SAM (contract-addressable) \u2014 {label} ($K)', tc_g)
        r += 1
        purpose_row(ws, r,
            'MRO SAM narrowed to the contract-addressable market. SDM is reduced to only '
            'the private-sector contract portion (Ship Maintenance By Contract + Equipment '
            'Maintenance By Contract, ~31% of SDM). Cont/Emerg, Mod, and Major LC are '
            'retained at full hull-level totals. This is the MRO market a company could '
            'realistically compete for.')
        r += 1; svc_r_g = r
        g_hull_nz, n_usn_g, n_cg_g = write_svc_header(ws, r, g_hull_nz, d['hull_svc'])

        r += 1
        wc(ws, r, 1, 'Work Type', font=F_HDR, border=B_HDR); cw(ws, 1, 45)
        for i, hl in enumerate(g_hull_nz):
            c = 2 + i
            wc(ws, r, c, hl, font=F_HDR, border=B_HDR); cw(ws, c, 14)
        wc(ws, r, tc_g, 'Total', font=F_HDR, border=B_HDR); cw(ws, tc_g, 14)

        contract_sum = '+'.join(contract_refs)
        first_g = r + 1

        # SDM — Contract row
        r += 1
        wc(ws, r, 1, 'SDM \u2014 Contract ($K)', font=F_DATA)
        for i, hl in enumerate(g_hull_nz):
            c = 2 + i
            sdm_ref = hull_sdm_cells.get(hl, '0')
            wc(ws, r, c,
               f'=IFERROR({sdm_ref}*({contract_sum})/{e_total_cell},0)',
               font=F_DATA, fmt=NUM_FMT)
        wc(ws, r, tc_g,
           f'=SUM({get_column_letter(2)}{r}:{get_column_letter(tc_g-1)}{r})',
           font=F_DATA, fmt=NUM_FMT)

        # Cont/Emerg, Mod, Major LC — full hull amounts
        for b, lbl in [('3', 'Cont/Emerg ($K)'), ('4', 'Mod ($K)'), ('5', 'Major LC ($K)')]:
            r += 1
            wc(ws, r, 1, lbl, font=F_DATA)
            for i, hl in enumerate(g_hull_nz):
                c = 2 + i
                wc(ws, r, c, cf(f'JB_B,{b}', f'JB_H,"{hl}"'), font=F_DATA, fmt=NUM_FMT)
            wc(ws, r, tc_g,
               f'=SUM({get_column_letter(2)}{r}:{get_column_letter(tc_g-1)}{r})',
               font=F_DATA, fmt=NUM_FMT)
        last_g = r

        # Total row
        r += 1
        wc(ws, r, 1, 'MRO SAM (contract-addressable)', font=F_TOTAL, border=B_TOT)
        for c in range(2, tc_g + 1):
            total_cell(ws, r, c,
                       f'=SUM({get_column_letter(c)}{first_g}:{get_column_letter(c)}{last_g})')
        span_top_border(ws, r, tc_g)
        apply_group_borders(ws, svc_r_g, r, 2, n_usn_g, n_cg_g)

    finish_sheet(ws)
    return {'mc': mc,
            'mekko_tc': mekko_tc,
            'mekko_row': mekko_row,
            'e_section_row': e_section_row,
            'f_section_row': f_section_row,
            'f_tc': f_tc,
            'g_section_row': g_section_row,
            'g_tc': g_tc,
            'sdm_excluded': d.get('sdm_excluded', [])}




# ── Competitive Dynamics ──────────────────────────────────────

# Data extracted from analysis/FY2026_Key_Programs_Contract_Awards.md
# Values in $M unless noted otherwise.

_PRIMES = [
    (1, 'Huntington Ingalls Industries', 'HII', 22000000, 'DDG-51 $11.4B, LPD $4.7B, LHA $4.5B, DDG-1000 $0.3B, depot'),
    (2, 'Raytheon / RTX', 'RTX Corp', 13700000, 'Standard Missile $5.7B, SPY-6 $3.5B, CIWS/RAM $2.0B, DDG-1000 $0.9B, CEC $0.6B'),
    (3, 'Bath Iron Works', 'General Dynamics', 8800000, 'DDG-51 FY23-27 $5.0B, FY18-22 $3.3B, FY13-17 $0.5B, DDG-1000 $0.2B'),
    (4, 'BAE Systems', 'BAE Systems plc', 5000000, 'DDG/LHD/LPD/LSD/CG depot maintenance — Norfolk, San Diego, Jacksonville, Hawaii'),
    (5, 'Lockheed Martin', 'LM', 1500000, 'SLQ-32(V)6 $0.8B, SSDS $0.4B, HELIOS $0.1B, LCS Freedom $0.3B'),
    (6, 'Metro Machine Corp', '--', 700000, 'LHD/DDG/LSD depot maintenance — Norfolk'),
    (7, 'NASSCO', 'General Dynamics', 600000, 'LHA/LHD/LSD/DDG/CG depot maintenance — San Diego'),
    (8, 'Vigor Marine LLC', '--', 500000, 'DDG/CG depot maintenance — Portland, OR'),
    (9, 'Northrop Grumman Systems', 'NG', 487000, 'Knifefish/LCS $0.35B, SEWIP Block 3 $0.14B, ALMDS $0.07B'),
    (10, 'QED Systems', '--', 300000, 'Third-party planning DDG/CG/LPD/LHD'),
    (11, 'Continental Maritime', '--', 300000, 'DDG depot — San Diego'),
    (12, 'Textron Systems', 'Textron', 242000, 'UISS $0.24B'),
    (13, 'Advanced Technology International', '--', 235000, 'OTA consortium: MK-54 MOD 2, SPY-6 prototype, HEL, LHD midlife'),
    (14, 'General Dynamics Mission Systems', 'General Dynamics', 210000, 'MK-54 MOD 1 $0.12B, Knifefish Block 1 $0.08B'),
    (15, 'SAIC', 'SAIC', 200000, 'DDG-1000 engineering, SSTD Nixie, torpedo defense'),
    (16, 'Marine Hydraulics International', '--', 200000, 'DDG/LSD depot + sub-prime on LHD primes'),
    (17, 'Bollinger Shipyards', 'Bollinger', 116000, 'MCM USV production'),
    (18, 'Ultra Electronics Ocean Systems', 'Ultra', 68000, 'MK 54 MOD 0 array kits + Nixie production'),
    (19, 'Penn State Applied Research Lab', 'Penn State', 60000, 'MK-54 MOD 2 engineering, HEL development'),
]

_SUBS = [
    (1, 'Fincantieri (Marinette Marine)', 2487000, 2, 1, 'LCS Freedom hull builder — actual constructor under LM prime'),
    (2, 'Timken Gears & Services', 516000, 22, 20, 'DDG/LHD/CG propulsion gearing'),
    (3, 'General Dynamics (rolled up)', 432000, 23, 14, 'GD Mission Systems $359M on SPY-6, GD-OTS, GDIT'),
    (4, 'L3Harris (Aerojet Rocketdyne)', 378000, 4, 4, 'Solid rocket motors for Standard Missile family'),
    (5, 'Rolls-Royce', 313000, 25, 25, 'Gas turbines, waterjets, gearing — DDG/LCS/LPD/LHA + depot'),
    (6, 'CAES (Cobham)', 217000, 10, 8, 'RF assemblies on SPY-6 $170M, SEWIP $110M, SLQ-32 $95M'),
    (7, 'Northrop Grumman (as sub)', 215000, 10, 9, 'SPY-6 RF $114M, CIWS, LPD/LHA construction'),
    (8, 'IMIA', 204000, 28, 28, 'Hull coatings / paint across 28 depot availabilities'),
    (9, 'L3Harris (excl. Aerojet)', 203000, 53, 45, 'Maritime Power, Cincinnati Electronics, Fuzing — broadest footprint'),
    (10, 'Johnson Controls (incl. York)', 172000, 31, 24, 'HVAC / chilled water on DDG/LPD/LHA + depot'),
    (11, 'GE Aerospace', 166000, 16, 14, 'LM2500 gas turbines on DDG-51 + GE Power Conversion LHA 9'),
    (12, 'Fairbanks Morse Defense', 161000, 25, 18, 'Diesel generators / propulsion on LPD/LHA + depot rebuilds'),
    (13, 'Leonardo DRS', 134000, 18, 15, 'Naval Power Systems, Network & Imaging, Laurel Technologies'),
    (14, 'Sparton DeLeon Springs', 108000, 2, 2, 'ASW sensors / payload on Raytheon Barracuda'),
    (15, 'RAM-System GmbH', 98000, 1, 1, 'RAM missile co-development with Raytheon (Germany)'),
    (16, 'Bay Metals & Fabrication', 93000, 9, 9, 'Norfolk metal fab serving 9 depot availabilities'),
    (17, 'US Joiner LLC', 88000, 12, 12, 'Interior outfitting — LPD construction + depot'),
    (18, 'United Rentals', 77000, 29, 29, 'Equipment rental across 29 depot availabilities'),
    (19, 'Honeywell', 71000, 12, 12, 'IMUs / IRUs on Standard Missile + CIWS'),
    (20, 'Anaren', 64000, 7, 7, 'RF combiners/dividers on SPY-6'),
]

_PROGRAMS = [
    ('DDG-51 New Construction', 530000, 'GE Aerospace $88M', 'Johnson Controls $76M', 'Timken Gears $41M'),
    ('DDG Mod — SPY-6', 1220000, 'GD Mission Systems $359M', 'CAES $170M', 'Northrop Grumman $114M'),
    ('DDG Mod — CIWS/RAM/SeaRAM', 560000, 'RAM-System GmbH $98M', 'L3Harris $39M', 'GD $35M'),
    ('LCS Freedom Construction', 2910000, 'Marinette Marine $2,487M', 'Rolls-Royce $196M', 'Airbus US $56M'),
    ('LCS MCM Mission Modules', 1230000, 'Trident Sensors $215M', 'Teledyne $55M', 'GD $13M'),
    ('LHA New Construction', 340000, 'Fairbanks Morse $40M', 'L3Harris $39M', 'Leonardo DRS $34M'),
    ('LPD Flight II New Construction', 340000, 'Fairbanks Morse $65M', 'US Joiner $36M', 'Caterpillar $20M'),
    ('Standard Missile Family', 930000, 'L3Harris (Aerojet) $390M', 'Honeywell $43M', 'Goodrich $32M'),
    ('DDG-1000', 130000, 'Red River Technology $61M', 'Air Masters $6M', '--'),
    ('DDG Mod — CEC', 30000, 'Sechan Electronics $7M', 'Action Electronics $5M', '--'),
    ('DDG Mod — SSDS', 40000, 'Mission Solutions $14M', 'PMAT $6M', '--'),
    ('DDG Mod — HELIOS', 60000, 'MZA Associates $25M', 'L3 Technologies $22M', '--'),
    ('Airborne MCM', 150000, 'Sparton $108M', 'Teledyne $17M', '--'),
    ('MK-54 Torpedo', 30000, 'L3Harris (Aerojet) $25M', 'J&E Precision Tool $3M', '--'),
    ('Depot Maintenance', 1500000, 'IMIA $204M', 'Bay Metals $93M', 'US Joiner $88M'),
]

_CROSSWALK = {
    'LHD': [
        ('OMN_Vol2 LHD/AMPHIBS depot', 346443, 'BAE Norfolk/SD, Metro Machine, NASSCO'),
        ('OMN_Vol2 LHD/AMPHIBS modernization', 631859, 'BAE Norfolk/SD, Metro Machine'),
        ('OPN BA1 Line 8 LHA/LHD Midlife', 123384, 'BAE / Metro / Gibbs & Cox'),
    ],
    'LPD': [
        ('SCN 3010 LPD Flight II procurement', 835037, 'HII — N0002424C2473 ($5.80B ceiling)'),
        ('SCN 3010 LPD Flight II AP', 275000, 'HII — N0002424C2473 AP scope'),
        ('OMN_Vol2 LPD depot', 149595, 'BAE Norfolk/SD'),
        ('OPN BA1 Line 15 LPD Class Support', 125542, 'HII Class Eng ($225M) + Raytheon LCE&S ($485M)'),
    ],
    'LSD': [
        ('OMN_Vol2 LSD depot', 94421, 'NASSCO, BAE, Metro Machine'),
        ('OMN_Vol2 LSD modernization', 25931, 'Alteration installation (small)'),
    ],
    'LHA': [
        ('SCN 3041 LHA Replacement', 634963, 'HII — N0002420C2437 LHA 9 ($3.14B ceiling)'),
        ('OMN_Vol2 LHA depot', 45012, 'NASSCO — LHA 6 DSRA ($198M ceiling)'),
    ],
    'DDG': [
        ('SCN 2122 DDG-51 procurement', 5410773, 'HII FY23-27 MYP + BIW FY23-27 MYP'),
        ('SCN 2122 DDG-51 AP', 1750000, 'HII + BIW — AP against FY23-27 MYPs'),
        ('SCN 2119 DDG-1000', 52358, 'HII Mod Planning + BIW Planning Yard'),
        ('OMN_Vol2 DDG depot', 952000, 'BAE / NASSCO / Continental / Vigor / MHI / Metro'),
        ('OMN_Vol2 DDG modernization', 1282612, 'Cross-vehicle alteration installation'),
        ('WPN 2234 Standard Missile', 1008875, 'Raytheon SM production vehicles'),
        ('OPN BA1 Line 5 DDG Mod', 878787, 'SPY-6, SEWIP, SSDS, CEC, CIWS, HELIOS, Aegis kits'),
        ('WPN 3215 MK-54 Torpedo Mods', 128513, 'GDMS MK 54 MOD 1 + Ultra MK 54 MOD 0 + ATI OTAs'),
        ('OPN BA1 Line 16 DDG-1000 Support', 115267, 'Raytheon DDG-1000 mission sys + BIW Planning Yard'),
    ],
    'LCS': [
        ('OMN_Vol2 LCS depot', 576389, 'LM Freedom ISEA + GDMS Independence ISEA + avails'),
        ('OMN_Vol2 LCS modernization', 383361, 'LM + GDMS combat system ISEA + ITT-2A'),
        ('OPN BA1 Line 37 LCS In-Service Mod', 189458, 'LM, GDMS, Textron, BAH PEO LCS support'),
        ('OPN BA1 Line 34 LCS MCM Modules', 91372, 'UISS (Textron), Knifefish (GDMS/NG), MCM USV (Bollinger)'),
        ('WPN 2292 Naval Strike Missile', 32238, 'Raytheon NSM integration + Kongsberg OEM via FMS'),
    ],
    'CG': [
        ('OMN_Vol2 CG depot', 115501, 'Vigor, BAE Norfolk, NASSCO, HII CG planning'),
        ('OMN_Vol2 CG modernization', 44940, 'Same vendor pool — alteration installation'),
    ],
}


def create_competitive_dynamics(wb):
    ws = wb.create_sheet('Competitive Dynamics')
    ws.sheet_properties.tabColor = '7B5B3A'  # muted brown — distinct from FY26 blue / FY27 green

    # Column widths
    cw(ws, 1, 30)  # A: Contractor / Program / SAM Line
    cw(ws, 2, 16)  # B: Parent / Total Subs / FY26 $K
    cw(ws, 3, 12)  # C: Window $M / Sub Total / Top Sub 1
    cw(ws, 4, 26)  # D: Key Programs / # Primes / Top Sub 2 / Vehicle
    cw(ws, 5, 26)  # E: # PIIDs / Top Sub 3

    r = 1; title_band(ws, r, 'Competitive Dynamics', 5)
    r = 2; purpose_row(ws, r, 'FY20-26 contract obligation analysis — primes, subcontractors, and FY26 vehicle crosswalk ($K)')

    # ── Section A: Top Prime Contractors ──
    r = 4; subsec_band(ws, r, '(A) Top Prime Contractors — FY20-26 Window Obligation', 4)
    r += 1; hdr_row(ws, r, [('Contractor', 30), ('Parent', 16), ('Window ($K)', 12), ('Key Programs', 44)])
    ft = r + 1
    for rank, name, parent, val, programs in _PRIMES:
        r += 1
        wc(ws, r, 1, name, font=F_DATA)
        wc(ws, r, 2, parent, font=F_GRAY)
        wc(ws, r, 3, val, font=F_BLUE, fmt=NUM_FMT)
        wc(ws, r, 4, programs, font=F_GRAY)
    r += 1; total_label(ws, r, 1, 'Total (top 19)')
    total_cell(ws, r, 3, f'=SUM(C{ft}:C{r-1})')
    span_top_border(ws, r, 4)

    # ── Section B: Top Subcontractors ──
    r += 3; subsec_band(ws, r, '(B) Top Subcontractors — Hidden Supply Chain', 5)
    r += 1; hdr_row(ws, r, [('Parent Company', 30), ('Sub Total ($K)', 16), ('# Primes', 12), ('# PIIDs', 10), ('Primary Programs', 44)])
    ft = r + 1
    for rank, name, val, pairs, piids, programs in _SUBS:
        r += 1
        wc(ws, r, 1, name, font=F_DATA)
        wc(ws, r, 2, val, font=F_BLUE, fmt=NUM_FMT)
        wc(ws, r, 3, pairs, font=F_BLUE, fmt=NUM_FMT)
        wc(ws, r, 4, piids, font=F_BLUE, fmt=NUM_FMT)
        wc(ws, r, 5, programs, font=F_GRAY)
    r += 1; total_label(ws, r, 1, 'Total (top 20)')
    total_cell(ws, r, 2, f'=SUM(B{ft}:B{r-1})')
    span_top_border(ws, r, 5)

    # ── Section C: Per-Program Subaward Summary ──
    r += 3; subsec_band(ws, r, '(C) Per-Program Subaward Summary', 5)
    r += 1; hdr_row(ws, r, [('Program', 30), ('Total Subs ($K)', 16), ('Top Sub 1', 26), ('Top Sub 2', 26), ('Top Sub 3', 26)])
    ft = r + 1
    for prog, val, s1, s2, s3 in _PROGRAMS:
        r += 1
        wc(ws, r, 1, prog, font=F_DATA)
        wc(ws, r, 2, val, font=F_BLUE, fmt=NUM_FMT)
        wc(ws, r, 3, s1, font=F_GRAY)
        wc(ws, r, 4, s2, font=F_GRAY)
        wc(ws, r, 5, s3, font=F_GRAY)
    r += 1; total_label(ws, r, 1, 'Total')
    total_cell(ws, r, 2, f'=SUM(B{ft}:B{r-1})')
    span_top_border(ws, r, 5)

    # ── Section D: FY26 Contract Vehicle Crosswalk ──
    r += 3; subsec_band(ws, r, '(D) FY2026 SAM Line Item to Contract Vehicle Crosswalk', 3)

    for hull, lines in _CROSSWALK.items():
        r += 2
        # Hull sub-header (sub-sub-section style) — 3 cols only
        for c in range(1, 4):
            cell = ws.cell(row=r, column=c)
            cell.fill = BG_GRAY; cell.alignment = _NOWRAP
        wc(ws, r, 1, hull, font=F_SUBSUBSEC, fill=BG_GRAY)

        r += 1; hdr_row(ws, r, [('FY26 SAM Line', 30), ('FY26 ($K)', 16), ('Contract Vehicle / Prime', 44)])
        ft = r + 1
        for line, val, vehicle in lines:
            r += 1
            wc(ws, r, 1, line, font=F_DATA)
            wc(ws, r, 2, val, font=F_BLUE, fmt=NUM_FMT)
            wc(ws, r, 3, vehicle, font=F_GRAY)
        r += 1; total_label(ws, r, 1, f'{hull} Total')
        total_cell(ws, r, 2, f'=SUM(B{ft}:B{r-1})')
        span_top_border(ws, r, 3)

    finish_sheet(ws)






# ── Named ranges ──────────────────────────────────────────────

def add_helper_columns(wb):
    """Extend helper-column formulas to cover the full named-range span (rows 6-4000).

    The data sheet (data_v2.xlsx) already contains the helper columns at AC-AE
    with headers and formulas for actual data rows.  This function fills any
    remaining rows up to 4000 so the named ranges cover the full range.

    Column layout (inserted between FY27 dollars and Reference):
      AC (29) — Row Class: "ADD" for additive, "ALT" for ALT_VIEW, blank otherwise
      AD (30) — FY26 Best ($K): cascade DAA Enacted > Total > Request
      AE (31) — FY27 Best ($K): cascade Total > Request
    """
    ws = wb['J Book Items Cons.']
    RC_COL, BV_START = 29, 30  # AC=29, AD=30

    for r in range(6, 4001):
        if ws.cell(row=r, column=RC_COL).value is None:
            ws.cell(row=r, column=RC_COL,
                    value=f'=IF(OR(E{r}="",E{r}="[PARENT]"),"ADD",IF(E{r}="[ALT_VIEW]","ALT",""))')
        for i, fyc in enumerate(FY_CONFIGS):
            c = BV_START + i
            if ws.cell(row=r, column=c).value is None:
                ws.cell(row=r, column=c,
                        value=f'={fyc["best_value_formula"].format(r=r)}')


def create_named_ranges(wb):
    JB = "'J Book Items Cons.'"
    shared = [
        ('JB_A', 'A'),     # Source Book (col 1)
        ('JB_B', 'I'),     # Bucket (col 9)
        ('JB_F', 'J'),     # Bucket Sub Category (col 10)
        ('JB_H', 'N'),     # Vessel Hull (col 14)
        ('JB_S', 'E'),     # Row Type (col 5)
        ('JB_U', 'K'),     # Vessel Service (col 11)
        ('JB_V', 'L'),     # Vessel Category (col 12)
        ('JB_W', 'M'),     # Vessel Type (col 13)
        ('JB_CE', 'AJ'),   # Cost Element (col 36 — shifted +3 by helper insert)
        ('JB_EX', 'G'),    # Exhibit Level (col 7)
        ('JB_CC', 'H'),    # Cost Category (col 8)
        ('JB_RC', 'AC'),   # Row Class helper (col 29)
    ]
    # Best-value helper columns: AD, AE, ...
    bv_col_letters = ['AD', 'AE', 'AF', 'AG']  # room for future FYs
    for i, fyc in enumerate(FY_CONFIGS):
        shared.append((fyc['best_value_range'], bv_col_letters[i]))

    fy_ranges = []
    for fyc in FY_CONFIGS:
        fy_ranges.extend(fyc['named_ranges'])

    for nm, col in shared + fy_ranges:
        if nm in wb.defined_names:
            del wb.defined_names[nm]
        wb.defined_names.add(DefinedName(nm, attr_text=f"{JB}!${col}$6:${col}$4000"))




# ── Validation sheet (see build/validation_sheet.py) ─────────




# ── Main ──────────────────────────────────────────────────────

def main():
    out_path = next_output_path()
    print(f'Source: {WORKBOOK_SRC}')
    print(f'Output: {out_path}')
    shutil.copy2(WORKBOOK_SRC, out_path)

    wb = openpyxl.load_workbook(out_path)

    # Strip all non-base sheets
    for nm in list(wb.sheetnames):
        if nm not in BASE_SHEETS:
            del wb[nm]
    print(f'Base sheets kept: {wb.sheetnames}')

    # Helper columns (Row Class, FY best-value) + named ranges
    add_helper_columns(wb)
    create_named_ranges(wb)

    # Build each fiscal year
    generated = []
    fy26_sam_info = None; fy26_mro_sam_info = None
    for fyc in FY_CONFIGS:
        label  = fyc['label']
        prefix = fyc['prefix']
        tab_color = fyc['tab_color']
        print(f'\n--- {label} ---')

        d = read_data(WORKBOOK_SRC, fyc['read_indices'])
        print(f'  TAM types: {len(d["all_tam_types"])}, SAM types: {len(d["all_sam_types"])}')
        print(f'  NB non-zero: {len(d["nb_nz"])}, NB SAM non-zero: {len(d["nb_sam_nz"])}')
        print(f'  MRO non-zero: {len(d["mro_nz"])}, MRO SAM non-zero: {len(d["mro_sam_nz"])}')
        print(f'  MRO SAM hulls: {len(d["mro_hull_sam_nz"])}')
        print(f'  MRO buckets with data: {d["mro_bkts_with_data"]}')
        print(f'  NB hulls non-zero: {len(d["nb_hull_nz"])}, P-5c hulls: {len(d["p5c_hull_types"])}, P-8a hulls: {len(d["p8a_hull_types"])}')

        # Divider tab
        divider_name = f'{label} >>'
        div_ws = wb.create_sheet(divider_name)
        div_ws.sheet_properties.tabColor = tab_color
        band(div_ws, 1, 6, label, F_TITLE, BG_BLACK)
        div_ws.sheet_view.showGridLines = False
        div_ws.row_dimensions[1].height = _RH

        cf, cf_neg, cf_inner = make_cascade(fyc['best_value_range'])
        acf, acf_inner = make_altview_cascade(fyc['best_value_range'])

        args = (wb, d, label, prefix, cf, cf_neg, cf_inner)
        create_total_funding(*args)
        create_newbuild_tam(*args)
        sam_info = create_newbuild_sam(*args, altview_cf=acf, altview_inner=acf_inner)
        mro_tam_info = create_mro_tam(*args)
        mro_sam_info = create_mro_sam(*args, mro_tam_info=mro_tam_info)

        if prefix == 'FY26':
            fy26_sam_info = sam_info
            fy26_mro_sam_info = mro_sam_info

        # Apply tab colors to all sheets in this FY group
        fy_sheets = [
            f'{prefix} Total Funding', f'{prefix} Newbuild TAM',
            f'{prefix} Newbuild SAM',  f'{prefix} MRO TAM',
            f'{prefix} MRO SAM',
        ]
        for sn in fy_sheets:
            wb[sn].sheet_properties.tabColor = tab_color

        generated.append(divider_name)
        generated.extend([
            f'{prefix} Total Funding', f'{prefix} Newbuild TAM',
            f'{prefix} Newbuild SAM',
            f'{prefix} MRO TAM',
            f'{prefix} MRO SAM',
        ])

    # Competitive Dynamics
    print('\n--- Competitive Dynamics ---')
    create_competitive_dynamics(wb)
    generated.append('Competitive Dynamics')

    # Validation sheet (built programmatically — see validation_sheet.py)
    from validation_sheet import create_validation_sheet
    print('\n--- Validation ---')
    create_validation_sheet(wb)
    generated.append('Validation')

    # Order: Sheet 1, then generated sheets in FY order, then remaining base sheets
    base_first = 'J Book Items Cons.'
    base_rest  = [s for s in wb.sheetnames if s in BASE_SHEETS and s != base_first]
    ordered    = [base_first] + generated + base_rest
    wb._sheets = [wb[s] for s in ordered]

    wb.save(out_path)

    # ── Post-save: inject sticky-note annotations via OOXML ──
    annotations = []

    if fy26_sam_info and fy26_sam_info.get('section_d_row'):
        er = fy26_sam_info['section_d_row'] - 1  # convert to 0-based
        tc = fy26_sam_info['tc_e']                # 1-based max col → 0-based "next col"
        annotations.append({
            'sheet': 'FY26 Newbuild SAM',
            'from_col': tc + 1, 'from_row': er,
            'to_col': tc + 4, 'to_row': er + 14,
            'lines': [
                '**P-5c Gross vs Net Budget Authority**',
                '',
                'P-5c exhibits report gross total-ship-estimate',
                'costs — what the ship actually costs to build,',
                'broken by component. These gross costs are',
                'larger than net budget authority because they',
                'include money that was funded through advance',
                'procurement in prior fiscal years.',
                '',
                '_Section (E) solves this: it takes the P-5c_',
                '_percentage mix and applies it to net BA,_',
                '_so cost categories reconcile to actual_',
                '_FY2026 funding._',
            ],
        })
    # Section (E) annotation — aggregate SDM cost breakdown
    if fy26_mro_sam_info and fy26_mro_sam_info.get('e_section_row'):
        dr = fy26_mro_sam_info['e_section_row'] - 1 + 2  # 0-based, +2 to clear purpose row
        excl = fy26_mro_sam_info.get('sdm_excluded', [])
        excl_lines = []
        if excl:
            excl_lines = [
                '',
                '**Excluded from this table:**',
                '',
            ]
            for title, val in excl:
                excl_lines.append(f'\u2022 {title}')
            excl_lines += [
                '',
                'These name specific hull programs in their',
                'titles, so applying them proportionally',
                'across all hulls would misattribute spending.',
                'They still count in SAM totals (sections',
                'A/B/C) via the OMN [PARENT] row.',
            ]
        annotations.append({
            'sheet': 'FY26 MRO SAM',
            'from_col': 3, 'from_row': dr,
            'from_col_off': 457200,
            'cx': 3657600,            # ~4 inches — wider to avoid wrapping excluded titles
            'cy': 3200400,            # ~3.5 inches — shorter to clear purpose row
            'lines': [
                '**What this table shows**',
                '',
                'The OMN justification book breaks Ship',
                'Maintenance (Bucket 2) into sub-components',
                '\u2014 labor, contracts, materials, weapons',
                'depot, MSC maintenance, etc. These are the',
                'raw book titles showing how SDM dollars are',
                'spent across the fleet, sorted by size.',
                '',
                'This is aggregate data (not per-hull) \u2014 it',
                'covers all vessels maintained under OMN,',
                'including SAM-excluded types. Section (E)',
                'applies these percentages to SAM-eligible',
                'hull SDM totals in section (F).',
            ] + excl_lines,
        })

    # Section (F) annotation — SDM cost-element estimate
    if fy26_mro_sam_info and fy26_mro_sam_info.get('f_section_row'):
        er = fy26_mro_sam_info['f_section_row'] - 1  # 0-based
        etc = fy26_mro_sam_info.get('f_tc', 4)
        annotations.append({
            'sheet': 'FY26 MRO SAM',
            'from_col': etc + 1, 'from_row': er,
            'from_col_off': 457200,
            'lines': [
                '**How this table works**',
                '',
                'Each hull\u2019s SDM total from section (D) is',
                'multiplied by the percentage mix from (E).',
                'For example, if 25% of aggregate SDM goes',
                'to private-sector contracts, DDG\u2019s $960M SDM',
                'is estimated as ~$240M contract work.',
                '',
                'Unlike Newbuild SAM where P-5c gives a',
                '_different_ cost mix per hull, MRO applies',
                'the _same_ aggregate mix to every hull.',
                'This is a directional approximation \u2014',
                'actual cost structure varies by hull class.',
                '',
                '_Estimated SDM Total reconciles back to_',
                '_each hull\u2019s SDM from section (D)._',
            ],
        })

    # Section (F) mekko annotation
    if fy26_mro_sam_info and fy26_mro_sam_info.get('mekko_tc'):
        mekko_tc = fy26_mro_sam_info['mekko_tc']
        mekko_r  = fy26_mro_sam_info['mekko_row'] - 1  # 0-based
        annotations.append({
            'sheet': 'FY26 MRO SAM',
            'from_col': mekko_tc, 'from_row': mekko_r,
            'from_col_off': 457200,
            'lines': [
                '**Work Type Mix by Hull**',
                '',
                'Shows how each hull\u2019s MRO SAM spend',
                'splits across the four outsourceable work',
                'types: SDM, Continuous/Emergent,',
                'Modernization, and Major Life-Cycle.',
            ],
        })
    # Section (G) annotation — contract-addressable SAM
    if fy26_mro_sam_info and fy26_mro_sam_info.get('g_section_row'):
        gr = fy26_mro_sam_info['g_section_row'] - 1 + 2  # 0-based, +2 to clear purpose row
        gtc = fy26_mro_sam_info.get('g_tc', 4)
        annotations.append({
            'sheet': 'FY26 MRO SAM',
            'from_col': gtc + 1, 'from_row': gr,
            'from_col_off': 457200,
            'lines': [
                '**Contract-Addressable MRO SAM**',
                '',
                'The narrowest SAM \u2014 the MRO market a',
                'company could realistically compete for.',
                '',
                'SDM is reduced from full hull totals to',
                'only the ~31% that flows to private-sector',
                'contracts (Ship Maintenance By Contract +',
                'Equipment Maintenance By Contract from',
                'section E). Government shipyard labor,',
                'materials, MSC, and weapons depot work',
                'are excluded from SDM.',
                '',
                'Cont/Emerg, Mod, and Major LC are retained',
                'at full hull-level totals \u2014 these work',
                'types do not have the same sub-component',
                'breakdown in the justification books.',
            ],
        })

    if annotations:
        add_sticky_notes(out_path, annotations)

    print(f'\nSaved {out_path}')
    print(f'Sheet order: {[s for s in wb.sheetnames]}')

    # Verify formula lengths
    wb2 = openpyxl.load_workbook(out_path)
    max_len = 0; max_cell = ''
    for nm in generated:
        for row in wb2[nm].iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith('=') and len(v) > max_len:
                    max_len = len(v); max_cell = f'{nm}!R{cell.row}C{cell.column}'
    wb2.close()
    status = 'OK' if max_len < 8192 else 'OVER LIMIT!'
    print(f'Max formula: {max_len} chars at {max_cell} — {status}')


if __name__ == '__main__':
    main()
