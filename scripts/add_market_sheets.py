#!/usr/bin/env python3
"""
add_market_sheets.py — Add 5 market-sizing sheets with live SUMIFS formulas.

Uses the v1.23 formula pattern: 6-term SUMIFS cascade for FY2026 values.
Named ranges (JB_B, JB_S, etc.) keep formulas readable.
MRO totals use cell references to intermediate rows to stay under Excel's 8192-char limit.
Each TAM/SAM sheet has: (A) Bridge, (B) All vessel types table, (C) Mekko (non-zero only).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from collections import defaultdict
import shutil

WORKBOOK_SRC = 'output/08APR2028_Newbuild_and_MRO_Spend_v2.1.xlsx'
WORKBOOK_DST = 'output/08APR2028_Newbuild_and_MRO_Spend_v2.2.xlsx'

TAM_CATEGORIES = {'Combatant Ships', 'Auxiliary Ships', 'Cutters', 'Unmanned Maritime Platforms'}
TAM_EXCLUDED_CATEGORIES = ['Combatant Crafts', 'Support Crafts']
SAM_EXCLUDED_TYPES = {'Submarines', 'Aircraft Carriers', 'Unmanned Undersea Vehicles'}

BUCKET_NAMES = {
    '1': 'New Construction', '2': 'Scheduled Depot Maintenance & Repair',
    '3': 'Continuous / Intermediate / Emergent Maintenance',
    '4': 'Modernization & Alteration Installation',
    '5': 'Major Life-Cycle Events / SLEP / MMA / RCOH',
    '6': 'Sustainment Engineering / Planning / Obsolescence Support',
    '7': 'Availability Support / Husbanding / Port Services',
}
BUCKET_SHORT = {
    '2': 'Scheduled Depot Maintenance', '3': 'Continuous / Emergent Maintenance',
    '4': 'Modernization', '5': 'Major Life-Cycle Events / RCOH',
    '6': 'Sustainment Engineering', '7': 'Availability Support',
}
MRO_BKTS = ['2','3','4','5','6','7']

# Styles
TITLE_FONT = Font(name='Calibri', size=14, bold=True)
SUBTITLE_FONT = Font(name='Calibri', size=11, italic=True, color='666666')
SECTION_FONT = Font(name='Calibri', size=11, bold=True)
HEADER_FONT = Font(name='Calibri', size=11, bold=True)
DATA_FONT = Font(name='Calibri', size=11)
BRIDGE_NEG_FONT = Font(name='Calibri', size=11, color='808080')
TOTAL_FONT = Font(name='Calibri', size=11, bold=True)
HEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
TB = Border(bottom=Side(style='thin'))
NUM_FMT = '#,##0'
PCT_FMT = '0.0%'
PCT2 = '0.00'


# ── Formula builder ──────────────────────────────────────────

def cf(*conditions):
    """6-term SUMIFS cascade. Returns '=SUMIFS(...)+...'"""
    cond = (',' + ','.join(conditions)) if conditions else ''
    t = []
    for rt in ['""', '"[PARENT]"']:
        t.append(f'SUMIFS(JB_R{cond},JB_S,{rt},JB_R,"<>")')
        t.append(f'SUMIFS(JB_N{cond},JB_S,{rt},JB_R,"",JB_N,"<>")')
        t.append(f'SUMIFS(JB_L{cond},JB_S,{rt},JB_R,"",JB_N,"")')
    return '=' + '+'.join(t)

def cf_neg(*conditions):
    return f'=-({cf(*conditions)[1:]})'

def cf_inner(*conditions):
    """Cascade without leading '=', for embedding in larger formulas."""
    return cf(*conditions)[1:]


# ── Cell helpers ─────────────────────────────────────────────

def cw(ws, c, w):
    ws.column_dimensions[get_column_letter(c)].width = w

def wc(ws, r, c, v, font=None, fill=None, fmt=None, border=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font: cell.font = font
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    if border: cell.border = border

def hdr_row(ws, r, headers, cs=1):
    for i, (t, w) in enumerate(headers):
        c = cs + i; wc(ws, r, c, t, font=HEADER_FONT, fill=HEADER_FILL); cw(ws, c, w)

def total_cell(ws, r, c, v):
    wc(ws, r, c, v, font=TOTAL_FONT, fill=TOTAL_FILL, fmt=NUM_FMT, border=TB)

def total_label(ws, r, c, v):
    wc(ws, r, c, v, font=TOTAL_FONT, fill=TOTAL_FILL, border=TB)


# ── Service grouping ─────────────────────────────────────────

def svc_group(types, svc_map):
    """Split types into USN/MSC group and USCG group, preserving order."""
    usn = [t for t in types if svc_map.get(t,'') in ('USN','MSC','')]
    cg = [t for t in types if svc_map.get(t,'') == 'USCG']
    return usn, cg

def write_svc_header(ws, r, types, svc_map, col_start=2):
    """Write USN / USCG label row. Returns reordered types."""
    usn, cg = svc_group(types, svc_map)
    ordered = usn + cg
    if usn: wc(ws, r, col_start, 'USN', font=HEADER_FONT)
    if cg: wc(ws, r, col_start + len(usn), 'USCG', font=HEADER_FONT)
    return ordered


# ── Data reading ─────────────────────────────────────────────

def fy26(row):
    for i in [17,13,11]:
        v = row[i].value
        if v is not None and isinstance(v, (int,float)): return float(v)
    return 0.0

def additive(rt):
    if rt is None or str(rt).strip() == '': return True
    return str(rt).strip() == '[PARENT]'

def read_data():
    wb = openpyxl.load_workbook(WORKBOOK_SRC, data_only=True)
    ws = wb['J Book Items Cons.']

    nb = defaultdict(float); mro = defaultdict(float)
    mro_bkt = defaultdict(lambda: defaultdict(float)); mro_b = defaultdict(float)
    type_svc = {}; type_cat = {}
    all_tam_types = set(); tf_src = defaultdict(float)
    nb_excl = defaultdict(float); mro_excl = defaultdict(float)

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        cat = str(row[24].value).strip() if row[24].value else ''   # Y: Vessel Category
        vt = str(row[25].value).strip() if row[25].value else ''    # Z: Vessel Type
        sv = str(row[23].value).strip() if row[23].value else ''    # X: Vessel Service

        # Collect all vessel types from TAM categories (regardless of row type or bucket)
        if cat in TAM_CATEGORIES and vt:
            all_tam_types.add(vt)
            if sv: type_svc[vt] = sv
            type_cat[vt] = cat

        rt = row[21].value                                          # V: Row Type
        if not additive(rt): continue
        bk = row[4].value
        if bk is None: continue
        bs = str(bk).strip()
        if bs not in ('1','2','3','4','5','6','7'): continue

        v = fy26(row)
        src = str(row[0].value).strip() if row[0].value else '(blank)'

        tf_src[src] += v

        if bs == '1':
            if cat in TAM_CATEGORIES: nb[vt] += v
            elif cat: nb_excl[cat] += v
        elif bs in MRO_BKTS:
            if cat in TAM_CATEGORIES:
                mro[vt] += v; mro_bkt[vt][bs] += v; mro_b[bs] += v
            elif cat: mro_excl[cat] += v

    wb.close()

    def sorted_svc(types, val_map):
        """Sort by service group (USN/MSC first, USCG second), then by $ desc within group."""
        usn = sorted([t for t in types if type_svc.get(t,'') in ('USN','MSC','')],
                      key=lambda t: -val_map.get(t, 0))
        cg = sorted([t for t in types if type_svc.get(t,'') == 'USCG'],
                     key=lambda t: -val_map.get(t, 0))
        return usn + cg

    all_sam_types = all_tam_types - SAM_EXCLUDED_TYPES
    nb_nz = [t for t in sorted_svc(all_tam_types, nb) if nb.get(t,0) > 0]
    nb_sam_nz = [t for t in nb_nz if t not in SAM_EXCLUDED_TYPES]
    mro_nz = [t for t in sorted_svc(all_tam_types, mro) if mro.get(t,0) > 0]
    mro_sam_nz = [t for t in mro_nz if t not in SAM_EXCLUDED_TYPES
                  and (mro_bkt[t].get('2',0) + mro_bkt[t].get('4',0)) > 0]

    return {
        'all_tam_types': sorted_svc(all_tam_types, nb),
        'all_sam_types': sorted_svc(all_sam_types, nb),
        'nb_nz': nb_nz, 'nb_sam_nz': nb_sam_nz,
        'mro_nz': mro_nz, 'mro_sam_nz': mro_sam_nz,
        'all_tam_mro_sorted': sorted_svc(all_tam_types, mro),
        'all_sam_mro_sorted': sorted_svc(all_sam_types, {t: mro_bkt[t].get('2',0)+mro_bkt[t].get('4',0) for t in all_sam_types}),
        'type_svc': type_svc,
        'mro_bkts_with_data': [b for b in MRO_BKTS if mro_b.get(b,0) > 0],
        'sources': [s for s,_ in sorted(tf_src.items(), key=lambda x:-x[1])],
        'nb_excl': [c for c,v in sorted(nb_excl.items(), key=lambda x:-x[1]) if v > 0],
        'mro_excl': [c for c,v in sorted(mro_excl.items(), key=lambda x:-x[1]) if v > 0],
    }


# ── Named ranges ─────────────────────────────────────────────

def create_named_ranges(wb):
    JB = "'J Book Items Cons.'"
    for nm, col in [('JB_A','A'),('JB_B','E'),('JB_F','F'),('JB_L','L'),
                    ('JB_N','N'),('JB_R','R'),('JB_S','V'),('JB_U','X'),
                    ('JB_V','Y'),('JB_W','Z')]:
        wb.defined_names.add(DefinedName(nm, attr_text=f"{JB}!${col}$5:${col}$3381"))


# ── Sheet 1: Total Funding ───────────────────────────────────

def create_total_funding(wb, d):
    ws = wb.create_sheet('Total Funding')
    r = 1; wc(ws, r, 1, 'Total Funding', font=TITLE_FONT)
    r = 2; wc(ws, r, 1, 'All Navy + USCG appropriations for vessel-related activity, FY2026 ($K)', font=SUBTITLE_FONT)

    r = 4; wc(ws, r, 1, '(A) FY2026 Total Funding by Work Type', font=SECTION_FONT)
    r = 5; hdr_row(ws, r, [('Work Type', 52), ('FY2026 ($K)', 18)])
    for b in ['1','2','3','4','5','6','7']:
        r += 1; wc(ws, r, 1, BUCKET_NAMES[b], font=DATA_FONT)
        wc(ws, r, 2, cf(f'JB_B,{b}'), font=DATA_FONT, fmt=NUM_FMT)
    r += 1; total_label(ws, r, 1, 'Total'); total_cell(ws, r, 2, f'=SUM(B6:B{r-1})')

    r += 2; wc(ws, r, 1, '(B) FY2026 Total Funding by Vessel Category', font=SECTION_FONT)
    r += 1; hdr_row(ws, r, [('Category', 52), ('FY2026 ($K)', 18), ('TAM Scope', 28)])
    cats = [('Combatant Ships','YES \u2192 TAM'),('Auxiliary Ships','YES \u2192 TAM'),
            ('Cutters','YES \u2192 TAM'),('Unmanned Maritime Platforms','YES \u2192 TAM'),
            ('Combatant Crafts','No \u2014 small craft'),('Support Crafts','No \u2014 non-oceangoing'),
            ('Boats','No \u2014 under 65 ft')]
    tam_rows = []; first = r + 1
    for cat, scope in cats:
        r += 1; wc(ws, r, 1, cat, font=DATA_FONT)
        wc(ws, r, 2, cf(f'JB_V,"{cat}"'), font=DATA_FONT, fmt=NUM_FMT)
        wc(ws, r, 3, scope, font=DATA_FONT)
        if 'YES' in scope: tam_rows.append(r)
    r += 1; wc(ws, r, 1, 'Unattributed (fleet-wide)', font=DATA_FONT)
    wc(ws, r, 2, cf('JB_V,""'), font=DATA_FONT, fmt=NUM_FMT)
    wc(ws, r, 3, 'No \u2014 not vessel-specific', font=DATA_FONT)
    r += 1; total_label(ws, r, 1, 'Total'); total_cell(ws, r, 2, f'=SUM(B{first}:B{r-1})')
    r += 1; blue = Font(name='Calibri', size=11, bold=True, color='2F5496')
    wc(ws, r, 1, 'TAM-eligible', font=blue)
    wc(ws, r, 2, '=' + '+'.join(f'B{tr}' for tr in tam_rows), font=blue, fmt=NUM_FMT)

    r += 2; wc(ws, r, 1, '(C) FY2026 Total Funding by Source Book', font=SECTION_FONT)
    r += 1; hdr_row(ws, r, [('Source Book', 52), ('FY2026 ($K)', 18)]); fs = r + 1
    for s in d['sources']:
        r += 1; wc(ws, r, 1, s, font=DATA_FONT); wc(ws, r, 2, cf(f'JB_A,"{s}"'), font=DATA_FONT, fmt=NUM_FMT)
    r += 1; total_label(ws, r, 1, 'Total'); total_cell(ws, r, 2, f'=SUM(B{fs}:B{r-1})')


# ── Sheet 2: Newbuild TAM ───────────────────────────────────

def create_newbuild_tam(wb, d):
    ws = wb.create_sheet('Newbuild TAM')
    r = 1; wc(ws, r, 1, 'Newbuild Total Addressable Market \u2014 FY2026 ($K)', font=TITLE_FONT)
    r = 2; wc(ws, r, 1, 'New Construction funding scoped to Combatant Ships, Auxiliary Ships, USCG Cutters & Unmanned Platforms', font=SUBTITLE_FONT)
    cw(ws, 1, 52); cw(ws, 2, 18)

    # (A) Bridge
    r = 4; wc(ws, r, 1, '(A) Bridge: Total NC Funding \u2192 Newbuild TAM', font=SECTION_FONT)
    r = 5; wc(ws, r, 1, 'Total New Construction Funding', font=DATA_FONT)
    wc(ws, r, 2, cf('JB_B,1'), font=DATA_FONT, fmt=NUM_FMT)
    r += 1; wc(ws, r, 1, '  Less: Unattributed (fleet-wide NC items)', font=BRIDGE_NEG_FONT)
    wc(ws, r, 2, cf_neg('JB_B,1', 'JB_V,""'), font=BRIDGE_NEG_FONT, fmt=NUM_FMT)
    for cat in TAM_EXCLUDED_CATEGORIES:
        r += 1; wc(ws, r, 1, f'  Less: {cat}', font=BRIDGE_NEG_FONT)
        wc(ws, r, 2, cf_neg('JB_B,1', f'JB_V,"{cat}"'), font=BRIDGE_NEG_FONT, fmt=NUM_FMT)
    r += 1
    total_label(ws, r, 1, 'Newbuild TAM')
    tam_f = '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES))
    total_cell(ws, r, 2, '=' + tam_f)

    # (B) All vessel types
    r += 2; wc(ws, r, 1, '(B) Newbuild TAM by Vessel Type', font=SECTION_FONT)
    r += 1; all_types = write_svc_header(ws, r, d['all_tam_types'], d['type_svc'])
    r += 1; hdr_row(ws, r, [('Vessel Type', 40), ('Service', 10), ('FY2026 ($K)', 16), ('% of TAM', 12)])
    ft = r + 1
    for vt in all_types:
        r += 1; wc(ws, r, 1, vt, font=DATA_FONT)
        wc(ws, r, 2, d['type_svc'].get(vt,''), font=DATA_FONT)
        wc(ws, r, 3, cf('JB_B,1', f'JB_W,"{vt}"'), font=DATA_FONT, fmt=NUM_FMT)
    lt = r; r += 1; tr = r
    total_label(ws, r, 1, 'Newbuild TAM'); total_label(ws, r, 2, '')
    total_cell(ws, r, 3, f'=SUM(C{ft}:C{lt})')
    wc(ws, r, 4, 1.0, font=TOTAL_FONT, fill=TOTAL_FILL, fmt=PCT_FMT, border=TB)
    for rn in range(ft, lt+1):
        wc(ws, rn, 4, f'=IFERROR(C{rn}/C${tr},0)', font=DATA_FONT, fmt=PCT_FMT)

    # (C) Mekko (non-zero only)
    r += 2; wc(ws, r, 1, '(C) Newbuild TAM \u2014 DD&C vs AP/LLTM by Vessel Type (Mekko)', font=SECTION_FONT)
    r += 1; mk = write_svc_header(ws, r, d['nb_nz'], d['type_svc'])
    r += 1; wc(ws, r, 1, 'Sub-Category', font=HEADER_FONT, fill=HEADER_FILL); cw(ws, 1, 40)
    for i, vt in enumerate(mk):
        c = 2+i; wc(ws, r, c, vt, font=HEADER_FONT, fill=HEADER_FILL); cw(ws, c, 16)
    tc = 2+len(mk); wc(ws, r, tc, 'Total', font=HEADER_FONT, fill=HEADER_FILL); cw(ws, tc, 16)
    dr = r+1; ar = r+2; orr = r+3; tkr = r+4

    # Total ($K)
    wc(ws, tkr, 1, 'Total ($K)', font=TOTAL_FONT, fill=TOTAL_FILL, border=TB)
    for i, vt in enumerate(mk):
        c = 2+i; total_cell(ws, tkr, c, cf('JB_B,1', f'JB_W,"{vt}"'))
    total_cell(ws, tkr, tc, f'=SUM({get_column_letter(2)}{tkr}:{get_column_letter(tc-1)}{tkr})')

    # DD&C
    ddc = ['"Full ship DD&C"','"Full ship DD&C / LLTM"','"Full vessel procurement"']
    wc(ws, dr, 1, 'Full Ship DD&C', font=DATA_FONT)
    for i, vt in enumerate(mk):
        c = 2+i; cl = get_column_letter(c)
        parts = '+'.join(cf_inner('JB_B,1', f'JB_W,"{vt}"', f'JB_F,{s}') for s in ddc)
        wc(ws, dr, c, f'=IFERROR(({parts})/{cl}${tkr},0)', font=DATA_FONT, fmt=PCT2)
    tcl = get_column_letter(tc)
    wc(ws, dr, tc, f'=IFERROR(1-{tcl}{ar},0)', font=DATA_FONT, fmt=PCT2)

    # AP/LLTM
    wc(ws, ar, 1, 'Advance Procurement / LLTM', font=DATA_FONT)
    for i, vt in enumerate(mk):
        c = 2+i; cl = get_column_letter(c)
        wc(ws, ar, c, f'=IFERROR(({cf_inner("JB_B,1", f"""JB_W,"{vt}" """.strip(), "JB_F,"+chr(34)+"Advance procurement / LLTM"+chr(34))})/{cl}${tkr},0)', font=DATA_FONT, fmt=PCT2)
    ap_cat_parts = '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"', 'JB_F,"Advance procurement / LLTM"') for c in sorted(TAM_CATEGORIES))
    wc(ws, ar, tc, f'=IFERROR(({ap_cat_parts})/{tcl}${tkr},0)', font=DATA_FONT, fmt=PCT2)

    # Other
    wc(ws, orr, 1, 'Other NC Sub-Categories', font=DATA_FONT)
    for i in range(len(mk)):
        c = 2+i; cl = get_column_letter(c)
        wc(ws, orr, c, f'=IFERROR(1-{cl}{dr}-{cl}{ar},0)', font=DATA_FONT, fmt=PCT2)
    wc(ws, orr, tc, f'=IFERROR(1-{tcl}{dr}-{tcl}{ar},0)', font=DATA_FONT, fmt=PCT2)


# ── Sheet 3: Newbuild SAM ───────────────────────────────────

def create_newbuild_sam(wb, d):
    ws = wb.create_sheet('Newbuild SAM')
    r = 1; wc(ws, r, 1, 'Newbuild Serviceable Addressable Market \u2014 FY2026 ($K)', font=TITLE_FONT)
    r = 2; wc(ws, r, 1, 'Newbuild TAM excluding vessel types not addressable by PA', font=SUBTITLE_FONT)
    cw(ws, 1, 52); cw(ws, 2, 18)

    # (A) Bridge
    r = 4; wc(ws, r, 1, '(A) Bridge: Newbuild TAM \u2192 Newbuild SAM', font=SECTION_FONT)
    r = 5; wc(ws, r, 1, 'Newbuild TAM', font=DATA_FONT)
    wc(ws, r, 2, '=' + '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES)), font=DATA_FONT, fmt=NUM_FMT)
    for vt in ['Submarines', 'Aircraft Carriers', 'Unmanned Undersea Vehicles']:
        r += 1; wc(ws, r, 1, f'  Less: {vt}', font=BRIDGE_NEG_FONT)
        wc(ws, r, 2, cf_neg('JB_B,1', f'JB_W,"{vt}"'), font=BRIDGE_NEG_FONT, fmt=NUM_FMT)
    r += 1; total_label(ws, r, 1, 'Newbuild SAM')
    total_cell(ws, r, 2, f'=SUM(B5:B{r-1})')

    # (B) All SAM vessel types
    r += 2; wc(ws, r, 1, '(B) Newbuild SAM by Vessel Type', font=SECTION_FONT)
    r += 1; all_types = write_svc_header(ws, r, d['all_sam_types'], d['type_svc'])
    r += 1; hdr_row(ws, r, [('Vessel Type', 40), ('Service', 10), ('FY2026 ($K)', 16), ('% of SAM', 12)])
    ft = r + 1
    for vt in all_types:
        r += 1; wc(ws, r, 1, vt, font=DATA_FONT)
        wc(ws, r, 2, d['type_svc'].get(vt,''), font=DATA_FONT)
        wc(ws, r, 3, cf('JB_B,1', f'JB_W,"{vt}"'), font=DATA_FONT, fmt=NUM_FMT)
    lt = r; r += 1; tr = r
    total_label(ws, r, 1, 'Newbuild SAM'); total_label(ws, r, 2, '')
    total_cell(ws, r, 3, f'=SUM(C{ft}:C{lt})')
    wc(ws, r, 4, 1.0, font=TOTAL_FONT, fill=TOTAL_FILL, fmt=PCT_FMT, border=TB)
    for rn in range(ft, lt+1):
        wc(ws, rn, 4, f'=IFERROR(C{rn}/C${tr},0)', font=DATA_FONT, fmt=PCT_FMT)

    # (C) Mekko (non-zero only)
    r += 2; wc(ws, r, 1, '(C) Newbuild SAM \u2014 DD&C vs AP/LLTM by Vessel Type (Mekko)', font=SECTION_FONT)
    r += 1; mk = write_svc_header(ws, r, d['nb_sam_nz'], d['type_svc'])
    r += 1; wc(ws, r, 1, 'Sub-Category', font=HEADER_FONT, fill=HEADER_FILL); cw(ws, 1, 40)
    for i, vt in enumerate(mk):
        c = 2+i; wc(ws, r, c, vt, font=HEADER_FONT, fill=HEADER_FILL); cw(ws, c, 16)
    tc = 2+len(mk); wc(ws, r, tc, 'Total', font=HEADER_FONT, fill=HEADER_FILL); cw(ws, tc, 16)
    dr = r+1; ar = r+2; tkr = r+3

    wc(ws, tkr, 1, 'Total ($K)', font=TOTAL_FONT, fill=TOTAL_FILL, border=TB)
    for i, vt in enumerate(mk):
        total_cell(ws, tkr, 2+i, cf('JB_B,1', f'JB_W,"{vt}"'))
    total_cell(ws, tkr, tc, f'=SUM({get_column_letter(2)}{tkr}:{get_column_letter(tc-1)}{tkr})')

    ddc = ['"Full ship DD&C"','"Full ship DD&C / LLTM"','"Full vessel procurement"']
    wc(ws, dr, 1, 'Full Ship DD&C', font=DATA_FONT)
    for i, vt in enumerate(mk):
        c = 2+i; cl = get_column_letter(c)
        parts = '+'.join(cf_inner('JB_B,1', f'JB_W,"{vt}"', f'JB_F,{s}') for s in ddc)
        wc(ws, dr, c, f'=IFERROR(({parts})/{cl}${tkr},0)', font=DATA_FONT, fmt=PCT2)
    tcl = get_column_letter(tc)
    wc(ws, dr, tc, f'=IFERROR(1-{tcl}{ar},0)', font=DATA_FONT, fmt=PCT2)

    wc(ws, ar, 1, 'Advance Procurement / LLTM', font=DATA_FONT)
    for i, vt in enumerate(mk):
        c = 2+i; cl = get_column_letter(c)
        wc(ws, ar, c, f'=IFERROR(({cf_inner("JB_B,1", f"""JB_W,"{vt}" """.strip(), "JB_F,"+chr(34)+"Advance procurement / LLTM"+chr(34))})/{cl}${tkr},0)', font=DATA_FONT, fmt=PCT2)
    ap_cat = '+'.join(cf_inner('JB_B,1', f'JB_V,"{c}"', 'JB_F,"Advance procurement / LLTM"') for c in sorted(TAM_CATEGORIES))
    wc(ws, ar, tc, f'=IFERROR(({ap_cat})/{tcl}${tkr},0)', font=DATA_FONT, fmt=PCT2)


# ── Sheet 4: MRO TAM ────────────────────────────────────────

def create_mro_tam(wb, d):
    ws = wb.create_sheet('MRO TAM')
    r = 1; wc(ws, r, 1, 'MRO Total Addressable Market \u2014 FY2026 ($K)', font=TITLE_FONT)
    r = 2; wc(ws, r, 1, 'MRO & sustainment funding scoped to Combatant Ships, Auxiliary Ships, USCG Cutters & Unmanned Platforms', font=SUBTITLE_FONT)
    cw(ws, 1, 52); cw(ws, 2, 18)

    bkts = d['mro_bkts_with_data']

    # (A) Bridge
    r = 4; wc(ws, r, 1, '(A) Bridge: Total MRO Funding \u2192 MRO TAM', font=SECTION_FONT)
    r = 5; wc(ws, r, 1, 'Total MRO Funding (Buckets 2\u20137)', font=DATA_FONT)
    wc(ws, r, 2, '=' + '+'.join(cf_inner(f'JB_B,{b}') for b in MRO_BKTS), font=DATA_FONT, fmt=NUM_FMT)
    r += 1; wc(ws, r, 1, '  Less: Unattributed (fleet-wide MRO items)', font=BRIDGE_NEG_FONT)
    wc(ws, r, 2, f'=-({"+".join(cf_inner(f"JB_B,{b}", "JB_V,"+chr(34)+chr(34)) for b in MRO_BKTS)})', font=BRIDGE_NEG_FONT, fmt=NUM_FMT)
    for cat in TAM_EXCLUDED_CATEGORIES:
        r += 1; wc(ws, r, 1, f'  Less: {cat}', font=BRIDGE_NEG_FONT)
        wc(ws, r, 2, f'=-({"+".join(cf_inner(f"JB_B,{b}", f"""JB_V,"{cat}" """.strip()) for b in MRO_BKTS)})', font=BRIDGE_NEG_FONT, fmt=NUM_FMT)
    r += 1; bridge_tam_row = r
    total_label(ws, r, 1, 'MRO TAM')
    # Will be filled AFTER Section B with =SUM reference

    # (B) By Work Type — provides intermediate rows for bridge total
    r += 2; wc(ws, r, 1, '(B) MRO TAM by Work Type', font=SECTION_FONT)
    r += 1; hdr_row(ws, r, [('Work Type', 52), ('FY2026 ($K)', 16), ('% of TAM', 12)])
    fb = r + 1
    for b in bkts:
        r += 1; wc(ws, r, 1, BUCKET_SHORT.get(b, BUCKET_NAMES[b]), font=DATA_FONT)
        parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES))
        wc(ws, r, 2, '=' + parts, font=DATA_FONT, fmt=NUM_FMT)
    lb = r; r += 1; btr = r
    total_label(ws, r, 1, 'MRO TAM')
    total_cell(ws, r, 2, f'=SUM(B{fb}:B{lb})')
    wc(ws, r, 3, 1.0, font=TOTAL_FONT, fill=TOTAL_FILL, fmt=PCT_FMT, border=TB)
    for rn in range(fb, lb+1):
        wc(ws, rn, 3, f'=IFERROR(B{rn}/B${btr},0)', font=DATA_FONT, fmt=PCT_FMT)

    # NOW fill bridge result as reference to work-type total
    total_cell(ws, bridge_tam_row, 2, f'=B{btr}')

    # (C) All vessel types
    r += 2; wc(ws, r, 1, '(C) MRO TAM by Vessel Type', font=SECTION_FONT)
    r += 1; all_types = write_svc_header(ws, r, d['all_tam_mro_sorted'], d['type_svc'])
    r += 1; hdr_row(ws, r, [('Vessel Type', 40), ('Service', 10), ('FY2026 ($K)', 16), ('% of TAM', 12)])
    ft = r + 1
    for vt in all_types:
        r += 1; wc(ws, r, 1, vt, font=DATA_FONT)
        wc(ws, r, 2, d['type_svc'].get(vt,''), font=DATA_FONT)
        mparts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_W,"{vt}"') for b in MRO_BKTS)
        wc(ws, r, 3, '=' + mparts, font=DATA_FONT, fmt=NUM_FMT)
    lt = r; r += 1; ttr = r
    total_label(ws, r, 1, 'MRO TAM'); total_label(ws, r, 2, '')
    total_cell(ws, r, 3, f'=SUM(C{ft}:C{lt})')
    wc(ws, r, 4, 1.0, font=TOTAL_FONT, fill=TOTAL_FILL, fmt=PCT_FMT, border=TB)
    for rn in range(ft, lt+1):
        wc(ws, rn, 4, f'=IFERROR(C{rn}/C${ttr},0)', font=DATA_FONT, fmt=PCT_FMT)

    # (D) Mekko (non-zero only)
    r += 2; wc(ws, r, 1, '(D) MRO TAM \u2014 Work Type by Vessel Type (Mekko)', font=SECTION_FONT)
    r += 1; mk = write_svc_header(ws, r, d['mro_nz'], d['type_svc'], col_start=2)
    r += 1; wc(ws, r, 1, 'Work Type', font=HEADER_FONT, fill=HEADER_FILL); cw(ws, 1, 40)
    for i, vt in enumerate(mk):
        c = 2+i; wc(ws, r, c, vt, font=HEADER_FONT, fill=HEADER_FILL); cw(ws, c, 16)
    tc = 2+len(mk); wc(ws, r, tc, 'Total', font=HEADER_FONT, fill=HEADER_FILL); cw(ws, tc, 16)
    nwt = len(bkts); tkr = r + 1 + nwt

    # Total ($K) — use SUM across type columns (not giant formula)
    wc(ws, tkr, 1, 'Total ($K)', font=TOTAL_FONT, fill=TOTAL_FILL, border=TB)
    for i, vt in enumerate(mk):
        c = 2+i; mparts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_W,"{vt}"') for b in MRO_BKTS)
        total_cell(ws, tkr, c, '=' + mparts)
    total_cell(ws, tkr, tc, f'=SUM({get_column_letter(2)}{tkr}:{get_column_letter(tc-1)}{tkr})')

    # Work type rows
    tcl = get_column_letter(tc)
    for bi, b in enumerate(bkts):
        wr = r + 1 + bi
        wc(ws, wr, 1, BUCKET_SHORT.get(b, BUCKET_NAMES[b]), font=DATA_FONT)
        for i, vt in enumerate(mk):
            c = 2+i; cl = get_column_letter(c)
            wc(ws, wr, c, f'=IFERROR(({cf_inner(f"JB_B,{b}", f"""JB_W,"{vt}" """.strip())})/{cl}${tkr},0)', font=DATA_FONT, fmt=PCT2)
        # Total col: sum across types for this bucket / total
        bkt_parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_V,"{c}"') for c in sorted(TAM_CATEGORIES))
        wc(ws, wr, tc, f'=IFERROR(({bkt_parts})/{tcl}${tkr},0)', font=DATA_FONT, fmt=PCT2)

    return {'bkt_total_row': btr}  # for cross-sheet reference


# ── Sheet 5: MRO SAM ────────────────────────────────────────

def create_mro_sam(wb, d, mro_tam_info=None):
    ws = wb.create_sheet('MRO SAM')
    r = 1; wc(ws, r, 1, 'MRO Serviceable Addressable Market \u2014 FY2026 ($K)', font=TITLE_FONT)
    r = 2; wc(ws, r, 1, 'Outsourceable MRO: SDM + Modernization, excluding non-addressable vessel types', font=SUBTITLE_FONT)
    cw(ws, 1, 56); cw(ws, 2, 18)

    # (A) Bridge — cross-sheet reference for MRO TAM to avoid >8192 char formula
    r = 4; wc(ws, r, 1, '(A) Bridge: MRO TAM \u2192 MRO SAM', font=SECTION_FONT)

    r = 5; wc(ws, r, 1, 'MRO TAM', font=DATA_FONT)
    tam_ref_row = mro_tam_info['bkt_total_row'] if mro_tam_info else 18
    wc(ws, r, 2, f"='MRO TAM'!B{tam_ref_row}", font=DATA_FONT, fmt=NUM_FMT)

    excl_rows = []
    for vt in ['Submarines', 'Aircraft Carriers', 'Unmanned Undersea Vehicles']:
        r += 1; excl_rows.append(r)
        wc(ws, r, 1, f'  Less: {vt}', font=BRIDGE_NEG_FONT)
        parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_W,"{vt}"') for b in MRO_BKTS)
        wc(ws, r, 2, f'=-({parts})', font=BRIDGE_NEG_FONT, fmt=NUM_FMT)

    r += 1; nowork_row = r; wc(ws, r, 1, '  Less: Non-outsourceable work types', font=BRIDGE_NEG_FONT)
    # Residual: filled after SAM result

    r += 1; sam_result_row = r
    total_label(ws, r, 1, 'MRO SAM')
    # SAM = outsourceable (bkt 2+4) for TAM cats minus excluded types
    sam_parts = '+'.join(cf_inner(f'JB_B,{b}', f'JB_V,"{c}"') for b in ['2','4'] for c in sorted(TAM_CATEGORIES))
    excl_parts = '+'.join(f'-({cf_inner(f"JB_B,{b}", f"""JB_W,"{vt}" """.strip())})' for b in ['2','4'] for vt in ['Submarines','Aircraft Carriers','Unmanned Undersea Vehicles'])
    total_cell(ws, r, 2, f'={sam_parts}{excl_parts}')

    # Non-outsourceable = residual
    excl_sum = '+'.join(f'B{er}' for er in excl_rows)
    wc(ws, nowork_row, 2, f'=-(B5+{excl_sum}-B{sam_result_row})', font=BRIDGE_NEG_FONT, fmt=NUM_FMT)

    # (B) All SAM vessel types
    r += 2; wc(ws, r, 1, '(B) MRO SAM by Vessel Type', font=SECTION_FONT)
    r += 1; all_types = write_svc_header(ws, r, d['all_sam_mro_sorted'], d['type_svc'])
    r += 1; hdr_row(ws, r, [('Vessel Type', 40), ('Service', 10), ('SDM ($K)', 14), ('Mod ($K)', 14), ('Total ($K)', 16), ('% of SAM', 12)])
    ft = r + 1
    for vt in all_types:
        r += 1; wc(ws, r, 1, vt, font=DATA_FONT)
        wc(ws, r, 2, d['type_svc'].get(vt,''), font=DATA_FONT)
        wc(ws, r, 3, cf('JB_B,2', f'JB_W,"{vt}"'), font=DATA_FONT, fmt=NUM_FMT)
        wc(ws, r, 4, cf('JB_B,4', f'JB_W,"{vt}"'), font=DATA_FONT, fmt=NUM_FMT)
        wc(ws, r, 5, f'=C{r}+D{r}', font=DATA_FONT, fmt=NUM_FMT)
    lt = r; r += 1; tr = r
    total_label(ws, r, 1, 'MRO SAM'); total_label(ws, r, 2, '')
    for c in [3,4,5]:
        total_cell(ws, r, c, f'=SUM({get_column_letter(c)}{ft}:{get_column_letter(c)}{lt})')
    wc(ws, r, 6, 1.0, font=TOTAL_FONT, fill=TOTAL_FILL, fmt=PCT_FMT, border=TB)
    for rn in range(ft, lt+1):
        wc(ws, rn, 6, f'=IFERROR(E{rn}/E${tr},0)', font=DATA_FONT, fmt=PCT_FMT)

    # (C) Mekko (non-zero only)
    r += 2; wc(ws, r, 1, '(C) MRO SAM \u2014 SDM vs Modernization by Vessel Type (Mekko)', font=SECTION_FONT)
    r += 1; mk = write_svc_header(ws, r, d['mro_sam_nz'], d['type_svc'], col_start=2)
    r += 1; wc(ws, r, 1, 'Work Type', font=HEADER_FONT, fill=HEADER_FILL); cw(ws, 1, 40)
    for i, vt in enumerate(mk):
        c = 2+i; wc(ws, r, c, vt, font=HEADER_FONT, fill=HEADER_FILL); cw(ws, c, 16)
    tc = 2+len(mk); wc(ws, r, tc, 'Total', font=HEADER_FONT, fill=HEADER_FILL); cw(ws, tc, 16)
    sr = r+1; mr = r+2; tkr = r+3

    wc(ws, tkr, 1, 'Total ($K)', font=TOTAL_FONT, fill=TOTAL_FILL, border=TB)
    for i, vt in enumerate(mk):
        c = 2+i; parts = cf_inner('JB_B,2', f'JB_W,"{vt}"') + '+' + cf_inner('JB_B,4', f'JB_W,"{vt}"')
        total_cell(ws, tkr, c, '=' + parts)
    total_cell(ws, tkr, tc, f'=SUM({get_column_letter(2)}{tkr}:{get_column_letter(tc-1)}{tkr})')

    tcl = get_column_letter(tc)
    wc(ws, sr, 1, 'Scheduled Depot Maintenance', font=DATA_FONT)
    for i, vt in enumerate(mk):
        c = 2+i; cl = get_column_letter(c)
        wc(ws, sr, c, f'=IFERROR(({cf_inner("JB_B,2", f"""JB_W,"{vt}" """.strip())})/{cl}${tkr},0)', font=DATA_FONT, fmt=PCT2)
    wc(ws, sr, tc, f'=IFERROR(1-{tcl}{mr},0)', font=DATA_FONT, fmt=PCT2)

    wc(ws, mr, 1, 'Modernization', font=DATA_FONT)
    for i, vt in enumerate(mk):
        c = 2+i; cl = get_column_letter(c)
        wc(ws, mr, c, f'=IFERROR(({cf_inner("JB_B,4", f"""JB_W,"{vt}" """.strip())})/{cl}${tkr},0)', font=DATA_FONT, fmt=PCT2)
    mod_parts = '+'.join(cf_inner('JB_B,4', f'JB_W,"{vt}"') for vt in mk)
    wc(ws, mr, tc, f'=IFERROR(({mod_parts})/{tcl}${tkr},0)', font=DATA_FONT, fmt=PCT2)


# ── Main ─────────────────────────────────────────────────────

def main():
    print('Reading J Book data...')
    d = read_data()
    print(f'  All TAM types: {len(d["all_tam_types"])}, All SAM types: {len(d["all_sam_types"])}')
    print(f'  NB non-zero: {len(d["nb_nz"])}, NB SAM non-zero: {len(d["nb_sam_nz"])}')
    print(f'  MRO non-zero: {len(d["mro_nz"])}, MRO SAM non-zero: {len(d["mro_sam_nz"])}')

    print(f'Copying {WORKBOOK_SRC} -> {WORKBOOK_DST}...')
    shutil.copy2(WORKBOOK_SRC, WORKBOOK_DST)

    wb = openpyxl.load_workbook(WORKBOOK_DST)

    # Remove pre-existing market sheets so we don't create duplicates
    for nm in ['Total Funding','Newbuild TAM','Newbuild SAM','MRO TAM','MRO SAM']:
        if nm in wb.sheetnames:
            del wb[nm]

    print('Creating named ranges...')
    create_named_ranges(wb)

    print('Creating sheets...')
    create_total_funding(wb, d)
    create_newbuild_tam(wb, d)
    create_newbuild_sam(wb, d)
    mro_tam_info = create_mro_tam(wb, d)
    create_mro_sam(wb, d, mro_tam_info)

    # Place market sheets right after 'J Book Items Cons.'
    market = ['Total Funding','Newbuild TAM','Newbuild SAM','MRO TAM','MRO SAM']
    other = [s for s in wb.sheetnames if s not in market]
    wb._sheets = [wb[other[0]]] + [wb[m] for m in market] + [wb[s] for s in other[1:]]

    wb.save(WORKBOOK_DST)
    print(f'Saved {WORKBOOK_DST}')

    # Verify no formula exceeds limit
    wb2 = openpyxl.load_workbook(WORKBOOK_DST)
    max_len = 0; max_cell = ''
    for nm in ['Total Funding','Newbuild TAM','Newbuild SAM','MRO TAM','MRO SAM']:
        for row in wb2[nm].iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith('=') and len(v) > max_len:
                    max_len = len(v); max_cell = f'{nm}!R{cell.row}C{cell.column}'
    wb2.close()
    status = 'OK' if max_len < 8192 else 'OVER LIMIT!'
    print(f'  Max formula: {max_len} chars at {max_cell} — {status}')
    print('Sheet order:', [s for s in wb.sheetnames])


if __name__ == '__main__':
    main()
