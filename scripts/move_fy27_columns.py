#!/usr/bin/env python3
"""
move_fy27_columns.py — Move FY2027 columns (AI-AK) to after column R in Sheet 1.

Repositions the three FY2027 budget columns so fiscal-year dollar data is contiguous:
  Before: ...R (FY2026 Enacted), S (Row Type), ... AH (PE Desc), AI/AJ/AK (FY2027)
  After:  ...R (FY2026 Enacted), S/T/U (FY2027), V (Row Type), ... AK (PE Desc)

Named ranges (JB_S, JB_U, JB_V, JB_W) are rebuilt to match the shifted columns.
"""

import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from copy import copy

WORKBOOK_SRC = 'output/08APR2028_Newbuild_and_MRO_Spend_v2.0.xlsx'
WORKBOOK_DST = 'output/08APR2028_Newbuild_and_MRO_Spend_v2.1.xlsx'

# Columns to move: AI(35), AJ(36), AK(37) → insert after R(18), i.e. at position 19
SRC_COLS = [35, 36, 37]
INSERT_AT = 19  # new position for the first moved column

# Named ranges — column letters AFTER the move
# Columns A-R (≤18) are unaffected; columns formerly at S+ shift right by 3
NAMED_RANGES_AFTER = [
    ('JB_A', 'A'),   # Source Book (col 1, unchanged)
    ('JB_B', 'E'),   # Bucket (col 5, unchanged)
    ('JB_F', 'F'),   # Bucket Sub Category (col 6, unchanged)
    ('JB_L', 'L'),   # FY2026 Request (col 12, unchanged)
    ('JB_N', 'N'),   # FY2026 Total (col 14, unchanged)
    ('JB_R', 'R'),   # FY2026 Enacted (col 18, unchanged)
    ('JB_S', 'V'),   # Row Type (was S/19, now V/22)
    ('JB_U', 'X'),   # Vessel Service (was U/21, now X/24)
    ('JB_V', 'Y'),   # Vessel Category (was V/22, now Y/25)
    ('JB_W', 'Z'),   # Vessel Type (was W/23, now Z/26)
]


def main():
    print(f'Loading {WORKBOOK_SRC}...')
    wb = openpyxl.load_workbook(WORKBOOK_SRC)
    ws = wb[wb.sheetnames[0]]
    max_row = ws.max_row

    # ── Step 1: Extract AI-AK data (values + styles) ────────────
    print(f'Extracting columns AI-AK ({max_row} rows)...')
    saved = {}
    for row in range(1, max_row + 1):
        for i, col in enumerate(SRC_COLS):
            cell = ws.cell(row=row, column=col)
            saved[(row, i)] = {
                'value': cell.value,
                'font': copy(cell.font),
                'fill': copy(cell.fill),
                'alignment': copy(cell.alignment),
                'border': copy(cell.border),
                'number_format': cell.number_format,
            }

    saved_widths = []
    for col in SRC_COLS:
        dim = ws.column_dimensions.get(get_column_letter(col))
        saved_widths.append(dim.width if dim and dim.width else None)

    # ── Step 2: Delete columns AI-AK (35-37) ────────────────────
    print('Deleting original columns AI-AK...')
    ws.delete_cols(SRC_COLS[0], len(SRC_COLS))

    # ── Step 3: Insert 3 blank columns at position 19 ───────────
    print(f'Inserting 3 columns at position {INSERT_AT} (after R)...')
    ws.insert_cols(INSERT_AT, len(SRC_COLS))

    # ── Step 4: Write extracted data to new positions ────────────
    print('Writing FY2027 data to new column positions...')
    for row in range(1, max_row + 1):
        for i in range(len(SRC_COLS)):
            data = saved[(row, i)]
            cell = ws.cell(row=row, column=INSERT_AT + i)
            cell.value = data['value']
            cell.font = data['font']
            cell.fill = data['fill']
            cell.alignment = data['alignment']
            cell.border = data['border']
            cell.number_format = data['number_format']

    # ── Step 5: Set column widths ────────────────────────────────
    for i, w in enumerate(saved_widths):
        if w is not None:
            ws.column_dimensions[get_column_letter(INSERT_AT + i)].width = w

    # ── Step 6: Rebuild named ranges ────────────────────────────
    print('Rebuilding named ranges...')
    JB = "'J Book Items Cons.'"

    # Delete all existing named ranges
    for name in list(wb.defined_names):
        del wb.defined_names[name]

    # Recreate with corrected column letters
    for nm, col in NAMED_RANGES_AFTER:
        wb.defined_names.add(
            DefinedName(nm, attr_text=f"{JB}!${col}$5:${col}$3381")
        )

    # ── Step 7: Save ─────────────────────────────────────────────
    print(f'Saving {WORKBOOK_DST}...')
    wb.save(WORKBOOK_DST)
    print('Done.')


if __name__ == '__main__':
    main()
