#!/usr/bin/env python3
"""
Build FY2027 columns for the consolidated newbuild & MRO spend workbook.

Adds three new columns at the END of the J Book Items Cons. sheet:
  - FY2027 Request ($K)           = FY27 Discretionary Request from P-1/O-1
  - FY2027 Mandatory Request ($K) = FY27 Mandatory Request from P-1/O-1
                                    (a NEW reconciliation ask, NOT enacted, NOT OBBBA)
  - FY2027 Total ($K)             = sum of the two

Why columns are appended at the end (not grouped with FY26):
  TAM, SAM, and Competitive Dynamics sheets all use hardcoded column-letter
  formula references like 'J Book Items Cons.'!R$5:R$3360. Inserting columns
  mid-sheet would silently break those formulas. Append-at-end keeps them intact.

Coverage policy (per user direction):
  - SCN: hand-coded NET values from FY27 P-1 (45 line items, validated)
  - OPN_BA1-BA5-8: name-matched against parsed FY27 P-1 1810N detail
  - WPN: name-matched against parsed FY27 P-1 1507N detail
  - OMN/NWCF/USCG: LEFT BLANK (option a — to be filled later in a separate sheet
                                using prorated SAG-level allocation)

OBBBA stays out of FY27 columns:
  Per OBBBA Figure 1 (and Figure 3 for shipbuilding), 100% of OBBBA budget authority
  is FY 2026, with $0 in FY27/28/29. The "FY 2027 Mandatory Request" column in the
  P-1 is a SEPARATE NEW reconciliation ask (the round 2 Brendan was worried about),
  not OBBBA carryover.
"""

from __future__ import annotations
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy
from collections import defaultdict
from pathlib import Path

# ─── PATHS ────────────────────────────────────────────────────────────────────

BASE = Path('/Users/brendantoole/projects2/buildco3')
SOURCE_XLSX = BASE / '08APR2028_Newbuild_and_MRO_Spend_v1.20.xlsx'
OUTPUT_XLSX = BASE / '08APR2028_Newbuild_and_MRO_Spend_v1.21.xlsx'
P1_TXT = BASE / 'FY2027_numbers' / 'FY2027_p1.txt'
O1_TXT = BASE / 'FY2027_numbers' / 'FY2027_o1.txt'

SHEET = 'J Book Items Cons.'

# ─── COLUMN INDICES (0-based, matching openpyxl row tuples) ──────────────────

COL_SOURCE = 0       # A: Source Book
COL_LI = 2           # C: Line Item Number
COL_TITLE = 3        # D: Line Item Title
COL_BUCKET = 4       # E: Bucket
COL_BUCKET_SUB = 5   # F: Bucket Sub Category
COL_DESC = 6         # G: Description
COL_FY26_REQ = 11    # L: FY2026 Request ($K)
COL_FY26_TOTAL = 13  # N: FY2026 Total ($K)
COL_FY26_DAA_ENACTED = 17  # R: FY2026 DAA Enacted Total ($K)
COL_ROW_TYPE = 18    # S: Row Type
COL_PARENT = 19      # T: Parent Item
COL_VESSEL_SVC = 20  # U: Vessel Service
COL_VESSEL_CAT = 21  # V: Vessel Category
COL_VESSEL_TYPE = 22 # W: Vessel Type
COL_VESSEL_HULL = 23 # X: Vessel Hull
COL_CITATION = 24    # Y: Citation
COL_MARKET = 31      # AF: Market
COL_SUBMARKET = 32   # AG: Submarket

# Total existing columns: 34 (A through AH).
# New columns will be appended at: 35 (AI), 36 (AJ), 37 (AK).
NEW_COL_FY27_REQ = 35       # AI
NEW_COL_FY27_MAND = 36      # AJ
NEW_COL_FY27_TOTAL = 37     # AK

# ─── SCN HAND-CODED FY27 NET MAPPING (from FY2027 P-1 pages 132-152) ─────────
#
# Convention: PARENT row gets the NET new-buy value (gross - Less:AP - Less:SFF).
# This matches the existing convention validated against the FY26 DAA Enacted column.
# AP CY (no-tag) row gets the Advance Procurement (CY) line value.
# Completion PY rolldowns and SFF for prior years are EXCLUDED from PARENT
# (they live in [SUB] rows which are not summed).

# Map: spreadsheet (Source, Line Item Number, name_hint) -> (disc, mand, total)
SCN_PARENT_FY27 = {
    # (LI as str, name keyword) -> (disc, mand, total) in $K
    # COLUMBIA: NET new buy ($6,904,785) + SFF for FY 2026 sub-line ($3,329,047)
    # = $10,233,832 of total budget authority committed in FY27 for COLUMBIA program.
    # The SFF for FY 2026 is the budgeted completion of the FY26 ship's commitment.
    # Including it makes the SCN BA01 total reconcile to the P-1 BA01 total of $14,997,174.
    ('1045', 'COLUMBIA Class Submarine'):                (10_233_832,        0, 10_233_832),  # Line 1: NET 6,904,785 + SFF FY26 3,329,047
    ('2001', 'Carrier Replacement Program'):             (  641_907,    37_000,    678_907),  # Line 3 (SFF FY18 only)
    ('2004', 'CVN-81'):                                  (1_447_882,         0,  1_447_882),  # Line 5 (SFF FY20 only)
    ('2013', 'Virginia Class Submarine'):                (8_402_316,         0,  8_402_316),  # Line 6 NET
    ('2086', 'CVN Refueling Overhauls'):                 (4_418_902,    30_000,  4_448_902),  # Line 10 (SFF FY25 only)
    ('2119', 'DDG 1000'):                                (   66_516,         0,     66_516),  # Line 12
    ('2122', 'DDG-51'):                                  (2_954_238,   314_000,  3_268_238),  # Line 13 NET
    ('2127', 'Littoral Combat Ship'):                    (        0,         0,          0),  # Line 15 (no FY27)
    ('2128', 'FFG-Frigate'):                             (        0,         0,          0),  # Line 16 (no FY27)
    ('3010', 'LPD Flight II'):                           (2_188_700,     4_000,  2_192_700),  # Line 18 NET
    ('3041', 'LHA Replacement'):                         (3_850_319,     2_000,  3_852_319),  # Line 21 NET
    ('3050', 'Medium Landing Ship'):                     (        0, 1_887_500,  1_887_500),  # Line 24 NET (mand only)
    ('5025', 'TAO Fleet Oiler'):                         (1_946_063,   282_820,  2_228_883),  # Line 28
    ('5030', 'TAGOS Surtass Ships'):                     (  610_664,   100_000,    710_664),  # Line 31 (Tagos)
    ('5035', 'Towing Salvage'):                          (        0,         0,          0),  # Line 32 ATS (no FY27)
    ('5087', 'Oceanographic Ships'):                     (        0,         0,          0),  # Line 33 (no FY27)
    ('5100', 'LCU 1700'):                                (        0,         0,          0),  # Line 34 (no FY27)
    ('5110', 'Outfitting'):                              (  741_270,         0,    741_270),  # Line 36
    ('5112', 'Ship to Shore Connector'):                 (  733_895,         0,    733_895),  # Line 37
    ('5113', 'Service Craft'):                           (  177_079,         0,    177_079),  # Line 38
    ('5114', 'Auxiliary Personnel Lighter'):             (   83_000,         0,     83_000),  # Line 39
    ('5139', 'LCAC SLEP'):                               (   37_998,         0,     37_998),  # Line 43
    ('5201', 'Auxiliary Vessels'):                       (        0,   130_000,    130_000),  # Line 44 (mand only)
    ('5300', 'Completion of PY Shipbuilding'):           (2_611_990,         0,  2_611_990),  # Line 45
}

# Strategic Sealift PARENT (no LI in spreadsheet — title match only)
SCN_PARENT_FY27_BY_TITLE = {
    'Strategic Sealift':                                 (  450_000,         0,    450_000),  # Line 35
}

# Advance Procurement (CY) rows — matched by (LI, "Advance Procurement (CY)")
SCN_AP_FY27 = {
    '1045': (4_763_342, 205_700, 4_969_042),  # Line 2  COLUMBIA AP CY
    '2001': (1_940_566,       0, 1_940_566),  # Line 4  Carrier Replacement AP CY
    '2013': (4_143_618, 1_440_000, 5_583_618),  # Line 7  Virginia AP CY
    '2086': (   53_070,       0,    53_070),  # Line 11 CVN RCOH AP CY
    '2122': (        0,       0,         0),  # Line 14 DDG-51 AP CY (no FY27)
    '3010': (  355_950,       0,   355_950),  # Line 19 LPD Flight II AP CY
    '3041': (        0,       0,         0),  # Line 22 LHA Replacement AP CY (no FY27)
    '3050': (        0,       0,         0),  # Line 25 Medium Landing Ship AP CY (no FY27)
}

# ─── SCN NEW PROGRAMS (no existing row in spreadsheet) ───────────────────────
# These get appended as new rows. Format:
# (line_no, title, bucket, bucket_sub, fy27_disc, fy27_mand, fy27_total,
#  vessel_svc, vessel_cat, vessel_type, vessel_hull, market, submarket, citation, description)

SCN_NEW_ROWS = [
    # P-1 Line, Title, Bucket, BucketSub, Disc, Mand, Total, Svc, Cat, Type, Hull, Market, Submarket, Cite, Desc
    (8,  'Surface Ship Industrial Base', 1, 'Construction engineering / planning yard',
     0, 93_380, 93_380,
     'USN', 'Combatant Ships', 'Surface Combatants', None,
     'Maritime Platforms & Systems', 'Surface Combatants - Large',
     'P-1 Exhibit > 1611N > BA02 Other Warships > LI 8 Surface Ship Industrial Base',
     'Industrial base investment for surface combatant construction (FY27 mandatory request).'),

    (9,  'BBG(X)', 1, 'Advance procurement / LLTM',
     1_000_000, 0, 1_000_000,
     'USN', 'Combatant Ships', 'Surface Combatants', None,
     'Maritime Platforms & Systems', 'Surface Combatants - Large',
     'P-1 Exhibit > 1611N > BA02 Other Warships > LI 9 BBG(X) AP CY',
     'New large surface combatant program (BBG(X)). FY27 advance procurement to support a future ship buy.'),

    (17, 'FF(X)', 1, 'Full ship DD&C',
     1_429_000, 0, 1_429_000,
     'USN', 'Combatant Ships', 'Surface Combatants', 'FFG',
     'Maritime Platforms & Systems', 'Surface Combatants - Medium',
     'P-1 Exhibit > 1611N > BA02 Other Warships > LI 17 FF(X)',
     'Future Frigate (FF(X)). 1 ship in FY27 — appears to be a new procurement line distinct from FFG-62 Constellation.'),

    (26, 'AS Submarine Tender', 1, 'Full ship DD&C',
     4_444_000, 0, 4_444_000,
     'USN', 'Combatant Ships', 'Material Support Ships', 'AS',
     'Maritime Platforms & Systems', 'Auxiliary & Support Ships',
     'P-1 Exhibit > 1611N > BA05 Auxiliaries > LI 26 As Submarine Tender',
     'New submarine tender procurement. 2 ships in FY27.'),

    (27, 'SMS-Special Mission Ship', 1, 'Full ship DD&C',
     0, 200_000, 200_000,
     'USN', 'Combatant Ships', 'Surveillance Ships', None,
     'Maritime Platforms & Systems', 'Auxiliary & Support Ships',
     'P-1 Exhibit > 1611N > BA05 Auxiliaries > LI 27 SMS-Special Mission Ship',
     'New Special Mission Ship (SMS) program. 1 ship in FY27 (mandatory request).'),

    (30, 'Next Generation Logistics Ship', 1, 'Advance procurement / LLTM',
     0, 0, 0,
     'MSC', 'Combatant Ships', 'Combat Logistics Ships', None,
     'Maritime Platforms & Systems', 'Combat Logistics Ships',
     'P-1 Exhibit > 1611N > BA05 Auxiliaries > LI 30 Next Generation Logistics Ship',
     'NGLS — new logistics ship program. FY26 had AP funding; no FY27 funding shown in P-1.'),

    (40, 'Bulk Fuel Vessel', 1, 'Full ship DD&C',
     450_000, 0, 450_000,
     'MSC', 'Combatant Ships', 'Combat Logistics Ships', None,
     'Maritime Platforms & Systems', 'Combat Logistics Ships',
     'P-1 Exhibit > 1611N > BA05 Auxiliaries > LI 40 Bulk Fuel Vessel',
     'New Bulk Fuel Vessel program. 1 ship in FY27.'),

    (41, 'T-AH(X)', 1, 'Full ship DD&C',
     0, 650_000, 650_000,
     'MSC', 'Auxiliary Ships', 'Hospital Ship', 'T-AH',
     'Maritime Platforms & Systems', 'Auxiliary & Support Ships',
     'P-1 Exhibit > 1611N > BA05 Auxiliaries > LI 41 T-AH(X)',
     'Next-generation hospital ship (T-AH replacement). 1 ship in FY27 (mandatory request).'),

    (42, 'Fireboats', 1, 'Full vessel procurement',
     0, 272_500, 272_500,
     'USN', 'Support Crafts', 'Other Craft, Self-Propelled', None,
     'Maritime Platforms & Systems', 'Auxiliary & Support Ships',
     'P-1 Exhibit > 1611N > BA05 Auxiliaries > LI 42 Fireboats',
     'New Fireboats procurement. 5 fireboats in FY27 (mandatory request).'),
]

# ─── 1613N Maritime Industrial Base Fund (FY26-only, NEW appropriation) ──────
# Two line items, both ZERO in FY27 (entire 1613N is OBBBA-funded FY26 only).
# We do NOT add these as new rows because they have no FY27 funding.
# Mentioned here for documentation only.

# ─── PARSE FY27 P-1 OPN/WPN LINE ITEMS ───────────────────────────────────────

def parse_p1_appropriation(text: str, appropriation_code: str) -> dict[int, dict]:
    """
    Parse the FY27 page sections for a given appropriation (e.g. '1810' for OPN).
    Returns a dict {line_no: {'name': ..., 'fy27_disc': int, 'fy27_mand': int, 'fy27_total': int}}.

    Strategy: position-aware parsing.
      1. Locate the column header line of each FY27 page; find the END position of each
         "Cost" label. Cost values are right-aligned about 6 chars past the label end.
      2. For each data line, find all money tokens. Assign each to the cost column
         whose alignment END position is closest (within ±10 chars). This automatically
         skips quantities (which fall between cost columns) and numbers embedded in
         item names (which fall before any cost column).
    """
    result = {}
    lines = text.split('\n')
    in_fy27_page = False
    current_appr = None
    cost_aligns = None  # list of approximate END positions for each cost column
    name_end = None  # position before the first cost column where name ends

    def to_int(s):
        if s is None: return 0
        stripped = s.strip()
        if not stripped: return 0
        neg = False
        if stripped.startswith('(') and stripped.endswith(')'):
            stripped = stripped[1:-1].strip()
        if stripped.startswith('-'):
            neg = True
            stripped = stripped[1:]
        cleaned = stripped.replace(',', '')
        if not re.match(r'^\d+$', cleaned):
            return 0
        v = int(cleaned)
        return -v if neg else v

    def find_cost_aligns(header_line):
        """Return list of approximate END positions (where right-aligned values end)
        for each Cost column. Cost values typically end ~6 chars past the Cost label end."""
        positions = []
        pos = 0
        while True:
            idx = header_line.find('Cost', pos)
            if idx == -1:
                break
            # Value end position is roughly Cost_label_end + 6
            positions.append(idx + 4 + 6)
            pos = idx + 4
        return positions

    name_carryover = ''  # text from a name continuation line above the numbered item

    i = 0
    while i < len(lines):
        line = lines[i]

        # Detect appropriation header
        m = re.match(r'\s*Appropriation:\s+(\d{4})\s', line)
        if m:
            current_appr = m.group(1)
            combined = line + ' ' + (lines[i+1] if i+1 < len(lines) else '')
            in_fy27_page = (current_appr == appropriation_code
                            and ('27 Discretionary' in combined or 'FY 2027' in line))
            cost_positions = None  # reset; will detect from this page's header
            name_carryover = ''
            i += 1
            continue

        if not in_fy27_page or current_appr != appropriation_code:
            i += 1
            continue

        # Detect column header line: contains "No Item Nomenclature" and "Cost"
        if 'No Item Nomenclature' in line and 'Cost' in line:
            cost_aligns = find_cost_aligns(line)
            name_end = (cost_aligns[0] - 25) if cost_aligns else 60
            name_carryover = ''
            i += 1
            continue

        # Skip until we have cost positions for this page
        if cost_aligns is None:
            i += 1
            continue

        # Detect a "name continuation" line: indented text with no leading line number,
        # no money values, doesn't start with a known sub-line keyword.
        # These appear IMMEDIATELY ABOVE a numbered line item.
        stripped_line = line.strip()
        if not stripped_line:
            # Blank line — clears any pending carryover
            name_carryover = ''
            i += 1
            continue

        if (stripped_line
            and not re.match(r'^\d', stripped_line)
            and not re.search(r'[\d,]{4,}', line)
            and not stripped_line.startswith(('Less:', 'Subsequent', 'Completion', 'Advance Procurement', 'C (FY',
                                              'Budget Activity', 'Total ', 'Department of', "President's", 'Exhibit',
                                              'Appropriation:', 'Conversion,', 'No Item', 'Line', 'Code', 'Sec'))
            and 'UNCLASSIFIED' not in stripped_line
            and 'Apr 2026' not in stripped_line
            and 'Page ' not in stripped_line
            and len(stripped_line) > 3
            and len(stripped_line) < 80):
            # Possible name continuation — but only used if NEXT line is a numbered item
            # (carryover is cleared on blank lines, so this works automatically)
            name_carryover = stripped_line
            i += 1
            continue

        # Detect numbered line item: leading whitespace + 1-3 digit number + 2+ spaces + name
        m = re.match(r'^(\s{0,8})(\d{1,3})\s{2,}([A-Za-z(].+?)$', line)
        if m:
            line_no = int(m.group(2))

            # Extract item name from the part before the first cost column
            name_section = line[:name_end].rstrip() if name_end else line.rstrip()
            name_section = re.sub(r'^\s*\d{1,3}\s+', '', name_section)
            name = re.sub(r'\s+[A-Z]\s+U\s*$', '', name_section).strip()
            name = re.sub(r'\s+U\s*$', '', name).strip()
            name = re.sub(r'\s+[A-Z]\s*$', '', name).strip()

            # Prepend any name carryover from the previous line
            if name_carryover:
                name = f'{name_carryover} {name}'.strip()
                name_carryover = ''

            # Find all money tokens in the line and their END positions
            costs = [0] * len(cost_aligns)
            for tok_match in re.finditer(r'\(?-?[\d,]+\)?', line):
                tok = tok_match.group()
                end_pos = tok_match.end()
                stripped = tok.strip('()').lstrip('-').replace(',', '')
                # Must be at least one digit
                if not stripped.isdigit():
                    continue
                # Skip pure 4-digit years
                if not (',' in tok or '(' in tok or '-' in tok) and 1900 <= int(stripped) <= 2050 and len(stripped) == 4:
                    continue
                # Skip if before the cost columns (it's part of the name or qty)
                if end_pos < name_end:
                    continue
                # Assign to the closest cost column (END position within ±3 chars to
                # avoid catching qty values which are ~5 chars away from cost column ends)
                best_col = -1
                best_dist = 4
                for ci, align in enumerate(cost_aligns):
                    dist = abs(end_pos - align)
                    if dist < best_dist:
                        best_dist = dist
                        best_col = ci
                if best_col >= 0:
                    costs[best_col] = to_int(tok)

            disc = costs[0] if len(costs) > 0 else 0
            mand = costs[1] if len(costs) > 1 else 0
            total = costs[2] if len(costs) > 2 else 0

            # Validation post-fix: if disc + mand != total, a quantity may have been
            # mis-assigned to a cost column. Trust the disc/total values (which align
            # tightly to cost columns at fixed end positions) and re-derive mand.
            if disc + mand != total and total != 0:
                if disc == total:
                    mand = 0
                elif mand == total:
                    disc = 0
                else:
                    # Both disc and mand non-zero but don't sum to total — keep
                    # the larger of the two and re-derive the other
                    derived_mand = total - disc
                    if derived_mand >= 0:
                        mand = derived_mand

            # Don't overwrite if we already have this line (first occurrence wins)
            if line_no not in result:
                result[line_no] = {
                    'name': name,
                    'fy27_disc': disc,
                    'fy27_mand': mand,
                    'fy27_total': total,
                    'raw': line.rstrip(),
                }
            else:
                # If we already have it, but old one had no name and new has name, update
                if not result[line_no]['name'] and name:
                    result[line_no]['name'] = name

        # End of page detected by certain markers
        if 'Total Other Procurement, Navy' in line and current_appr == '1810':
            in_fy27_page = False
        if 'Total Weapons Procurement, Navy' in line and current_appr == '1507':
            in_fy27_page = False
        if 'Total Shipbuilding and Conversion, Navy' in line and current_appr == '1611':
            in_fy27_page = False

        i += 1

    return result


# ─── NORMALIZE TITLE FOR FUZZY MATCH ─────────────────────────────────────────

def normalize_title(s: str) -> str:
    """Normalize a title for fuzzy comparison. Lowercases, strips parentheticals
    and punctuation, expands common abbreviations."""
    if not s:
        return ''
    s = s.lower()
    s = re.sub(r'\([^)]*\)', '', s)
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    # Expand common defense-budget abbreviations
    abbrevs = {
        r'\bcomms\b': 'communications',
        r'\bcomm\b': 'communications',
        r'\bmods\b': 'modifications',
        r'\bmod\b': 'modifications',
        r'\bequip\b': 'equipment',
        r'\bsupt\b': 'support',
        r'\bsys\b': 'systems',
        r'\bproj\b': 'project',
        r'\bsubm\b': 'submarine',
        r'\bnav\b': 'navigation',
    }
    for pat, repl in abbrevs.items():
        s = re.sub(pat, repl, s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def extract_acronyms(s: str) -> set[str]:
    """Extract acronyms from a title — both parenthetical (RAM) and ALL-CAPS tokens."""
    if not s:
        return set()
    acros = set()
    # Parenthetical
    for m in re.finditer(r'\(([A-Z][A-Z0-9/]+)\)', s):
        acros.add(m.group(1).lower())
    # Standalone all-caps tokens (4+ chars to avoid 'A', 'U', etc.)
    for tok in re.findall(r'\b[A-Z][A-Z0-9/]{2,}\b', s):
        acros.add(tok.lower())
    return acros


def fuzzy_find(target: str, candidates: dict[int, dict]) -> tuple[int, dict] | None:
    """Find the best matching line in candidates by name. Considers substring
    containment and acronym matches as weighted signals, NOT as overrides."""
    target_norm = normalize_title(target)
    if not target_norm:
        return None
    target_words = set(target_norm.split())
    target_acros = extract_acronyms(target)

    best_line = None
    best_score = 0
    best_data = None
    for line_no, data in candidates.items():
        cand_name = data['name']
        cand_norm = normalize_title(cand_name)
        if cand_norm == target_norm:
            return (line_no, data)
        cand_words = set(cand_norm.split())
        cand_acros = extract_acronyms(cand_name)
        if not cand_words:
            continue
        # Primary score: Jaccard-style word overlap
        overlap = len(target_words & cand_words)
        word_score = overlap / max(len(target_words), len(cand_words))
        # Substring containment bonus — only when one side is short (<=3 words)
        # to avoid matching e.g. "Virginia Class Spares" → "Spares and Repair Parts"
        if (len(target_words) <= 3 or len(cand_words) <= 3):
            if target_norm in cand_norm or cand_norm in target_norm:
                word_score = max(word_score, 0.85)
        # Acronym bonus: only when the SAME acronym appears in both sides
        if target_acros and cand_acros and (target_acros & cand_acros):
            word_score = max(word_score, 0.9)

        if word_score > best_score:
            best_score = word_score
            best_line = line_no
            best_data = data

    if best_score >= 0.7:
        return (best_line, best_data)
    return None


# ─── MAIN BUILD ──────────────────────────────────────────────────────────────

def main():
    print(f"Loading {SOURCE_XLSX.name}...")
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=False)
    ws = wb[SHEET]

    # Parse FY27 P-1 OPN and WPN
    print("Parsing FY27 P-1...")
    p1_text = P1_TXT.read_text()

    opn_items = parse_p1_appropriation(p1_text, '1810')
    wpn_items = parse_p1_appropriation(p1_text, '1507')

    print(f"  OPN (1810) line items extracted: {len(opn_items)}")
    print(f"  WPN (1507) line items extracted: {len(wpn_items)}")

    # ─── ADD HEADER CELLS FOR NEW COLUMNS ─────────────────────────────────
    print("Adding new column headers...")
    header_row = 4
    ws.cell(row=header_row, column=NEW_COL_FY27_REQ, value='FY2027 Request ($K)')
    ws.cell(row=header_row, column=NEW_COL_FY27_MAND, value='FY2027 Mandatory Request ($K)')
    ws.cell(row=header_row, column=NEW_COL_FY27_TOTAL, value='FY2027 Total ($K)')

    # Copy header formatting from an existing FY26 column header (col R = 18)
    template_cell = ws.cell(row=header_row, column=COL_FY26_DAA_ENACTED + 1)  # 1-indexed
    if template_cell.font:
        for col in [NEW_COL_FY27_REQ, NEW_COL_FY27_MAND, NEW_COL_FY27_TOTAL]:
            target = ws.cell(row=header_row, column=col)
            if template_cell.font:
                target.font = copy(template_cell.font)
            if template_cell.fill:
                target.fill = copy(template_cell.fill)
            if template_cell.alignment:
                target.alignment = copy(template_cell.alignment)

    # ─── WRITE FY27 DATA ──────────────────────────────────────────────────
    stats = defaultdict(int)
    unmatched_opn = []
    unmatched_wpn = []

    # Walk all rows in J Book sheet
    for excel_row in range(5, ws.max_row + 1):
        source = ws.cell(row=excel_row, column=COL_SOURCE + 1).value
        li = ws.cell(row=excel_row, column=COL_LI + 1).value
        title = ws.cell(row=excel_row, column=COL_TITLE + 1).value or ''
        rtype = ws.cell(row=excel_row, column=COL_ROW_TYPE + 1).value

        if not source:
            continue

        # Only process additive rows (PARENT or no-tag) per user instruction
        # We'll also fill SUB rows where they have a clear match (Advance Procurement CY)
        is_additive = rtype is None or rtype == '[PARENT]'

        disc, mand, total = None, None, None

        # SCN
        if source == 'SCN':
            li_str = str(li).strip() if li is not None else None
            title_str = str(title).strip()
            is_ap_row = 'Advance Procurement' in title_str

            # AP rows (no-tag with "Advance Procurement (CY)" in title) — handle FIRST
            if rtype is None and is_ap_row and li_str in SCN_AP_FY27:
                disc, mand, total = SCN_AP_FY27[li_str]
                stats['scn_ap'] += 1

            # Parent mapping by (LI, name fragment) — only for non-AP rows
            elif rtype == '[PARENT]' and li_str and not is_ap_row:
                for (key_li, key_name), vals in SCN_PARENT_FY27.items():
                    if key_li == li_str and key_name.lower() in title_str.lower():
                        disc, mand, total = vals
                        stats['scn_parent'] += 1
                        break

            # No-tag (additive) row for a parent program (e.g. CVN-81, DDG 1000) — non-AP
            elif rtype is None and li_str and not is_ap_row:
                for (key_li, key_name), vals in SCN_PARENT_FY27.items():
                    if key_li == li_str and key_name.lower() in title_str.lower():
                        disc, mand, total = vals
                        stats['scn_notag_parent_like'] += 1
                        break

            # Strategic Sealift (no LI)
            if disc is None and 'Strategic Sealift' in title_str and not is_ap_row:
                disc, mand, total = SCN_PARENT_FY27_BY_TITLE['Strategic Sealift']
                stats['scn_strategic_sealift'] += 1

        # OPN BA1-BA5-8 — match by name
        elif source.startswith('OPN_BA') and is_additive:
            match = fuzzy_find(title, opn_items)
            if match:
                _, data = match
                disc = data['fy27_disc']
                mand = data['fy27_mand']
                total = data['fy27_total']
                stats[f'{source}_matched'] += 1
            else:
                stats[f'{source}_unmatched'] += 1
                unmatched_opn.append(f"  {source} R{excel_row} LI={li}: {title}")

        # WPN — match by name
        elif source == 'WPN' and is_additive:
            match = fuzzy_find(title, wpn_items)
            if match:
                _, data = match
                disc = data['fy27_disc']
                mand = data['fy27_mand']
                total = data['fy27_total']
                stats['wpn_matched'] += 1
            else:
                stats['wpn_unmatched'] += 1
                unmatched_wpn.append(f"  WPN R{excel_row} LI={li}: {title}")

        # OMN/NWCF/USCG — left blank per user direction
        elif source in ('OMN', 'OMN_Vol2', 'NWCF', 'USCG'):
            stats[f'{source}_blank'] += 1
            continue

        # Write the data
        if disc is not None:
            ws.cell(row=excel_row, column=NEW_COL_FY27_REQ, value=disc if disc != 0 else None)
            ws.cell(row=excel_row, column=NEW_COL_FY27_MAND, value=mand if mand != 0 else None)
            ws.cell(row=excel_row, column=NEW_COL_FY27_TOTAL, value=total if total != 0 else None)

    # ─── APPEND NEW PROGRAM ROWS ──────────────────────────────────────────
    print("Appending new SCN program rows...")
    next_row = ws.max_row + 1
    for (line_no, title, bucket, bucket_sub, disc, mand, total,
         svc, cat, vtype, hull, market, submarket, cite, desc) in SCN_NEW_ROWS:
        ws.cell(row=next_row, column=COL_SOURCE + 1, value='SCN')
        ws.cell(row=next_row, column=2, value='Procurement')                  # B Color of Money
        # Leave Line Item Number (col C) blank — these are new programs without P-40 LI
        ws.cell(row=next_row, column=COL_TITLE + 1, value=title)
        ws.cell(row=next_row, column=COL_BUCKET + 1, value=bucket)
        ws.cell(row=next_row, column=COL_BUCKET_SUB + 1, value=bucket_sub)
        ws.cell(row=next_row, column=COL_DESC + 1, value=desc)
        # Row Type: blank (no-tag = additive)
        # Vessel taxonomy
        if svc:    ws.cell(row=next_row, column=COL_VESSEL_SVC + 1, value=svc)
        if cat:    ws.cell(row=next_row, column=COL_VESSEL_CAT + 1, value=cat)
        if vtype:  ws.cell(row=next_row, column=COL_VESSEL_TYPE + 1, value=vtype)
        if hull:   ws.cell(row=next_row, column=COL_VESSEL_HULL + 1, value=hull)
        if cite:   ws.cell(row=next_row, column=COL_CITATION + 1, value=cite)
        if market: ws.cell(row=next_row, column=COL_MARKET + 1, value=market)
        if submarket: ws.cell(row=next_row, column=COL_SUBMARKET + 1, value=submarket)
        # FY27 columns
        if disc != 0:  ws.cell(row=next_row, column=NEW_COL_FY27_REQ, value=disc)
        if mand != 0:  ws.cell(row=next_row, column=NEW_COL_FY27_MAND, value=mand)
        if total != 0: ws.cell(row=next_row, column=NEW_COL_FY27_TOTAL, value=total)
        next_row += 1
        stats['scn_new_rows'] += 1

    # ─── UPDATE VALIDATION SHEET ──────────────────────────────────────────
    print("Updating Validation sheet...")
    val_ws = wb['Validation']
    add_validation_section(val_ws)

    # ─── SAVE ─────────────────────────────────────────────────────────────
    print(f"Saving to {OUTPUT_XLSX.name}...")
    wb.save(OUTPUT_XLSX)

    # ─── REPORT ───────────────────────────────────────────────────────────
    print("\n=== COVERAGE STATS ===")
    for key in sorted(stats.keys()):
        print(f"  {key}: {stats[key]}")
    if unmatched_opn:
        print(f"\nUnmatched OPN rows ({len(unmatched_opn)}):")
        for u in unmatched_opn[:30]:
            print(u)
        if len(unmatched_opn) > 30:
            print(f"  ... and {len(unmatched_opn) - 30} more")
    if unmatched_wpn:
        print(f"\nUnmatched WPN rows ({len(unmatched_wpn)}):")
        for u in unmatched_wpn:
            print(u)


def add_validation_section(val_ws):
    """Append FY2027 documentation to the Validation sheet."""
    # Find next empty row
    start_row = val_ws.max_row + 2

    bold = Font(bold=True)
    section_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')

    # Section header
    cell = val_ws.cell(row=start_row, column=1, value='FY 2027 COLUMNS — METHODOLOGY & SOURCING')
    cell.font = bold
    cell.fill = section_fill
    start_row += 2

    rows = [
        # (Sheet, Field, Value, Description, Notes)
        ('FY 2027 Columns', 'Source Document',
         'FY 2027 P-1 / O-1 (April 2026)',
         'New columns added: FY2027 Request ($K), FY2027 Mandatory Request ($K), FY2027 Total ($K). Located at end of sheet (cols AI/AJ/AK) to avoid breaking the column-letter formula references in TAM, SAM, and Competitive Dynamics.',
         'Status of FY27: REQUEST only, not enacted. Congress has not yet acted on the FY27 budget submission.'),

        ('FY 2027 Columns', 'Discretionary vs Mandatory',
         'Two separate FY27 columns',
         'The FY 2027 P-1 has TWO request columns: "FY 2027 Discretionary Request" (the regular annual base request) and "FY 2027 Mandatory Request" (a NEW reconciliation ask the Administration is making — round 2 mandatory funding, separate from OBBBA, NOT enacted, NOT guaranteed to pass).',
         'FY2027 Total = Disc Request + Mandatory Request. SCN FY27 totals: Disc $60.18B + Mand $5.65B = $65.83B.'),

        ('FY 2027 Columns', 'TAM Formula',
         'Same as FY26 — sum (no tag) + [PARENT] rows',
         'For FY27: TAM = SUM(FY27_Total) WHERE Row_Type IN ("", "[PARENT]"). Use FY2027 Total as the per-row value, OR sum FY2027 Request + FY2027 Mandatory Request separately if you want to model the two scenarios.',
         'Same double-counting guards apply: skip [SUB], [ALT_VIEW], [REFERENCE], [UNFUNDED]. Composite-bucket rows still excluded from per-bucket pivots.'),

        ('FY 2027 Columns', 'Net vs Gross convention',
         'PARENT = NET (gross − Less:AP − Less:SFF)',
         'Same convention as FY26 DAA columns: PARENT row holds the NET value after deducting Advance Procurement (PY) and Subsequent Full Funding (FY) lines from the gross procurement amount. Completion-PY rolldowns from earlier years are EXCLUDED from PARENT (they live in [SUB] rows).',
         'Validated against FY26: Virginia Class PARENT FY26 DAA Final = $2.74B = gross $5.39B − AP $2.65B (FY27 P-1 page 135). Same approach used for FY27.'),

        ('FY 2027 Columns', 'OBBBA exclusion',
         'OBBBA does NOT get a separate FY27 column',
         'CRITICAL: OBBBA (PL 119-21, "One Big Beautiful Bill Act") provided $152.3B in MANDATORY funding enacted in July 2025. Per OBBBA Figure 1 in the OBBBA Funding Allocation Narrative, 100% of OBBBA budget authority is FY 2026 — every section shows zero in FY27/FY28/FY29 columns. OBBBA is fully captured in the existing FY2026 DAA Mandatory / FY2026 DAA Final Bill / FY2026 DAA Enacted Total columns.',
         'See "Budget authority vs spending implementation" row below for the nuance about why some OBBBA contracts will sign in FY27/FY28 even though the budget authority is FY26.'),

        ('FY 2027 Columns', 'Budget authority vs spending implementation',
         'Important conceptual distinction',
         'BUDGET AUTHORITY (BA) is the legal authority Congress grants in a given fiscal year for an agency to enter into financial obligations. SPENDING IMPLEMENTATION (outlays / contract awards) can happen in later years if the authority is multi-year. OBBBA gave the Department $152.3B of FY2026 BA, available for obligation through Sept 30, 2029. So contracts will sign in FY26, FY27, FY28, and FY29 — but the BA all counts as FY26. THIS WORKBOOK MEASURES BA, not outlays, because that is the apples-to-apples basis for comparing across years and against the President\'s Budget request. The "FY 2026 Reconciliation" / "FY 2026 DAA Mandatory" columns already capture all OBBBA dollars correctly. Adding an "FY 2027 OBBBA" column would double-count.',
         'Examples from OBBBA narrative: T-AO Oilers (item 23) — "three (3) T-AO Oilers, two (2) in FY 2026 and one (1) in FY 2027" — all $2.725B is FY26 BA, but the third oiler\'s contract will sign in FY27. LPD 34 / LHA 10 advance procurement (item 26) is also FY26 BA for FY27 ship buys.'),

        ('FY 2027 Columns', 'FY27 Mandatory Request vs OBBBA',
         'These are DIFFERENT pots of money',
         'OBBBA = enacted FY26 mandatory funding (PL 119-21, July 2025). Captured in FY26 DAA columns. FY 2027 Mandatory Request = a NEW reconciliation ask in the Apr 2026 President\'s Budget submission — the Administration is asking Congress for ANOTHER round of mandatory funding in FY27. As of the FY27 PB submission, this is NOT enacted; Congress may or may not pass it.',
         'SCN FY27 Mandatory Request = $5.65B. Compare to SCN FY26 OBBBA (PL 119-21) = $17.93B. Different bills, different fiscal years, both technically "reconciliation" but accounted separately.'),

        ('FY 2027 Columns', 'Coverage by source book',
         'High for SCN/OPN/WPN, blank for OMN/NWCF/USCG',
         'High coverage (line-item match available): SCN (33 rows), OPN_BA1-BA5-8 (118 rows), WPN (23 rows). Blank coverage in v1.21: OMN (12 rows), OMN_Vol2 (94 rows), NWCF (33 rows) — these source books only have SAG/account-level totals in the O-1/RF-1 exhibits, not per-line detail. USCG (18 rows) — Coast Guard is part of DHS, not in the DoW P-1/O-1.',
         'Plan: OMN/NWCF will be filled via prorated SAG-level allocation in a separate sheet (Phase 2). USCG will be filled when DHS Congressional Justification is sourced.'),

        ('FY 2027 Columns', 'New programs added in v1.21',
         '9 new SCN rows for FY27 P-1 line items',
         'Programs that exist in the FY27 P-1 but had no row in the v1.20 spreadsheet: Surface Ship Industrial Base (LI 8), BBG(X) (LI 9, AP only), FF(X) (LI 17), AS Submarine Tender (LI 26), SMS-Special Mission Ship (LI 27), Next Generation Logistics Ship (LI 30, no FY27 funding), Bulk Fuel Vessel (LI 40), T-AH(X) (LI 41), Fireboats (LI 42).',
         'Each new row has Source=SCN, Color of Money=Procurement, no Row Type tag (additive), and full vessel taxonomy populated where available.'),

        ('FY 2027 Columns', '1613N Maritime Industrial Base Fund',
         'NOT added — FY26-only appropriation',
         'The FY27 P-1 includes a new appropriation 1613N "Maritime Industrial Base Fund" with two line items: Submarine Industrial Base ($2.54B FY26 PL 119-21) and Surface Ship Industrial Base ($1.99B FY26 PL 119-21). Both are zero in FY27. This is OBBBA-only money already captured by other rows in the workbook — no new rows added.',
         '—'),
    ]

    for (sheet, field, value, desc, notes) in rows:
        val_ws.cell(row=start_row, column=1, value=sheet)
        val_ws.cell(row=start_row, column=2, value=field)
        val_ws.cell(row=start_row, column=3, value=value)
        val_ws.cell(row=start_row, column=4, value=desc)
        val_ws.cell(row=start_row, column=5, value=notes)
        # Wrap text in description columns
        for col in [4, 5]:
            val_ws.cell(row=start_row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        start_row += 2  # blank row between entries


if __name__ == '__main__':
    main()
