"""Apply DD&C / LLTM breakdown sub-rows to SAM Mekko Table 1 and TAM Mekko Table 6."""
import re
import openpyxl
from openpyxl.utils import column_index_from_string
from copy import copy

XLSX = '/Users/brendantoole/projects2/buildco3/08APR2028_Newbuild_and_MRO_Spend_v1.17.xlsx'
J_BOOK = "'J Book Items Cons.'"


def shift_formula_rows(formula, min_row, shift):
    """Add `shift` to row numbers >= min_row in same-sheet refs. Cross-sheet refs untouched."""
    if not isinstance(formula, str) or not formula.startswith('='):
        return formula

    pattern = re.compile(
        r"(?:('[^']+'|[A-Za-z_]\w*)!)?(\$?[A-Z]+\$?\d+)(?::(\$?[A-Z]+\$?\d+))?"
    )

    def shift_cell(cell_ref):
        m = re.match(r"(\$?)([A-Z]+)(\$?)(\d+)", cell_ref)
        if not m:
            return cell_ref
        col_dol, col, row_dol, row = m.groups()
        row = int(row)
        if row >= min_row:
            return f"{col_dol}{col}{row_dol}{row + shift}"
        return cell_ref

    def replacer(m):
        sheet = m.group(1)
        cell1 = m.group(2)
        cell2 = m.group(3)
        if sheet is not None:
            return m.group(0)  # cross-sheet, leave alone
        if cell2:
            return f"{shift_cell(cell1)}:{shift_cell(cell2)}"
        return shift_cell(cell1)

    return pattern.sub(replacer, formula)


def make_breakdown_formula(vessel_service, vessel_type, sub_category, total_cell):
    """Build the 6-SUMIFS R→N→L coalesce formula for a sub-row cell, divided by total cell."""
    parts = []
    for rt in ['', '[PARENT]']:
        # T1: sum R where R<>""
        parts.append(
            f'SUMIFS({J_BOOK}!R$5:R$3360'
            f',{J_BOOK}!E$5:E$3360,1'
            f',{J_BOOK}!F$5:F$3360,"{sub_category}"'
            f',{J_BOOK}!S$5:S$3360,"{rt}"'
            f',{J_BOOK}!U$5:U$3360,"{vessel_service}"'
            f',{J_BOOK}!W$5:W$3360,"{vessel_type}"'
            f',{J_BOOK}!R$5:R$3360,"<>")'
        )
        # T2: sum N where R="" and N<>""
        parts.append(
            f'SUMIFS({J_BOOK}!N$5:N$3360'
            f',{J_BOOK}!E$5:E$3360,1'
            f',{J_BOOK}!F$5:F$3360,"{sub_category}"'
            f',{J_BOOK}!S$5:S$3360,"{rt}"'
            f',{J_BOOK}!U$5:U$3360,"{vessel_service}"'
            f',{J_BOOK}!W$5:W$3360,"{vessel_type}"'
            f',{J_BOOK}!R$5:R$3360,""'
            f',{J_BOOK}!N$5:N$3360,"<>")'
        )
        # T3: sum L where R="" and N=""
        parts.append(
            f'SUMIFS({J_BOOK}!L$5:L$3360'
            f',{J_BOOK}!E$5:E$3360,1'
            f',{J_BOOK}!F$5:F$3360,"{sub_category}"'
            f',{J_BOOK}!S$5:S$3360,"{rt}"'
            f',{J_BOOK}!U$5:U$3360,"{vessel_service}"'
            f',{J_BOOK}!W$5:W$3360,"{vessel_type}"'
            f',{J_BOOK}!R$5:R$3360,""'
            f',{J_BOOK}!N$5:N$3360,"")'
        )
    return f'=IF({total_cell}=0,0,({"+".join(parts)})/{total_cell})'


# (col_letter, vessel_service, vessel_type, blank_breakdown)
sam_mekko_columns = [
    ('B', 'USN',  'Surface Combatants',              False),
    ('C', 'USN',  'Amphibious Warfare Ships',        False),
    ('D', 'USN',  'Expeditionary & Seabasing Ships', False),
    ('E', 'MSC',  'Combat Logistics Ships',          False),
    ('F', 'USN',  'Unmanned Surface Vehicles',       True),   # Full vessel procurement
    ('G', 'USN',  'Amphibious Warfare Craft',        False),
    ('H', 'MSC',  'Cargo & Vehicle Cargo Ships',     False),
    ('I', 'MSC',  'Surveillance Ships',              False),
    ('J', 'USN',  'Command Ships',                   False),  # 0 NewCon
    ('K', 'USN',  'Mine Warfare',                    False),  # 0 NewCon
    ('L', 'USCG', 'Offshore Patrol Cutter',          True),   # Combined bucket
    ('M', 'USCG', 'Fast Response Cutters',           False),
    ('N', 'USCG', 'Icebreakers, Oceangoing',         True),   # Combined bucket
    ('O', 'USCG', 'Waterways Commerce Cutters',      True),   # Combined bucket
    ('P', 'USCG', 'National Security Cutters',       False),  # 0 NewCon
]

tam_mekko_columns = [
    ('B', 'USN',  'Submarines',                      False),
    ('C', 'USN',  'Surface Combatants',              False),
    ('D', 'USN',  'Aircraft Carriers',               False),
    ('E', 'USN',  'Amphibious Warfare Ships',        False),
    ('F', 'USN',  'Expeditionary & Seabasing Ships', False),
    ('G', 'MSC',  'Combat Logistics Ships',          False),
    ('H', 'USN',  'Unmanned Surface Vehicles',       True),
    ('I', 'USN',  'Amphibious Warfare Craft',        False),
    ('J', 'MSC',  'Cargo & Vehicle Cargo Ships',     False),
    ('K', 'MSC',  'Surveillance Ships',              False),
    ('L', 'USN',  'Unmanned Undersea Vehicles',      True),
    ('M', 'USN',  'Command Ships',                   False),
    ('N', 'USN',  'Mine Warfare',                    False),
    ('O', 'USCG', 'Offshore Patrol Cutter',          True),
    ('P', 'USCG', 'Fast Response Cutters',           False),
    ('Q', 'USCG', 'Icebreakers, Oceangoing',         True),
    ('R', 'USCG', 'Waterways Commerce Cutters',      True),
    ('S', 'USCG', 'National Security Cutters',       False),
]


def insert_breakdown_rows(wb, sheet_name, newcon_row, total_row_after_insert,
                           total_col_letter, columns_config):
    """Insert two breakdown sub-rows directly after newcon_row."""
    ws = wb[sheet_name]
    insert_at = newcon_row + 1  # rows shift starting here

    # Step 1: shift all existing same-sheet formula references with row >= insert_at by +2
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell.value = shift_formula_rows(cell.value, insert_at, 2)

    # Step 2: physically insert 2 blank rows
    ws.insert_rows(insert_at, 2)

    # Step 3: copy number formats from the NewCon row above for the new rows' data cells
    for col_letter, _, _, _ in columns_config + [(total_col_letter, '', '', False)]:
        col_idx = column_index_from_string(col_letter)
        src = ws.cell(row=newcon_row, column=col_idx)
        for offset in (0, 1):
            dst = ws.cell(row=insert_at + offset, column=col_idx)
            dst.number_format = src.number_format
            if src.has_style:
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.protection = copy(src.protection)

    # Step 4: write labels in column A (copy style from NewCon row's A cell)
    a_src = ws.cell(row=newcon_row, column=1)
    a_dd = ws.cell(row=insert_at, column=1, value='  Full Ship DD&C')
    a_lltm = ws.cell(row=insert_at + 1, column=1, value='  Advance Procurement / LLTM')
    if a_src.has_style:
        for c in (a_dd, a_lltm):
            c.font = copy(a_src.font)
            c.fill = copy(a_src.fill)
            c.border = copy(a_src.border)
            c.alignment = copy(a_src.alignment)
            c.number_format = a_src.number_format

    # Step 5: write data cells (formulas or blanks)
    for col_letter, service, vt, blank in columns_config:
        col_idx = column_index_from_string(col_letter)
        total_cell = f"{col_letter}{total_row_after_insert}"
        if blank:
            ws.cell(row=insert_at,     column=col_idx, value=None)
            ws.cell(row=insert_at + 1, column=col_idx, value=None)
        else:
            ws.cell(row=insert_at,     column=col_idx,
                    value=make_breakdown_formula(service, vt, 'Full ship DD&C', total_cell))
            ws.cell(row=insert_at + 1, column=col_idx,
                    value=make_breakdown_formula(service, vt, 'Advance procurement / LLTM', total_cell))

    # Step 6: total column SUMPRODUCT formulas
    first_col = columns_config[0][0]
    last_col = columns_config[-1][0]
    total_col_idx = column_index_from_string(total_col_letter)
    sumprod_dd = (
        f'=IF({total_col_letter}${total_row_after_insert}=0,0,'
        f'SUMPRODUCT({first_col}{insert_at}:{last_col}{insert_at},'
        f'{first_col}${total_row_after_insert}:{last_col}${total_row_after_insert})'
        f'/{total_col_letter}${total_row_after_insert})'
    )
    sumprod_lltm = (
        f'=IF({total_col_letter}${total_row_after_insert}=0,0,'
        f'SUMPRODUCT({first_col}{insert_at + 1}:{last_col}{insert_at + 1},'
        f'{first_col}${total_row_after_insert}:{last_col}${total_row_after_insert})'
        f'/{total_col_letter}${total_row_after_insert})'
    )
    ws.cell(row=insert_at,     column=total_col_idx, value=sumprod_dd)
    ws.cell(row=insert_at + 1, column=total_col_idx, value=sumprod_lltm)


def main():
    wb = openpyxl.load_workbook(XLSX)

    # SAM Mekko Table 1: NewCon at row 7, Total moves from row 14 -> 16
    insert_breakdown_rows(
        wb, 'SAM',
        newcon_row=7,
        total_row_after_insert=16,
        total_col_letter='Q',
        columns_config=sam_mekko_columns,
    )

    # TAM Mekko Table 6: NewCon at row 70, Total moves from row 77 -> 79
    insert_breakdown_rows(
        wb, 'TAM',
        newcon_row=70,
        total_row_after_insert=79,
        total_col_letter='T',
        columns_config=tam_mekko_columns,
    )

    wb.save(XLSX)
    print("Saved.")


if __name__ == '__main__':
    main()
