#!/usr/bin/env python3
"""
Apply three updates to v1.21 → save as v1.22:
  1. Consolidate duplicate Submarket values (en-dash → hyphen variants)
  2. Add Validation section flagging duplicate OPN_BA3/OPN_BA5-8 rows
  3. Add Validation section describing OMN/NWCF prorated allocation methodology
     (NOT the implementation — just the spec for a future sheet)
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE = Path('/Users/brendantoole/projects2/buildco3')
SOURCE_XLSX = BASE / '08APR2028_Newbuild_and_MRO_Spend_v1.21.xlsx'
OUTPUT_XLSX = BASE / '08APR2028_Newbuild_and_MRO_Spend_v1.22.xlsx'

print(f"Loading {SOURCE_XLSX.name}...")
wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=False)

# ─── 1. CONSOLIDATE SUBMARKET DUPLICATES ─────────────────────────────────────
print("Consolidating Submarket en-dash → hyphen variants...")
ws = wb['J Book Items Cons.']

EN_DASH_TO_HYPHEN = {
    'Surface Combatants – Large': 'Surface Combatants - Large',
    'Surface Combatants – Medium': 'Surface Combatants - Medium',
}

submarket_col = 33  # 1-indexed col AG
fixed = 0
for r in range(5, ws.max_row + 1):
    cell = ws.cell(row=r, column=submarket_col)
    if cell.value in EN_DASH_TO_HYPHEN:
        cell.value = EN_DASH_TO_HYPHEN[cell.value]
        fixed += 1
print(f"  Consolidated {fixed} cells.")


# ─── 2. ADD VALIDATION SECTIONS ──────────────────────────────────────────────
print("Adding Validation sheet sections...")
val_ws = wb['Validation']
start_row = val_ws.max_row + 2

bold = Font(bold=True)
section_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')  # orange for flags

def write_section_header(row, title, fill_color='FFD966'):
    cell = val_ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

def write_row(row, sheet, field, value, desc, notes):
    val_ws.cell(row=row, column=1, value=sheet)
    val_ws.cell(row=row, column=2, value=field)
    val_ws.cell(row=row, column=3, value=value)
    val_ws.cell(row=row, column=4, value=desc)
    val_ws.cell(row=row, column=5, value=notes)
    for col in [4, 5]:
        val_ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')


# ─── SECTION A: DATA QUALITY FLAGS ───────────────────────────────────────────
write_section_header(start_row, 'DATA QUALITY FLAGS — KNOWN DUPLICATES & INCONSISTENCIES', fill_color='F4B084')
start_row += 2

flag_rows = [
    ('Data Quality', 'Submarket value duplicates (RESOLVED in v1.22)',
     'Surface Combatants - Large / Medium',
     'The Submarket column previously had two punctuation variants for the same category: hyphen ("Surface Combatants - Large") and en-dash ("Surface Combatants – Large"). Same for Medium. As of v1.22, all en-dash variants have been replaced with hyphen variants so SAM filters and SUMIFS formulas treat them as a single bucket.',
     'If you reload data from a source that uses en-dashes, re-run the consolidation step.'),

    ('Data Quality', 'Duplicate rows across OPN_BA3 and OPN_BA5-8',
     '5 line items appear in both source books',
     'These five P-1 line items have a row in BOTH the OPN_BA3 source book and the OPN_BA5-8 source book of the J Book Items Cons. sheet. When summed across all OPN_BAx source books for TAM, they double-count. The duplication is pre-existing in v1.20 and earlier — likely an artifact of how the spreadsheet was originally compiled (BA03 captures aviation support items; BA08 captures spares; some items legitimately span both, but the duplicates here are the same line being categorized twice rather than split).',
     'Estimated double-count impact in FY26 and FY27: roughly +$1.0M to +$1.3M depending on year. To resolve, decide which source book is canonical and BLANK the other set\'s rows (do not delete — keep row positions stable for cross-sheet formula references).'),

    ('Data Quality', 'Duplicate row #1: LI 116 Advanced Arresting Gear (AAG)',
     'OPN_BA3 R2890 + OPN_BA5-8 R3208',
     'Both rows reference the same P-1 line item (Advanced Arresting Gear). Citation in OPN_BA3 says "P-40 / BA03 / BSA03 Aircraft Support Equipment / LI 4217". Citation in OPN_BA5-8 says "Exhibit P-1 / BA03 Aviation Support Equipment / Aircraft Support Equipment".',
     'Both populate FY27 = $23,551K disc. Pick one as canonical.'),

    ('Data Quality', 'Duplicate row #2: LI 117 Electromagnetic Aircraft Launch System (EMALS)',
     'OPN_BA3 R2901 + OPN_BA5-8 R3209',
     'Both rows reference the same P-1 line item (EMALS). Same dual-citation pattern as AAG.',
     'Both populate FY27 = $36,908K disc. Pick one as canonical.'),

    ('Data Quality', 'Duplicate row #3: LI 122 UMCS-Unman Carrier Aviation (UCA) Mission Control Station',
     'OPN_BA3 R2947 + OPN_BA5-8 R3210',
     'Both rows reference the same P-1 line item (UMCS-UCA). Same dual-citation pattern.',
     'Both populate FY27 = $211,216K disc. Pick one as canonical.'),

    ('Data Quality', 'Duplicate row #4: LI 176 Spares and Repair Parts',
     'OPN_BA3 R2973 + OPN_BA5-8 R3220',
     'Both rows reference P-1 BA08 line 176 "Spares and Repair Parts". OPN_BA3 has it because the source book includes BA08 spares; OPN_BA5-8 also covers BA08. Same line, two rows.',
     'Both populate FY27 = $765,711K disc. Pick one as canonical. Largest single-line double-count.'),

    ('Data Quality', 'Duplicate row #5: LI 177 VIRGINIA Class (VACL) Spares and Repair Parts',
     'OPN_BA3 R2974 + OPN_BA5-8 R3221',
     'Both rows have title "VIRGINIA Class (VACL) Spares and Repair Parts" with citation pointing to BA08 BLIN 9021. In FY27, the P-1 OPN does NOT have a separate Virginia-specific spares line — only the consolidated BA08 "Spares and Repair Parts" line 176. So in v1.21, neither of these rows received a FY27 value (no clean P-1 match). Both rows are blank in FY27 columns.',
     'In FY26, both rows have data populated. For FY27, decide whether VIRGINIA Spares should be allocated as a fixed share of the consolidated BA08 line 176 ($765,711K) or left blank. If allocated, follow the same prorated approach planned for OMN/NWCF.'),
]

for sheet, field, value, desc, notes in flag_rows:
    write_row(start_row, sheet, field, value, desc, notes)
    start_row += 2  # blank row spacing


# ─── SECTION B: OMN/NWCF PRORATED ALLOCATION METHODOLOGY ─────────────────────
start_row += 1
write_section_header(start_row, 'PHASE 2 — OMN / NWCF / USCG FY27 ALLOCATION METHODOLOGY (planned, not yet implemented)', fill_color='C6E0B4')
start_row += 2

phase2_rows = [
    ('Phase 2 Plan', 'Why a separate sheet',
     'New sheet "FY27 Allocation Workbench"',
     'The FY 2027 P-1 (procurement) gives line-item detail, so SCN/OPN/WPN can be populated directly in J Book Items Cons. (done in v1.21). The FY 2027 O-1 (operation & maintenance) and RF-1 (revolving funds), however, only publish SAG/account-level totals — they do NOT break out individual ship availabilities or NSWC/NUWC sub-allocations. The OMN, OMN_Vol2, NWCF, and USCG source-book rows in this workbook have FAR more granular detail than the FY27 source documents support. To populate FY27 for those rows, we need to APPORTION the SAG total across the line items using a defensible rule, and we want to do that in a separate workbench sheet that can be reviewed independently before merging values back into J Book Items Cons.',
     'USCG cannot be filled from this DoW exhibit set at all — it would require the DHS Coast Guard Congressional Justification, sourced separately. Phase 2 covers OMN/OMN_Vol2/NWCF only.'),

    ('Phase 2 Plan', 'Allocation rule',
     'Pro-rate by FY26 share of SAG total',
     'For each spreadsheet row to be filled, compute its FY26 share of its SAG\'s total, then apply that share to the FY27 SAG total. Formula per row: FY27_row = (FY26_row / FY26_SAG_total) * FY27_SAG_total. Use FY26 DAA Enacted Total (col R) as the FY26 baseline because it is the most current authoritative figure (Validation rule: column priority is DAA Enacted → Total → Request).',
     'This assumes the relative distribution of dollars across line items within a SAG is stable year-over-year. That is a reasonable first-pass assumption for routine maintenance accounts but breaks down for one-time/emergent items. Flag any row where the prorated FY27 differs from FY26 by more than ±25% for human review.'),

    ('Phase 2 Plan', 'Source SAG totals (FY27)',
     'From FY27 O-1 / RF-1',
     'OMN (1804N) FY27 SAG totals come from FY27_o1.txt. The relevant OMN ship-related SAGs are: 1A5A Aircraft Depot Maintenance ($2,219,583K), 1B1B Mission and Other Ship Operations ($7,424,752K), 1B2B Ship Operations Support & Training ($1,713,065K), 1B4B Ship Depot Maintenance ($14,292,873K), 1B5B Ship Depot Operations Support ($2,597,722K). NWCF (493002N) FY27 = $266,212K (only one line: Naval Surface Warfare Centers).',
     'These totals all have FY27 Disc Request only — no FY27 Mandatory Request for OMN/NWCF in the FY27 PB (mandatory is concentrated in procurement accounts).'),

    ('Phase 2 Plan', 'Workbench sheet structure',
     'One row per J Book row to be allocated',
     'Columns: (A) J Book row #, (B) Source Book, (C) LI, (D) Title, (E) SAG, (F) FY26 DAA Enacted, (G) FY26 SAG total, (H) FY26 share %, (I) FY27 SAG total, (J) FY27 prorated, (K) ±25% flag, (L) Notes / manual override. The FY27 prorated value in col J = G * F / G — i.e., F/G * I. Allow manual overrides in col L which take precedence when set.',
     'After review, copy col J back into the FY27 Request column of J Book Items Cons. for each row. Leave Mandatory Request blank for OMN/NWCF (no FY27 mandatory ask in those accounts).'),

    ('Phase 2 Plan', 'SAG → row mapping',
     'Use the Citation column',
     'Each OMN / OMN_Vol2 row in J Book Items Cons. has a Citation that includes the SAG code (e.g., "OMN > BA01 Operating Forces > 1B4B Ship Depot Maintenance > ..."). Parse the SAG code from the Citation to group rows by SAG. NWCF rows similarly cite their account.',
     'For rows with ambiguous or missing citations, manual SAG assignment is required.'),

    ('Phase 2 Plan', 'Validation after Phase 2',
     'Sum check vs FY27 SAG total',
     'After allocating, the sum of all prorated FY27 values within a SAG should equal the FY27 SAG total (modulo rounding). If it does not, the prorated approach has lost or duplicated dollars. The Phase 2 sheet should include a per-SAG summary row with: SAG total, sum of prorated rows, variance.',
     'Same TAM rules apply: only PARENT and (no tag) rows are additive. Sub rows can be filled for context but should not be summed in the SAG check.'),

    ('Phase 2 Plan', 'When NOT to prorate',
     'Manually populated rows',
     'If you discover that a specific OMN/NWCF row has a directly-stated FY27 amount in some other source (e.g., a Q&A document, an exhibit P-3a, an OSD memo), use the manual override (col L of the workbench) instead of the prorated value. The prorated approach is a placeholder, not a substitute for actual data when actual data is available.',
     '—'),

    ('Phase 2 Plan', 'USCG handling',
     'Out of scope for Phase 2',
     'USCG rows (18 PARENT/no-tag rows) cannot be populated from the FY27 P-1/O-1 because Coast Guard is part of DHS, not DoW. To fill these, source the FY27 DHS Coast Guard Congressional Justification (a separate document, typically released around the same time as the DoW PB). Then build a similar mapping (likely simpler — USCG has fewer cutters/boats than the Navy has ship maintenance availabilities). This is Phase 3.',
     'Until Phase 3, USCG FY27 columns remain blank.'),
]

for sheet, field, value, desc, notes in phase2_rows:
    write_row(start_row, sheet, field, value, desc, notes)
    start_row += 2


# ─── SAVE ─────────────────────────────────────────────────────────────────────
print(f"Saving to {OUTPUT_XLSX.name}...")
wb.save(OUTPUT_XLSX)
print("Done.")
